import numpy as np
# Visualisation
import matplotlib.pyplot as plt
import matplotlib
import numpy as np
import os
import pandas as pd
from tqdm import tqdm, trange
import parmap
import scipy
from openpyxl import load_workbook
from scipy.signal import savgol_filter
from scipy.interpolate import interp1d
from dtw import dtw
import yaml
from scipy.stats import pearsonr
from tqdm import tqdm
from scipy.signal import savgol_filter
from scipy.signal import butter, sosfilt
from scipy import stats
import pickle
import dill

#
import binarize2pcalcium.binarize2pcalcium as binca

#
def smooth_ca_time_series4(diff):
	#
	''' This returns the last value. i.e. no filter

	'''

	temp = (diff[-1]*0.4+
			diff[-2] * 0.25 +
			diff[-3] * 0.15 +
			diff[-4] * 0.10 +
			diff[-5] * 0.10)

	return temp

#
class ProcessCalcium():

    def __init__(self, root_dir, 
                 animal_id 
                 ):

        #
        self.fps = 30


        #
        self.root_dir = root_dir
        self.animal_id = animal_id
        
        #
        self.fname_animal = os.path.join(self.root_dir,
                                    self.animal_id,
                                    self.animal_id+'.dill')
        
        # load yaml file
        self.fname_yaml = os.path.join(self.root_dir,
                            self.animal_id,
                            animal_id+'.yaml')
        
        # load yaml file
        with open(self.fname_yaml) as file:
            doc = yaml.load(file, Loader=yaml.FullLoader)

        #
        self.cohort_year = doc['cohort_year']
        self.dob = doc['dob']
        self.animal_name = doc['name']
        self.sex = doc['sex']
        self.rec_type = doc['rec_type']

        #
        self.session_ids = doc['session_ids']

        #
        self.verbose = True

        # also make a results folder in the animal id folder
        self.results_dir = os.path.join(self.root_dir,
                                        self.animal_id,
                                        'results')
        if not os.path.exists(self.results_dir):
            os.makedirs(self.results_dir)

        #
        self.clrs=['blue','lightblue','red','pink']

        # load session types by looping over all session
        self.session_types = []
        for session_id in self.session_ids:
            # load the yaml file in that directory
            fname_yaml = os.path.join(self.root_dir,
                                    self.animal_id,
                                    str(session_id),
                                    str(session_id)+'.yaml')
            with open(fname_yaml) as file:
                doc = yaml.load(file, Loader=yaml.FullLoader)
                self.session_types.append(doc['session_type'])

    def save_animal(self):
        #
        import dill

        #
        with open(self.fname_animal, 'wb') as file:
            dill.dump(self, file)

    def load_animal(self):
        
        #
        print ("...loading animal from disk...")
        #
        with open(self.fname_animal, 'rb') as file:
            a = dill.load(file)

        return a

    #
    def make_correlation_dirs(self):

        # Since i have no idea how to solve the problem with the missing wheel_flag i decided to do i like that
        # You should look deeper into it
        # Checking if variable wheel_flag is defined in locals or globals
        
        #
        self.data_dir = os.path.join(self.root_dir,
                                    self.animal_id,
                                    str(self.session_ids[self.session_id]))

        # make sure the data dir is correct
        if self.shuffle_data:
            data_dir = os.path.join(self.data_dir,'correlations_shuffled')
        else:
            data_dir = os.path.join(self.data_dir,'correlations')
        self.make_dir(data_dir)


        #
        self.corr_dir = data_dir


    def make_dir(self,data_dir):

        # check if dir exists or make it
        if os.path.exists(data_dir)==False:
            os.mkdir(data_dir)

    #
    def compute_correlations(self, min_number_bursts=0):

        ############## COMPUTE CORRELATIONS ###################

        # turn off intrinsic parallization or this step goes too slow
        os.environ['OPENBLAS_NUM_THREADS'] = '1'
        os.environ['OMP_NUM_THREADS']= '1'

        # compute correlations between neurons
        rasters_DFF = self.sessions[self.session_id].F_filtered   # use fluorescence filtered traces
        rasters = self.sessions[self.session_id].F_upphase_bin
        self.min_number_bursts = min_number_bursts
        # here we shuffle data as a control
        if self.shuffle_data:
            rasters, rasters_DFF = self.shuffle_rasters(rasters, rasters_DFF)

        # # if we subselect for moving periods only using wheel data velcoity
        # wheel_flag = False
        
        # # select moving
        # text = 'all_states'
        # if self.subselect_moving_only and wheel_flag:
        #     rasters = rasters[:, self.idx_run]
        #     rasters_DFF = rasters_DFF[:, self.idx_run]

        #     # add moving flag to filenames
        #     text = 'moving'

        # elif self.subselect_quiescent_only and wheel_flag:
        #     rasters = rasters[:, self.idx_quiescent]
        #     rasters_DFF = rasters_DFF[:, self.idx_quiescent]

        #     # add moving flag to filenames
        #     text = 'quiescent'

        

        # select only good ids 
        #rasters = rasters[self.clean_cell_ids]
        #rasters_DFF = rasters_DFF[self.clean_cell_ids]

        # self.corrs = compute_correlations(rasters, self)
        self.corrs = compute_correlations_parallel(self.corr_dir,
                                                    rasters,
                                                    rasters_DFF,
                                                    self.n_cores,
                                                    # self.correlation_method,
                                                    self.binning_window,
                                                    self.subsample,
                                                    self.scale_by_DFF,
                                                    self.corr_parallel_flag,
                                                    self.zscore,
                                                    self.n_tests_zscore,
                                                    self.recompute_correlation,
                                                    self.min_number_bursts)
        

    # #
    def load_day0_mask(self):

    
        fname = os.path.join(
                        self.root_dir,
                        self.animal_id,
                        'day0',
                        'rois_pixels_and_thresholds_day0.npz')

        #
        try:
            data = np.load(fname, allow_pickle=True)
        except:
            print ("Could not Day0 masks ... trying day1 ")
            fname = os.path.join(
                                self.root_dir,
                                self.animal_id,
                                self.session_ids[1],
                                'rois_pixels_and_thresholds.npz')

            data = np.load(fname, allow_pickle=True)
        #
        contours_all_cells = data['contours_all_cells']
        self.cell_ids = data['cell_ids']

        self.contours_ROIs = contours_all_cells[self.cell_ids]

        
    #
    def compute_reward_centered_traces(self):

        
        if True:
            parmap.map(get_reward_centered_traces,
                        self.session_ids,
                        self.root_dir,
                        self.animal_id,
                        self.reward_window,
                        self.filter,
                        pm_processes=10,
                        pm_pbar=True)

    ###################
    def plot_network_graph(self):

        print (self.reward_centered_traces.shape)

        #
        temp = self.reward_centered_traces.copy().transpose(1,0,2)
        print (temp.shape)

        temp = temp.reshape(temp.shape[0],-1)
        print (temp.shape)

        # remove means from each cell
        temp = temp - np.median(temp,1)[:,np.newaxis]

        # compute the correlation matrix
        corr_matrix = np.corrcoef(temp)
        print (corr_matrix.shape)

        # set threshold for correlation matrix at 0.3
        corr_matrix[corr_matrix<0.3]=0

        ###################################################
        # use the corr_matrix as an adjacency matrix and compute the graph
        import networkx as nx
        G = nx.from_numpy_array(corr_matrix)

        # remove self-loops
        G.remove_edges_from(nx.selfloop_edges(G))

        # from pylab import rcParams
        # rcParams['figure.figsize'] = 14, 10
        pos = nx.spring_layout(G, scale=20, k=3/np.sqrt(G.order()))
        d = dict(G.degree)
        nx.draw(G, pos, 
                node_color='lightblue', 
                with_labels=False, 
                node_size = 2,
                width = 0.15,
                nodelist=d, 
                alpha=1,
                #node_size=[d[k]*300 for k in d]
                ax=self.ax
                )


        # title should show graph density and also the number of nodes
        density = 2*G.number_of_edges() / (G.number_of_nodes() *(G.number_of_nodes() - 1))

        self.ax.set_title("\n\n\n"+self.animal_id+",  "+self.session_ids[self.session_id]+'\n' 
                    +"# of cells: "+ str(corr_matrix.shape[0])+'\n'
                        +'Graph density: '+str(np.round(density,3))+'\n'
                        +'Number of nodes: '+str(G.number_of_nodes())+'\n'
                        +'Number of edges: '+str(G.number_of_edges())+'\n'
                        +'Number of isolates: '+str(nx.number_of_isolates(G)),
                        fontsize=14)


    def plot_reward_centered_traces_ROIs_all_days(self):


        #
        clrs = ['blue','lightblue','red','pink']
        clrs2 = ['black','green','magenta','orange']

        # make a viridis based discrete colormap
        import matplotlib.colors as colors
        cmap = plt.cm.viridis(np.linspace(0,1,len(self.session_ids)-1))
        
        
        #
        plt.figure()
        for cell_id in range(4):
            #
            ax=plt.subplot(2,2,cell_id+1)

            # plot title about cell_id
            ax.set_title('Cells matching ROI: '+str(cell_id), fontsize=14)

            
            # loop over all the session_ids
            ctr_session = 0
            for session_id in range(1,len(self.session_ids),1):

                # get the roi for the session      
                closest_ROI_cells = self.session_contours[session_id]
                #print ("closest_ROI_cells: ", closest_ROI_cells)

                # load reward centered traces for the session
                fname = os.path.join(
                                    self.root_dir,
                                    self.animal_id,
                                    self.session_ids[session_id],
                                    'reward_centered_traces_rewardwindow_'+str(self.reward_window)+'.npy')

                reward_centered_traces = np.load(fname, allow_pickle=True)
        
                # plot the mean and std for all the cells spacing them out by 3
                t = np.arange(reward_centered_traces.shape[2])/30-(self.reward_window/30)

                #
                cell_traces = []
                for k in range(reward_centered_traces.shape[1]):


                    # loop over all the closest_ROI_cells and check if k is matching
                    for i in range(len(closest_ROI_cells)):
                        matching_mask = closest_ROI_cells[i][0]
                        matching_cell = closest_ROI_cells[i][1]
                        
                        #print ("cell_id: ", cell_id, 'k: ', k, ' matching_cell: ', matching_cell, ' matching_mask: ', matching_mask)
                        if k == matching_cell and matching_mask == cell_id:

                            # grab the data                            #
                            temp = reward_centered_traces[:,k]

                            # subtract the median from each trace
                            # compute f0 as the median of the stacked trace
                            #f0 = np.median(temp.flatten())
                            #print ("f0: ", f0)
                            #temp = (temp-f0)/f0*100

                            #
                            #import scipy
                            #temp_mean = scipy.stats.mode(temp,0)
                            
                            #print ("temp_mean: ", temp_mean.shape)

                            # compute the std of temp in the first axis
                            temp_mean = np.mean(temp,0)
                            
                            cell_traces.append(temp_mean)
                            
                            


                # 
                cell_traces = np.array(cell_traces)

                # pick the cell trace with the highest ptp peak to peak
                #print ("cell_traces: ", cell_traces.shape)

                if cell_traces.shape[0]==0:
                    continue
               
                elif cell_traces.shape[0]==1:
                    cell_trace = cell_traces[0]

                elif cell_traces.shape[0]>1:
                    ptps = np.ptp(cell_traces,1)

                    idx = np.argmax(ptps)
                    cell_trace = cell_traces[idx]


                # plot legend only if cell_id ==0:
                if cell_id==0:
                    plt.plot(t,cell_trace,
                            linewidth=3,
                            c=cmap[ctr_session],
                            label=self.session_ids[session_id]+ ", # rew: "+str(reward_centered_traces[:,k].shape[0]))
                else:
                    plt.plot(t,cell_trace,
                            linewidth=3,
                            c=cmap[ctr_session],
                            )


                ctr_session+=1


            # plot horizontal line at 0
            plt.plot(t, temp_mean*0,'--',
                    c='grey')    
            
            # plot vertical line at 0
            #plt.plot([0,0],[-50,100],'--',
            #        c='grey',
            #        linewidth=3)

            # 
            #plt.plot([0,0],[-.50,100],'--',
            #         c='grey',
            #         linewidth=3)
            #
            plt.ylabel("DFF")
            #
            plt.legend()

        # label y axis using ylabels 
        #plt.yticks(ylabels, np.arange(temp.shape[0]))
        
        #
        plt.suptitle(self.animal_id)
        plt.show()

    #
    def find_matching_cells_to_master_mask(self):

        fname_out = os.path.join(self.root_dir, 
                                self.animal_id, 
                                'results',
                                'best_matches.npy')
        
        #
        if os.path.exists(  fname_out)==False or self.recompute_match_master_mask:
            #
            sessions_F_filtered = []
            for session in self.sessions:
                sessions_F_filtered.append(session.F_filtered)

            #
            cell_names = ['roi_pos1','roi_pos2','roi_neg1','roi_neg2']
            res = parmap.map(find_best_match_bmi_vs_mastermask,
                             cell_names,
                             self.root_dir,
                             self.animal_id,
                             sessions_F_filtered,
                             self.session_ids,
                             pm_processes=4,
                             pm_pbar=True)

            #
            print ("res: ", res)

            # make 4 panels with histogram of best match
            plt.figure(figsize=(10,10))
            best_matches = []
            for k in range(4):
                plt.subplot(2,2,k+1)

                #
                y = np.histogram(res[k], bins=np.arange(0,np.max(res[k])+3,1))

                # find locations where y[1]>0
                ctr =0
                tags = []
                best_match = [0,0]
                for q in range(y[1].shape[0]-1):
                    yy = y[0][q]
                    if yy>0:
                        plt.bar(ctr,yy,0.9, label=str(q))
                        tags.append(q)

                        #
                        if yy>best_match[0]:
                            best_match[0]=yy
                            best_match[1]=q


                        ctr+=1     

                # relabel x axis
                plt.xticks(np.arange(ctr), y[1][tags])

                #
                plt.title(cell_names[k])
                plt.legend()

                #
                best_matches.append(best_match)

            #    
            plt.suptitle(self.animal_id+ "\nBest match to master mask for ensemble cells over all sessions")
            plt.savefig(os.path.join(self.root_dir, 
                                    self.animal_id, 
                                    'results',
                                    'best_match_ensemble_cells.png'),dpi=300)

            # 
            plt.show()
            
            # also save the res array in the same folder
            np.save(os.path.join(self.root_dir, 
                                    self.animal_id, 
                                    'results',
                                    'best_match_ensemble_cells.npy'),res)
            

            
            # also save best_matches array
            np.save(fname_out,best_matches)

            #
            self.best_matches = best_matches    

        #
        else:
            self.best_matches = np.load(fname_out, allow_pickle=True)
        
        #
        self.best_matches = np.array(self.best_matches)


    def load_best_matches(self):
        fname = os.path.join(self.root_dir, 
                                self.animal_id, 
                                'results',
                                'best_matches.npy')
        
        #           
        self.best_matches = np.load(fname, allow_pickle=True)[:,1]

        print ("self.best_matches: ", self.best_matches)
    #
    def plot_reward_centered_traces_ROIs_only(self):

        #
        clrs = ['blue','lightblue','red','pink']
        clrs2 = ['black','green','magenta','orange','yellow','cyan','purple','brown']


        # plot the mean and std for all the cells spacing them out by 3
        t = np.arange(self.reward_centered_traces.shape[2])/30-(self.reward_window/30)

        #
        closest_ROI_cells = self.session_contours[self.session_id]
        print ("closest_ROI_cells: ", closest_ROI_cells)
    
        #
        linetypes = ['solid',':','--','dashdot']

        plt.figure()
        scale = 100
        for cell_id in range(4):
            #
            ax=plt.subplot(2,2,cell_id+1)
            
            # plot title about cell_id
            ax.set_title('Cells matching ROI: '+str(cell_id), fontsize=14)


            #
            ctr = 0
            for k in range(self.reward_centered_traces.shape[1]):

                #
                temp = self.reward_centered_traces[:,k]

                # subtract the median from each trace
                #temp = temp - np.median(temp,1)[:,np.newaxis]
                #print ("temp: ", temp.shape)
                temp_mean = np.mean(temp,0)

                # compute the std of temp in the first axis
                #temp_mean = np.mean(temp,0)

                # loop over all the closest_ROI_cells and check if k is matching
                for i in range(len(closest_ROI_cells)):
                    matching_mask = closest_ROI_cells[i][0]
                    matching_cell = closest_ROI_cells[i][1]
                    
                    #print ("cell_id: ", cell_id, 'k: ', k, ' matching_cell: ', matching_cell, ' matching_mask: ', matching_mask)
                    if k == matching_cell and matching_mask == cell_id:
                    
                        # if k is the first two values in cell_ids plot using blue
                        plt.plot(t,temp_mean,
                                linewidth=3,
                                #linestyle=linetypes[ctr],
                               # alpha=0.5,
                                c=clrs2[ctr],
                                label=str(matching_cell))

                        ctr+=1
                
            # plot horizontal line at 0
            plt.plot(t, temp_mean*0,'--',
                    c='grey')    
            
            # plot vertical line at 0
            #plt.plot([0,0],[-100,500],'--',
            #        c='grey',
            #        linewidth=3)

            # 
            #plt.plot([0,0],[-100,500],'--',
            #         c='grey',
            #         linewidth=3)

            #
            plt.legend()

        # label y axis using ylabels 
        #plt.yticks(ylabels, np.arange(temp.shape[0]))
        plt.xlabel('Time (s)',fontsize=20)

        #
        plt.suptitle(self.session_ids[self.session_id])
        plt.show()

    #
    def plot_rewards(self, ax, 
                     y_scale=1,
                     alpha=1.0):

        r_times = []
        for k in range(len(self.reward_times)):
            if k>0:
                temp = self.reward_times[k]+k*90000
                r_times.extend(temp)
        r_times = np.array(r_times)

        #
        ax.vlines(r_times/self.fps,-0.3,y_scale,
                        color='green',
                        #linestyles='dashed',
                        linewidth=1,
                        alpha=alpha,
                        label="Reward times")

    #
    def make_multi_session_traces(self):
        
        #
        fps = 30

        # visualize the ensembel cells over all time
        roi_match_ids = self.best_matches[:,1]

        #

        # loop over the sessions and plot the ensemble cells
        if self.plotting:
            plt.figure(figsize=(20,10))
            plt.suptitle(self.animal_id)
            ax=plt.subplot(111)
        
        #
        t_start = 0
        window = 120*30

        #
        traces = [[],[],[],[]]
        traces_bin = [[],[],[],[]]
        traces_realtime = [[],[],[],[]]
        for session in range(len(self.sessions)):
            ca = self.sessions[session].F_filtered
            ca_bin = self.sessions[session].F_upphase_bin

            # plot all 4 rois from the master mask
            t = np.arange(t_start, t_start+ca.shape[1],1)
            ctr=0
            for roi_id in roi_match_ids:
                temp = ca[roi_id].copy()
                temp_bin = ca_bin[roi_id].copy()

                # cmpute mode of temp using scipy or stats
                traces[ctr].extend(temp)            

                # the same but for upphase
                traces_bin[ctr].extend(temp_bin)
                
                #
                ctr+=1

            # plot rois from the .npz file
            if self.show_realtime_bmi_ensembles:
                
                if session==0:
                    temp = np.zeros(ca.shape[1])
                    traces_realtime[0].extend(temp)
                    traces_realtime[1].extend(temp)
                    traces_realtime[2].extend(temp)
                    traces_realtime[3].extend(temp)
                else:
                    # # load results.npz file
                    rois1, rois2, _ = load_results_npz_standalone(self.root_dir,
                                                                    self.animal_id,
                                                                    session,
                                                                    self.session_ids)
                    #
                    rois1[:,:300]=0
                    rois2[:,:300]=0
                    
                    #
                    traces_realtime[0].extend(rois1[0])
                    traces_realtime[1].extend(rois1[1])
                    traces_realtime[2].extend(rois2[0])
                    traces_realtime[3].extend(rois2[1])
                                            

            #
            t_start+=ca.shape[1]

        #
        traces = np.array(traces)
        traces_bin = np.array(traces_bin)
        traces_realtime = np.array(traces_realtime)

        #
        if self.plotting:
            t=np.arange(traces[0].shape[0])/30. 

            # loop over all sessions
            for k in range(traces.shape[0]):
                
                # plot merged-mask traces
                temp = traces[k]
                ax.plot(t,
                        temp+k*self.y_scale,
                        c=self.clrs[k],
                        label="[ca] merged mask" if k==0 else None
                        )
                
                # plot BMI realtime ensemble signals
                if self.show_realtime_bmi_ensembles:
                    temp = traces_realtime[k]
                    ax.plot(t,
                            temp+k*self.y_scale,
                            '--',
                            c=self.clrs[k],
                            label="Realtime BMI signal" if k==0 else None
                            )
                
                # plot binarization traces
                if self.show_upphases:
                    temp = traces_bin[k]
                    ax.plot(t,
                            temp+k*self.y_scale,
                            #'-.',
                            linewidth=1,
                            alpha=0.4,
                            c='black',
                            label="Upphases" if k==0 else None
                            )
                    
            #######################################################
            #######################################################
            #######################################################

            # plot vertical red lines at reward times
            if self.show_rewards:
                self.plot_rewards(ax, 4*self.y_scale,
                                  alpha=0.2)
    
            #######################################################
            #######################################################
            #######################################################
            # make light grey background shading every 90000 frames
            ctr=0
            for k in range(0, t_start, 2*90000):
                try:
                    ax.axvspan(t[k], t[k+90000], color='grey', alpha=0.05)
                except:
                    pass

            # plot horizontal lines at zero
            ctr = 0
            for k in range(4):
                temp = traces[k].copy()
                temp = np.round(temp,3)
                mode1 = scipy.stats.mode(temp).mode

                ax.plot([t[0], t[-1]], 
                            [ctr*self.y_scale+mode1, ctr*self.y_scale+mode1], c='black')
                ctr+=1
            
            # add xtick custom labels
            xlabels = ['day0']
            for k in range(1,len(self.session_ids)):
                xlabels.append(self.session_types[k])
            
            #
            plt.xlabel("Time (sec)")
            plt.xlim(t[0], t[-1])

            # relabel yticks with neuron ids
            neuron_ids = ['roi_pos1','roi_pos2','roi_neg1','roi_neg2']
            plt.yticks(np.arange(4)*self.y_scale+0.75, neuron_ids,
                    rotation=90)

            plt.legend()


            #################################################################
            #################################################################
            #################################################################
            # add the xlabels to the xticks
            ax2 = ax.twiny()
            ax2.set_xticks(np.arange(len(xlabels))*90000/fps+90000/fps/2, 
                    xlabels, rotation=45)
            
            ax2.set_xlim(t[0], t[-1])

            #

            #
            plt.show()

        #
        self.traces = traces
        self.traces_bin = traces_bin 
        self.traces_realtime = traces_realtime 

    #
    def make_multi_session_rasters_allcells(self):
        
        #s
        self.fps=30

        #
        rows, cols = self.sessions[0].F_upphase_bin.shape
        rasters = np.zeros((rows,cols*len(self.sessions)))
        self.rasters = np.zeros((rows,cols*len(self.sessions)))
        print ("rasters: ", rasters.shape)

        #
        t_start = 0
        for session in trange(len(self.sessions), desc='loading sessions'):
            ca_bin = self.sessions[session].F_upphase_bin
            print ("ca_bin: ", ca_bin.shape)
            # find where ca_bin is 1
            for p in range(ca_bin.shape[0]):
                idx = np.where(ca_bin[p]==1)[0]
                self.rasters[p,t_start+idx]=1

                # for visualization only 
                for q in range(5):
                    rasters[p,t_start+q+idx]=1

            #
            t_start+=cols

        # 
        print ("...ploting...")
        plt.figure()
        ax=plt.subplot(111)
        plt.imshow(rasters, 
                   aspect='auto',
                   interpolation='none',
                   # plot x-extent from 0 to rasters.shape[1]//30
                   extent=[0,rasters.shape[1]//30,rasters.shape[0],0],
                   cmap='gray_r',)
   
        #
        if self.show_rewards:
            self.plot_rewards(ax, 
                              rasters.shape[0],
                              alpha=0.4)

        ################################################################
        # plot horizontal lines where the self.best_matches are
        matches = self.best_matches[:,1]
        for k in range(matches.shape[0]):
            ax.plot([0, rasters.shape[1]/self.fps], 
                    [matches[k],matches[k]],
                    color=self.clrs[k],
                    linewidth=3,
                    alpha=0.4,
                    )
                
        ############################################################
        # 
        ax.set_ylim(0,rasters.shape[0])
        # add xtick custom labels
        xlabels = ['day0']
        for k in range(1,len(self.session_ids)):
            xlabels.append(self.session_types[k])
            
        ax.set_xlabel("Time (secs)")
        ax.set_ylabel("Cell #")


        #################################################################
        #################################################################
        #################################################################
        # make light grey background shading every 90000 frames
        ctr=0
        t = np.arange(rasters.shape[1])
        print ("t_sart: ", t_start)
        t_start-=cols
        for k in range(0, t_start, 2*90000):
            ax.axvspan(t[k]/self.fps, t[k+90000]/self.fps, color='grey', alpha=0.2)

        # add the xlabels to the xticks
        ax2 = ax.twiny()
        xx = np.arange(len(xlabels))*90000/self.fps+90000/self.fps/2
        print (xx)
        ax2.set_xticks(xx, 
                        xlabels, rotation=45)
        
        ax2.set_xlim(0, xx[-1]+90000/self.fps/2)

        #
        plt.show()
        
    #
    def plot_reward_centered_traces(self):
        
        #
        clrs = ['blue','lightblue','red','pink']

        # load reward centered traces for the session
        fname = os.path.join(
                            self.root_dir,
                            self.animal_id,
                            self.session_ids[self.session_id],
                            'reward_centered_traces_rewardwindow_'+str(self.reward_window)+'.npy')
        #
        fname_shuffled = os.path.join(
                            self.root_dir,
                            self.animal_id,
                            self.session_ids[self.session_id],
                            'reward_centered_traces_rewardwindow_'+str(self.reward_window)+'_shuffled.npy')

        #
        self.reward_centered_traces = np.load(fname, allow_pickle=True)
        self.reward_centered_traces_shuffled = np.load(fname_shuffled, allow_pickle=True)

        # compute the mean of temp in the first axis
        temp = np.mean(self.reward_centered_traces,0)
        temp_shuffled = np.mean(self.reward_centered_traces_shuffled,0)

        # compute the std of temp in the first axis
        #temp_std = np.std(self.reward_centered_traces,0)

        # plot the mean and std for all the cells spacing them out by 3
        t = np.arange(temp.shape[1])/30-(self.reward_window/30)

        #
        closest_ROI_cells = self.session_contours[self.session_id]
        print ("closest_ROI_cells: ", closest_ROI_cells)
    
        #
        # looping over all suite2p cells
        traces = []
        traces_shuffled = []
        clrs2 = []
        print ("loopin gover all suite2p cells: ", temp.shape[0])
        for k in range(temp.shape[0]):
            
            #
            temp2 = temp[k]

            #
            temp2 = temp2 - np.median(temp2)

            # loop over all the closest_ROI_cells and check if k is matching
            traces.append(temp2)

            #
            temp2_shuffled = temp_shuffled[k]
            temp2_shuffled = temp2_shuffled - np.median(temp2_shuffled)
            traces_shuffled.append(temp2_shuffled)
            
            #
            found_match = False
            for i in range(len(closest_ROI_cells)):
                matching_mask = closest_ROI_cells[i][0]
                matching_cell = closest_ROI_cells[i][1]

                if k == matching_cell:
                    print ('k: ', k, ' matching_cell: ', matching_cell, ' matching_mask: ', matching_mask)
                    clrs2.append(clrs[matching_mask])
                    found_match = True
                    break
 
            if found_match == False:
                clrs2.append('black')

        #
        traces = np.vstack(traces)
        traces_shuffled = np.vstack(traces_shuffled)
        clrs2 = np.vstack(clrs2)
        cell_ids = np.arange(traces.shape[0])

        #
        if self.sort_by_peak:
            peak_idxs = np.argmax(traces,1)
            
            # sort by location of ptp
            sorted_idx = np.argsort(peak_idxs)[::-1]

            # sort traces
            traces = traces[sorted_idx]
            clrs2 = clrs2[sorted_idx]
            cell_ids = cell_ids[sorted_idx]

            # same for traces_shuffled
            peak_idxs = np.argmax(traces_shuffled,1)
            sorted_idx = np.argsort(peak_idxs)
            traces_shuffled = traces_shuffled[sorted_idx]
            

        ###################################
        plt.figure()
        ax=plt.subplot(131)
        plt.ylabel("Neuron ID")
        # plotthe traces
        ctr = 0
        labels = []
        img = []
        img_shuffled = []
        for k in range(traces.shape[0]):
            
            # grab the shuffled traces first 
            temp4 = traces_shuffled[k]
            max = np.max(np.abs(temp4))
            if max >= self.min_ca_peak:
                img_shuffled.append(temp4)

            #
            temp3 = traces[k]
            max = np.max(np.abs(temp3))

            #
            if max < self.min_ca_peak:
                continue

            #
            img.append(temp3)

            # 
            plt.plot(t,temp3+ctr*self.scale,
                    linewidth=3,
                    c=clrs2[k][0])
            
            # plot horizontal line at temp3 = 0
            plt.plot(t, temp3*0+ctr*self.scale,'--',
                    c='grey',alpha=0.5)
            

            # print colors
            if clrs2[k][0]!='black':
                print ('k: ', k, ' clrs2[k][0]: ', clrs2[k][0])
            
            # plot the peak using peak_idx
            peak_idx = np.argmax(temp3)
            plt.plot(t[peak_idx],temp3[peak_idx]+ctr*self.scale,'o',
                    c='red')   

            #
            labels.append(cell_ids[k])

            #            
            ctr+=1
        
        # plot a vertical line at 0
        plt.plot([0,0],[0,ctr*self.scale],'--',c='grey',linewidth=3, alpha=0.2)


        #
        img = np.vstack(img)[::-1]

        #
        img_shuffled = np.vstack(img_shuffled)


        # 
        #plt.plot([0,0],
        #         [0,(ctr+1)*self.scale],
        #         'r--',linewidth=3)
        
        plt.title("# of cells: "+str(img.shape[0]))

        # label y axis using ylabels 
        plt.yticks(np.arange(ctr)*self.scale, labels)
        
        plt.xlabel('Time (s)',fontsize=20)
        plt.suptitle(self.animal_id+' '+self.session_ids[self.session_id]+"\n # rewards: "
                     +str(self.reward_centered_traces.shape[0]),fontsize=20)
        # limit plot to size of image
        plt.xlim(t[0],t[-1])
        plt.ylim(0,ctr*self.scale)

        ##########################################
        ##########################################
        ##########################################
        ax=plt.subplot(132)
        plt.imshow(img,
                       aspect='auto',
                       cmap='jet',
                       interpolation='none',
                       extent=[t[0],t[-1],0,img.shape[0]],
                       vmin=-self.vmax,
                       vmax=self.vmax
                       )
        # plot a vertical line at 0
        plt.plot([0,0],[0,img.shape[0]],'--',c='grey',linewidth=3, alpha=0.2)

        #plt.colorbar()
        plt.xlabel('Time (s)',fontsize=20)

        ##########################################
        ##########################################
        ##########################################
        ax=plt.subplot(133)
        plt.title('Shuffled')
        # plotthe traces shuffled
        if img_shuffled.shape[0]<=img.shape[0]:
            img_c = img.copy()

            img_c[:] = np.nan
            img_c[-img_shuffled.shape[0]:,] = img_shuffled
        else:
            img_c = img_shuffled.copy()
        #
        print (img_c.shape)
        plt.imshow(img_c,
                       aspect='auto',
                       cmap='jet',
                       interpolation='none',
                       extent=[t[0],t[-1],0,img_c.shape[0]],
                       vmin=-self.vmax,
                       vmax=self.vmax
                       )
        # plot a vertical line at 0
        plt.plot([0,0],[0,img_c.shape[0]],'--',c='grey',linewidth=3, alpha=0.2)

        # plot colorbar with label "DFF"
        plt.colorbar(label="DFF")


        plt.xlabel('Time (s)',fontsize=20)


        plt.show()

    #
    def plot_tone_triggered_ca(self):

        tone_ids = np.arange(2,self.unique_tones.shape[0])
        print ("tone_ids: ", tone_ids)

        #
        window = 10*30
        min_space_starts = window
        vmax = self.vmax

        #
        plt.figure()

        for tone_id in tone_ids:

            #
            idx = np.where(np.int32(self.tones)==self.unique_tones[tone_id])[0]
            n_tones = idx.shape[0]
            print ("tone: ", self.unique_tones[tone_id], " # of tones: ", n_tones)

            # find the starts in each sequence of idx values
            starts = []
            for k in range(idx.shape[0]-1):
                if idx[k+1]-idx[k]>1:
                    starts.append(idx[k+1])
            starts = np.array(starts) #+3000

            # # plot the calcium events for the first tone
            t = np.arange(-window,window)/30

            # remove starts that are too close to each other
            ss = []
            ss.append(starts[0])
            for k in range(1,starts.shape[0]):
                # skip k if too close to previous
                if (starts[k]-starts[k-1])>min_space_starts:
                    ss.append(starts[k])

            starts = np.array(ss)

            ######################################
            # make image stack
            img = []
            img_shuffled = []
            for cell_id in range(self.sessions[self.session_id].F.shape[0]):
                traces = []
                traces_shuffled = []
                for s in starts:
                    #temp = c.sessions[session_id].F_upphase_bin[cell_id,s-window:s+window]
                    temp = self.sessions[self.session_id].F_filtered[cell_id,s-window:s+window]
                    if temp.shape[0]==2*window:
                        traces.append(temp)
                    
                    # same but shuffle the tone times
                    s_shuffled = np.random.randint(300,self.sessions[self.session_id].F.shape[1]-2*window)
                    temp = self.sessions[self.session_id].F_filtered[cell_id,s_shuffled-window:s_shuffled+window]
                    if temp.shape[0]==2*window:
                        traces_shuffled.append(temp)

                # plot the average of traces
                img.append(np.mean(traces,axis=0))
                img_shuffled.append(np.mean(traces_shuffled,axis=0))

            # Reorder by peak
            img = np.array(img)

            # order img by ptp
            if False:
                ptp = np.max(img,axis=1)
                idx = np.argsort(ptp)[::-1]
                img = img[idx,:]
            #
            good_cell_ids = np.where(np.max(np.abs(img),axis=1)>self.min_ca_val)[0]
            img = img[good_cell_ids]

            # also order them by location of their max/min
            #max_val = np.argmax(img,axis=1)
            max_val = np.max(img,axis=1)
            idx = np.argsort(max_val)[::-1]
            img = img[idx,:]
            good_cell_ids = good_cell_ids[idx]


            # same for img shuffled
            img_shuffled = np.array(img_shuffled)
            if False:
                ptp = np.max(img_shuffled,axis=1)
                idx = np.argsort(ptp)[::-1]
                img_shuffled = img_shuffled[idx,:]
            
            #
            idx = np.where(np.max(np.abs(img_shuffled),axis=1)>self.min_ca_val)[0]
            img_shuffled = img_shuffled[idx]

            max_val = np.max(img_shuffled,axis=1)
            idx = np.argsort(max_val)[::-1]
            img_shuffled = img_shuffled[idx,:]


            ######################################
            ######################################
            ######################################
            top_n = 1000

            #
            ax=plt.subplot(2,15,tone_id-1)
            img = img[:top_n]

            # arrange img by location of max value argument
            if False:
                idx = np.argmax(img,axis=1)
                idx = np.argsort(idx)
                img = img[idx,:]

            plt.imshow(img,
                    aspect='auto',
                    extent=[t[0],t[-1],0,img.shape[0]],
                    interpolation='none',cmap=self.cmap,
                    vmin=self.vmin,
                    vmax=self.vmax
                    )
            
            # plot yticks as good_cell_ids
            plt.yticks(np.arange(img.shape[0])+0.5,good_cell_ids)
            

            if tone_id == 2:
                plt.ylabel("Cell #")
            #else:
            #    plt.yticks([])

            plt.title(str(self.unique_tones[tone_id])+"hz" + ", #: "+str(len(starts)))

            # draw verical line at tone onset
            plt.plot([0,0],[0,img.shape[0]],'k--')
            plt.xlabel("Time (s)")

            #####################################
            # sem for shuffled
            img_shuffled = img_shuffled[:top_n]
            ax=plt.subplot(2,15,tone_id-1+15)
            if tone_id == 2:
                plt.ylabel("Cell # (SHUFFLED)")
            else:
                plt.yticks([])

            #
            plt.imshow(img_shuffled,
                       aspect='auto'
                       ,extent=[t[0],t[-1],0,img_shuffled.shape[0]],
                        interpolation='none',cmap=self.cmap,
                        vmin=self.vmin,
                        vmax=self.vmax)
            


            # draw verical line at tone onset
            plt.plot([0,0],[0,img_shuffled.shape[0]],'k--')


            plt.xlabel("Time (s)")

        plt.colorbar()
        #plt.plot(t,np.mean(traces,axis=0))
        #plt.plot(c.sessions[session_id].F[:,idx])
        #plt.title("Tone: "+str(tone))

        # plot suptitle with animal id, session id and number of rewards
        plt.suptitle("Animal: "+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(self.n_rewards))

        plt.show()

    #
    def get_tone_triggered_ca_single_cell(self):

        tone_ids = np.arange(2,self.unique_tones.shape[0])
        print ("tone_ids: ", tone_ids)

        #
        window = self.window
        min_space_starts = self.window
        vmax = self.vmax

        #
        plt.figure()
        # make image stack
        img = []
        img_shuffled = []
        traces = []
        traces_shuffled = []
        cell_id = self.cell_id
        for tone_id in tone_ids:

            #
            idx = np.where(np.int32(self.tones)==self.unique_tones[tone_id])[0]
            n_tones = idx.shape[0]
            #print ("tone: ", self.unique_tones[tone_id], " # of tones: ", n_tones)

            # find the starts in each sequence of idx values
            starts = []
            for k in range(idx.shape[0]-1):
                if idx[k+1]-idx[k]>1:
                    starts.append(idx[k+1])
            starts = np.array(starts) #+3000

            # # plot the calcium events for the first tone
            t = np.arange(-window,window)/30

            # remove starts that are too close to each other
            ss = []
            ss.append(starts[0])
            for k in range(1,starts.shape[0]):
                # skip k if too close to previous
                if (starts[k]-starts[k-1])>min_space_starts:
                    ss.append(starts[k])

            starts = np.array(ss)

            ######################################

            for s in starts:
                #temp = c.sessions[session_id].F_upphase_bin[cell_id,s-window:s+window]
                temp = self.sessions[self.session_id].F_filtered[cell_id,s-window:s+window]
                if temp.shape[0]==2*window:
                    traces.append(temp)
                
                # same but shuffle the tone times
                s_shuffled = np.random.randint(300,self.sessions[self.session_id].F.shape[1]-2*window)
                temp = self.sessions[self.session_id].F_filtered[cell_id,s_shuffled-window:s_shuffled+window]
                if temp.shape[0]==2*window:
                    traces_shuffled.append(temp)

            # plot the average of traces
            img.append(np.mean(traces,axis=0))
            img_shuffled.append(np.mean(traces_shuffled,axis=0))

        # Reorder by peak
        img = np.array(img)

        # # order img by ptp
        # if False:
        #     ptp = np.max(img,axis=1)
        #     idx = np.argsort(ptp)[::-1]
        #     img = img[idx,:]
        # #
        # good_cell_ids = np.where(np.max(np.abs(img),axis=1)>self.min_ca_val)[0]
        # img = img[good_cell_ids]

        # # also order them by location of their max/min
        # #max_val = np.argmax(img,axis=1)
        # max_val = np.max(img,axis=1)
        # idx = np.argsort(max_val)[::-1]
        # img = img[idx,:]
        # good_cell_ids = good_cell_ids[idx]


        # same for img shuffled
        img_shuffled = np.array(img_shuffled)
        if False:
            ptp = np.max(img_shuffled,axis=1)
            idx = np.argsort(ptp)[::-1]
            img_shuffled = img_shuffled[idx,:]
        
        # #
        # idx = np.where(np.max(np.abs(img_shuffled),axis=1)>self.min_ca_val)[0]
        # img_shuffled = img_shuffled[idx]

        # max_val = np.max(img_shuffled,axis=1)
        # idx = np.argsort(max_val)[::-1]
        # img_shuffled = img_shuffled[idx,:]


        ######################################
        ######################################
        ######################################
        top_n = 1000

        #
        ax=plt.subplot(1,2,1)
        img = img[:top_n]

        # arrange img by location of max value argument
        if False:
            idx = np.argmax(img,axis=1)
            idx = np.argsort(idx)
            img = img[idx,:]

        plt.imshow(img,
                aspect='auto',
                extent=[t[0],t[-1],0,img.shape[0]],
                interpolation='none',cmap=self.cmap,
                vmin=self.vmin,
                vmax=self.vmax
                )
        
        # plot yticks as unique tones
        plt.yticks(np.arange(img.shape[0])+0.5,self.unique_tones[tone_ids])
        #plt.yticks(np.arange(img.shape[0])+0.5,good_cell_ids)
        

        # if tone_id == 2:
        #     plt.ylabel("Cell #")
        # #else:
        # #    plt.yticks([])

        # draw verical line at tone onset
        plt.plot([0,0],[0,img.shape[0]],'k--')
        plt.xlabel("Time (s)")

        #####################################
        # sem for shuffled
        img_shuffled = img_shuffled[:top_n]
        ax=plt.subplot(1,2,2)
        if tone_id == 2:
            plt.ylabel("Cell # (SHUFFLED)")
        else:
            plt.yticks([])

        #
        plt.imshow(img_shuffled,
                    aspect='auto'
                    ,extent=[t[0],t[-1],0,img_shuffled.shape[0]],
                    interpolation='none',cmap=self.cmap,
                    vmin=self.vmin,
                    vmax=self.vmax)
        
        # draw verical line at tone onset
        plt.plot([0,0],[0,img_shuffled.shape[0]],'k--')

        #
        plt.xlabel("Time (s)")

        #
        plt.colorbar()

        # plot suptitle with animal id, session id and number of rewards
        plt.suptitle("Cell_id" + str(self.cell_id )+ "\nAnimal: "+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(self.n_rewards)
                    + "\n"+
                    str(self.unique_tones[tone_id])+"hz" + ", #: "+str(len(starts)))

        plt.show()

        plt.savefig(os.path.join(self.root_dir,
                                 self.animal_id,
                                 self.session_ids[self.session_id],
                                 'cells',
                                 str(cell_id)+'.png'))
        plt.close()

    #
    def get_mean_warped_trace(self):    

        #
        if self.burst_sign == 1:
            self.compute_mean_warped_trace_positive()
        else:
            self.compute_mean_warped_trace_negative()

    #
    def compute_mean_warped_trace_negative(self):
        
        #
        fps = 30

        # plot the ensemble state for each of the starts_ends separating into 2 plots by sign
        if self.plotting:
            plt.figure(figsize=(10,5))

            ax1 = plt.subplot(1,2,1)
            ax1.set_title('negative')
            plt.xlabel('time (s)')
            # plot the thresh value
            ax1.axhline(y=np.mean(self.thresh), color='grey', linestyle='--')

            #
            ax2 = plt.subplot(1,2,2)
            ax2.set_title('warped')
            plt.xlabel('time (s)')

            # make a colormap of fixed colors in viridis for 10 values
            # count # of positive and negative vals in signs
            neg = np.sum(self.signs == -1)            
            cmap_r = plt.cm.viridis(np.linspace(0, 1, neg))

        #
        ctr_neg = 0
        self.snippets = []
        self.snippets_warped = []
        print ("TODO: remove the worst negative bursts from the list...")
        for k in range(self.starts_ends.shape[0]):
            temp = self.e_state[self.starts_ends[k,0]:self.starts_ends[k,1]]

            # if time warp
            #if self.time_warp == True:           # this just stretches the snippet to be the same length - 30 timesteps
            #    temp = resample_snippet(temp)

            if self.signs[k] == -1:
        
                #
                self.snippets.append(temp)

                # 
                temp2 = resample_snippet(temp.copy())

                #                         
                self.snippets_warped.append(temp2)

            #
                if self.plotting:
                    ax1.plot(np.arange(temp.shape[0])/fps, temp,
                        c=cmap_r[ctr_neg])

                    #
                    ax2.plot(np.arange(temp2.shape[0])/fps, 
                             temp2, c=cmap_r[ctr_neg])
                ctr_neg += 1

        #
        # plot average of temps
        self.snippets_warped = np.array(self.snippets_warped)
        mean = np.mean(self.snippets_warped,axis=0)
        self.mean = mean

        #
        self.snippets = np.array(self.snippets, dtype=object)
        if self.plotting:
            ax2.plot(np.arange(mean.shape[0])/30, mean ,
                    c='red', linewidth=3)

        #
        self.cmap_r = cmap_r

        # plot the thresh value
        if self.plotting:   
            plt.show()

    #
    def compute_mean_warped_trace_positive(self):

        # plot the ensemble state for each of the starts_ends separating into 2 plots by sign
        if self.plotting:
            plt.figure(figsize=(20,5))

            ax1 = plt.subplot(1,4,1)
            ax1.set_title('positive')
            plt.xlabel('time (s)')
            # plot the thresh value
            ax1.axhline(y=np.mean(self.thresh), color='grey', linestyle='--')

            #
            ax2 = plt.subplot(1,4,2)
            ax2.set_title('negative')
            plt.xlabel('time (s)')

            #
            ax3 = plt.subplot(1,4,3)
            ax3.set_title('good snippets (> threshold; start from 0.5 max)')
            ax3.axhline(y=np.mean(self.thresh), color='grey', linestyle='--')
            plt.xlabel('Time (warped)')

            #
            ax4 = plt.subplot(1,4,4)
            ax4.set_title('snippets only to threshold')
            ax4.axhline(y=np.mean(self.thresh), color='grey', linestyle='--')
            plt.xlabel('Time (warped)')
            ax4.set_ylim(top=1)

            #
            cmap_r = plt.cm.viridis(np.linspace(0, 1, len(self.ids_snippets_rewarded)))


        # count # of positive and negative vals in signs
        pos = np.sum(self.signs == 1)
        neg = np.sum(self.signs == -1)

        # make a colormap of fixed colors in viridis for 10 values
        cmap_pos = plt.cm.viridis(np.linspace(0, 1, pos))
        cmap_neg = plt.cm.viridis(np.linspace(0, 1, neg))

        #
        ctr_pos = 0
        ctr_neg = 0
        # this just counts for purposes of making a nice list below
        for k in range(self.starts_ends.shape[0]):
            temp = self.e_state[self.starts_ends[k,0]:self.starts_ends[k,1]]

            # if time warp
            #if self.time_warp == True:
            #    temp = resample_snippet(temp)

            if self.signs[k] == 1:
                if self.plotting:
                    ax1.plot(np.arange(temp.shape[0])/30, temp,
                        c=cmap_pos[ctr_pos])
                ctr_pos += 1
            else:
                if self.plotting:
                    ax2.plot(np.arange(temp.shape[0])/30, temp,
                        c=cmap_neg[ctr_neg])
                ctr_neg += 1

        
        #
        self.snippets = []
        self.snippets_warped = []
        self.snippets_to_thresh = []
        self.snippets_to_thresh_warped = []
        for k in range(len(self.ids_snippets_rewarded)):
            idx = self.ids_snippets_rewarded[k]
            temp1 = self.e_state[self.starts_ends[idx,0]:self.starts_ends[idx,1]]

            #
            self.snippets.append(temp1)

            #
            temp2 = resample_snippet(temp1.copy())
            #print ("temp: ", temp1.shape, " temp2.shape: ", temp2.shape)
                                      
            self.snippets_warped.append(temp2)

            # find the index up to where temp is below thresh
            temp_thresh = np.mean(self.thresh)
            idx = np.where(temp1<temp_thresh)[0]
            snippet_to_threshold = temp1[idx]
            #print ("snippet_to_threshold: ", snippet_to_threshold.shape)
            if snippet_to_threshold.shape[0]==0:
                continue

            self.snippets_to_thresh.append(snippet_to_threshold)

            #
            snippet_to_threshold_warped = resample_snippet(snippet_to_threshold.copy())
            self.snippets_to_thresh_warped.append(snippet_to_threshold_warped)

            #
            if self.plotting:
                ax3.plot(np.arange(temp2.shape[0])/30, temp2, c=cmap_r[k])

                #
                ax4.plot(np.arange(snippet_to_threshold_warped.shape[0])/30, 
                         snippet_to_threshold_warped, c=cmap_r[k])

        # plot average of temps
        self.snippets_warped = np.array(self.snippets_warped)
        print ("snippets_warped: ", self.snippets_warped.shape)
        mean = np.mean(self.snippets_warped,axis=0)
        self.mean = mean

        #
        self.snippets_to_thresh = np.array(self.snippets_to_thresh)
        self.snippets_to_thresh_warped = np.array(self.snippets_to_thresh_warped)
        self.mean_to_thresh = np.mean(self.snippets_to_thresh_warped,axis=0)

        #
        self.snippets = np.array(self.snippets, dtype=object)
        #print ("temps: ", temps.shape)
        if self.plotting:
            ax3.plot(np.arange(mean.shape[0])/30, mean ,
                    c='red', linewidth=3)

            #
            ax4.plot(np.arange(self.mean_to_thresh.shape[0])/30, 
                     self.mean_to_thresh ,
                    c='red', linewidth=3)

        # plot the thresh value
        if self.plotting:   
            plt.show()

    #
    def get_tone_triggered_ca_single_cell_static(self):

        tone_ids = np.arange(2,self.unique_tones.shape[0])
        #print ("tone_ids: ", tone_ids)

        #
        if self.F_type=='upphase':
            F = self.sessions[self.session_id].F_upphase_bin
        elif self.F_type == 'F_filtered':
            F = self.sessions[self.session_id].F_filtered

        # 
        vector = []
        for tone_id in tone_ids:
            
            #
            idx = np.where(np.int32(self.tones)==self.unique_tones[tone_id])[0]

            temp = np.zeros((self.tones.shape[0]))
            temp[idx] = 1

            # FIND BEGINNIGN AND ENDS OF tones
            from scipy.signal import chirp, find_peaks, peak_widths
            peaks, _ = find_peaks(temp)  # middle of the pluse/peak
            widths, heights, starts, ends = peak_widths(temp, peaks)
            starts = np.int32(starts)
            ends = np.int32(ends)

            #print ("duration of tones: ", widths/30, "sec")

            # index into temp between tone starts and ends
            cell_id_ca = []
            for k in range(len(starts)):
                temp = F[self.cell_id][starts[k]:ends[k]]
                cell_id_ca.append(temp)
            cell_id_ca = np.hstack(cell_id_ca)

            #
            #print ("cell_id_ca: ", cell_id_ca.shape)
            vector.append(np.mean(cell_id_ca,0))
       
        ######################################
        ######################################
        ######################################
        #
        vector = np.vstack(vector)
        #plt.figure()
        #plt.imshow(vector,
        #           aspect='auto',
        #           extent = [0,1,0,len(tone_ids)])
        plt.plot(vector,
                  )
        
        # plot xticks as unique tones and rotated 45 degrees
        plt.xticks(np.arange(len(tone_ids)),self.unique_tones[tone_ids],rotation=45)

        
        #
       # plt.show()

    #
    def get_ca_triggered_tone_all_cells(self):
        
        #
        
        #
        window = self.window
        vmax = self.vmax
        shuffle = self.shuffle

        #
        #reward_times = self.reward_times

        #


        # submit a particular type of trigger: tone_starts vs. all
        if self.tone_type == 'starts':
            tone_starts = self.tone_starts
        elif self.tone_type == 'all':
            tone_starts = self.tone_all

            #print ("selected all: ", tone_starts)

        ##
        if self.F_type=='upphase':
            F = self.sessions[self.session_id].F_upphase_bin
        elif self.F_type == 'F_filtered':
            F = self.sessions[self.session_id].F_filtered

        ######################################
        cell_ids = np.arange(self.sessions[self.session_id].F.shape[0])
        #cell_ids = np.arange(4)
        imgs = parmap.map(get_img_tone,
                          cell_ids,
                          F,
                          window,
                          self.tones,
                          shuffle,
                          self.smooth,
                          self.tone_type,
                          self.remove_base_tone,
                          pm_processes=16,
                          pm_pbar=True)
       
        ######################################
        ######################################
        ######################################

        #imgs = np.vstack(imgs)

        ctr=0
        k=0
        plt.figure(figsize=(12,12))
        while True:
            if ctr%100==0 and ctr>0:

                self.savefig_tones(ctr)

                plt.figure(figsize=(12,12))
                k=0

            ###################################### 
            ax=plt.subplot(10,10,k+1)
            if self.vmax==None:
                vmax = np.max(imgs[ctr])
                vmin = -vmax
                plt.imshow(imgs[ctr],
                        aspect='auto',
                        interpolation='none',cmap=self.cmap,
                        vmin=vmin,
                        vmax=vmax
                        #vmax=self.vmax
                        )
    
            else:
                #if self.vmax < np.max(imgs[ctr]):
                #    vmax = np.max(imgs[ctr])
                #else:
                vmax = self.vmax
                vmin = -vmax

                plt.imshow(imgs[ctr],
                        aspect='auto',
                        interpolation='none',cmap=self.cmap,
                        vmin=-vmax,
                        vmax=vmax
                        )
                
            plt.xticks([])
            plt.yticks([])
            #plt.ylim(-0.5,imgs[k].shape[0]-0.5)
            # plot y ticks as tones values
            if k == 0:
                plt.ylabel("Tone value (hz)")
                
                # make a list of strings that contains the tone values and the number of times they appear
                tone_labels = np.zeros((self.unique_tones.shape[0]-2,2))
                tone_labels[:,0] = self.unique_tones[2:]
                tone_labels[:,1] = self.counts[2:]

                ls = []
                for q in range(2, self.unique_tones.shape[0],1):
                    x = self.unique_tones[q]
                    y = tone_starts[q-2].shape[0]
                    #print ("x: ", x, " y: ", y  )
                    ls.append([str(x)+"hz (# "+str(y)+")"])

                ls = np.array(ls)
                #
                #print (np.arange(imgs[k].shape[0]))
                #print ("tone_labels: ", tone_labels)
                #plt.yticks(np.arange(imgs[k].shape[0]),
                #           ls[:,0],
                #            fontsize=5)
                

                # plot also time ticks on x axis but only the first and last and zero
                #plt.xticks([0,imgs[k].shape[1]/2,imgs[k].shape[1]-1],
                #            [str(-window/30),0,str(window/30)],
                #            fontsize=5)
                
                # print xlabel also but with padding so it's close to the image
                plt.xlabel("Time (s)",fontsize=5,labelpad=0.1)

                # plot colorbar with label "DFF" and without shrinking it and with fontsize 5 and with only 3 ticks
                clb = plt.colorbar(shrink=1, ticks=[-vmin,0,vmax])
                clb.ax.tick_params(labelsize=5)

                # put the colorbar at the top of the plot
                clb.ax.set_title('DFF')

            # plot title with padding to be close to the image
            if k!=10:
                plt.title(str(ctr),pad=0.1)

            # plot a vertical line at 0
            #plt.plot([window,window],[0,imgs[k].shape[0]],'--',c='grey',linewidth=3, alpha=0.4)

            #####################
            ctr+=1
            k+=1

            # 
            if ctr == len(imgs):

                self.savefig_tones(ctr)

                break

  

    #
    def get_tone_triggered_ca_all_cells(self):
        
        #
        
        #
        window = self.window
        vmax = self.vmax
        shuffle = self.shuffle

        #
        #reward_times = self.reward_times

        #


        # submit a particular type of trigger: tone_starts vs. all
        if self.tone_type == 'starts':
            tone_starts = self.tone_starts
        elif self.tone_type == 'all':
            tone_starts = self.tone_all

            #print ("selected all: ", tone_starts)

        ##
        if self.F_type=='upphase':
            F = self.sessions[self.session_id].F_upphase_bin
        elif self.F_type == 'F_filtered':
            F = self.sessions[self.session_id].F_filtered

        ######################################
        cell_ids = np.arange(self.sessions[self.session_id].F.shape[0])
        imgs = parmap.map(get_img_ca,
                          cell_ids,
                          F,
                          window,
                          tone_starts,
                          shuffle,
                          self.smooth,
                          pm_processes=16,
                          pm_pbar=True)
       
        ######################################
        ######################################
        ######################################

        # do a 10 x 10 grid showing each element of imgs using imshow
        ctr=0
        k=0
        plt.figure(figsize=(12,12))
        while True:
            if ctr%100==0 and ctr>0:

                self.savefig(ctr)

                plt.figure(figsize=(12,12))
                k=0

            ###################################### 
            ax=plt.subplot(10,10,k+1)
            if self.vmax==None:
                vmax = np.max(imgs[ctr])
                vmin = -vmax
                plt.imshow(imgs[ctr],
                        aspect='auto',
                        interpolation='none',cmap=self.cmap,
                        vmin=vmin,
                        vmax=vmax
                        #vmax=self.vmax
                        )
    
            else:
                #if self.vmax < np.max(imgs[ctr]):
                #    vmax = np.max(imgs[ctr])
                #else:
                vmax = self.vmax
                vmin = -vmax

                plt.imshow(imgs[ctr],
                        aspect='auto',
                        interpolation='none',cmap=self.cmap,
                        vmin=-vmax,
                        vmax=vmax
                        )
                
            plt.xticks([])
            plt.yticks([])
            plt.ylim(-0.5,imgs[k].shape[0]-0.5)
            # plot y ticks as tones values
            if k == 0:
                plt.ylabel("Tone value (hz)")
                
                # make a list of strings that contains the tone values and the number of times they appear
                tone_labels = np.zeros((self.unique_tones.shape[0]-2,2))
                tone_labels[:,0] = self.unique_tones[2:]
                tone_labels[:,1] = self.counts[2:]

                ls = []
                for q in range(2, self.unique_tones.shape[0],1):
                    x = self.unique_tones[q]
                    y = tone_starts[q-2].shape[0]
                    #print ("x: ", x, " y: ", y  )
                    ls.append([str(x)+"hz (# "+str(y)+")"])

                ls = np.array(ls)
                #
                #print (np.arange(imgs[k].shape[0]))
                #print ("tone_labels: ", tone_labels)
                plt.yticks(np.arange(imgs[k].shape[0]),
                           ls[:,0],
                            fontsize=5)
                

                # plot also time ticks on x axis but only the first and last and zero
                plt.xticks([0,imgs[k].shape[1]/2,imgs[k].shape[1]-1],
                            [str(-window/30),0,str(window/30)],
                            fontsize=5)
                
                # print xlabel also but with padding so it's close to the image
                plt.xlabel("Time (s)",fontsize=5,labelpad=0.1)

                # plot colorbar with label "DFF" and without shrinking it and with fontsize 5 and with only 3 ticks
                clb = plt.colorbar(shrink=1, ticks=[-vmin,0,vmax])
                clb.ax.tick_params(labelsize=5)

                # put the colorbar at the top of the plot
                clb.ax.set_title('DFF')

            # plot title with padding to be close to the image
            if k!=10:
                plt.title(str(ctr),pad=0.1)

            # plot a vertical line at 0
            plt.plot([window,window],[0,imgs[k].shape[0]],'--',c='grey',linewidth=3, alpha=0.4)

            #####################
            ctr+=1
            k+=1

            # 
            if ctr == len(imgs):

                self.savefig(ctr)

                break


    def savefig_tones(self, ctr):   
        
        # plot suptitle with animal id, session id and number of rewards
        plt.suptitle("Animal: "+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(self.n_rewards)
                    #+ "\n"+
                    #str(self.unique_tones[tone_id])+"hz" + ", #: "+str(len(starts))
                    )

        plt.show()
        
        #
        if self.shuffle == False:
            fname_out = os.path.join(self.root_dir,
                                self.animal_id,
                                self.session_ids[self.session_id],
                                'all_tones_'+str(ctr)+"_"+
                                str(self.F_type)+
                                "_reward_lockout_"+str(self.remove_rewards)+    
                                "_window_"+str(self.window//30)+
                                "_tone_type_"+self.tone_type+
                                '.png')
        else:
            fname_out = os.path.join(self.root_dir,
                                self.animal_id,
                                self.session_ids[self.session_id],
                                'all_tones_shuffled_'+str(ctr)+
                                "_"+str(self.F_type)+
                                "_reward_lockout_"+str(self.remove_rewards)+                                
                                "_window_"+str(self.window//30)+
                                "_tone_type_"+self.tone_type+
                                '.png')
        #
        plt.savefig(fname_out, dpi=300)
        plt.close()
        

    def savefig(self, ctr):   
        
        # plot suptitle with animal id, session id and number of rewards
        plt.suptitle("Animal: "+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(self.n_rewards)
                    #+ "\n"+
                    #str(self.unique_tones[tone_id])+"hz" + ", #: "+str(len(starts))
                    )

        plt.show()
        
        #
        if self.shuffle == False:
            fname_out = os.path.join(self.root_dir,
                                self.animal_id,
                                self.session_ids[self.session_id],
                                'all_cells_'+str(ctr)+"_"+
                                str(self.F_type)+
                                "_reward_lockout_"+str(self.remove_rewards)+    
                                "_window_"+str(self.window//30)+
                                "_tone_type_"+self.tone_type+
                                '.png')
        else:
            fname_out = os.path.join(self.root_dir,
                                self.animal_id,
                                self.session_ids[self.session_id],
                                'all_cells_shuffled_'+str(ctr)+
                                "_"+str(self.F_type)+
                                "_reward_lockout_"+str(self.remove_rewards)+                                
                                "_window_"+str(self.window//30)+
                                "_tone_type_"+self.tone_type+
                                '.png')
        #
        plt.savefig(fname_out, dpi=300)
        plt.close()
        
    #
    def get_unique_tones(self, plotting=False):

        #       
        # get all unique values in tone and the number of times they appear
        self.unique_tones, self.counts = np.unique(np.int32(self.tones), return_counts=True)

        # make a list of all unique tone indices
        unique_tones_idx = []
        for k in range(self.unique_tones.shape[0]):
            idx = np.where(self.tones==self.unique_tones[k])[0]
            unique_tones_idx.append(idx)

        # delete values from tones that occurs less than 200 times
        for k in range(len(unique_tones_idx)):
            if unique_tones_idx[k].shape[0]<200:
                self.tones[unique_tones_idx[k]]=0

        # recompute unique tones based on what's left
        self.unique_tones, self.counts = np.unique(self.tones, return_counts=True)

        #
        self.tone_all = []
        tone_starts = []
        reward_times = self.reward_times
        tone_ids = np.arange(2,self.unique_tones.shape[0])
        min_space_starts = self.window
        for tone_id in tone_ids:

            #
            ##############################################################
            idx = np.where(np.int32(self.tones)==self.unique_tones[tone_id])[0]
            
            # exclude tone timesthat are too close to rewards
            if self.remove_rewards:
                ss = []
                for k in range(idx.shape[0]):
                    if np.min(np.abs(idx[k]-reward_times))>min_space_starts:
                        ss.append(idx[k])
                temp = np.array(ss)
                #print ("tone: ", self.unique_tones[tone_id], " # of tones: ", temp.shape[0])
                self.tone_all.append(temp)

            ##############################################################
            # find the starts in each sequence of idx values
            starts = []
            for k in range(idx.shape[0]-1):
                if idx[k+1]-idx[k]>1:
                    starts.append(idx[k+1])
            starts = np.array(starts) #+3000

            # # plot the calcium events for the first tone
            #t = np.arange(-window,window)/30

            # remove starts that are too close to each other
            ss = []
            ss.append(starts[0])
            for k in range(1,starts.shape[0]):
                # skip k if too close to previous
                if (starts[k]-starts[k-1])>min_space_starts:
                    ss.append(starts[k])

            starts = np.array(ss)

            # remove starts that are too close to rewards
            if self.remove_rewards:
                ss = []
                for k in range(starts.shape[0]):
                    if np.min(np.abs(starts[k]-reward_times))>min_space_starts:
                        ss.append(starts[k])
                starts = np.array(ss)

            # 
            tone_starts.append(starts)

        #
        self.tone_starts = tone_starts


        # make a histogram of the tone values
        if plotting:
            plt.figure()
            y = np.histogram(self.tones, bins=np.arange(1,16500,100))
            plt.bar(y[1][:-1],y[0], width=100)

            #
            plt.xlabel("Tone value (hz)")
            plt.ylabel("# of frames")

            #
            plt.suptitle(self.animal_id+ " - " + self.session_ids[1])

            plt.show()

        # load the calcium events for all cells
        #print ("# cells: ", self.sessions[self.session_id].F.shape)
       # print ("# cells: ", self.sessions[self.session_id].F_upphase_bin.shape)


    #
    def load_trials(self):
        # make an fname_out directory if not already present
        session_name = self.session_ids[self.session_id]

        # #load the n_rewards
        # fname = os.path.join(self.root_dir,
        #                         self.animal_id,
        #                         session_name,
        #                         'results.npz')
        
        # #
        # results = np.load(fname, allow_pickle=True)
        # rewards = results['reward_times'].T
        # idx = np.where(rewards[:,1]>0)[0]
        # self.n_rewards = idx.shape[0]
        # reward_times = rewards[idx,1]
        #print ("# of rewards: ", self.n_rewards, " rewar[:100])

        # Load the workbook
        wb = load_workbook(os.path.join(self.root_dir,
                                        self.animal_id,
                                        session_name,
                                        'results_fixed.xlsx'), read_only=False)

        # Access the worksheet
        ws = wb[wb.sheetnames[0]]  # assuming you want the first sheet, change as needed

        # get white noise state  
        rows = [3,4,7,8,9,10,11]          # white_noise, tone_state, ensemble_state, reward_state
        self.threshold_state = []       # 3rd column
        self.white_noise_state = []    # 4th column
        self.tone_state = []           # 7th
        self.ensemble_state = []       # 8th
        self.reward_state = []          # 9th
        self.trials = []               # 10th
        self.trials_rewards = []       # 11th  this is the flag which indicates if trial was rewarded
        for row in ws.iter_rows(min_row=2,values_only=False):
            self.threshold_state.append(row[rows[0]-1].value)
            self.white_noise_state.append(row[rows[1]-1].value)
            self.tone_state.append(row[rows[2]-1].value)
            self.ensemble_state.append(row[rows[3]-1].value)
            self.reward_state.append(row[rows[4]-1].value)
            self.trials.append(row[rows[5]-1].value)
            self.trials_rewards.append(row[rows[6]-1].value)

        # make all lists into arrays
        self.threshold_state = np.array(self.threshold_state)
        self.white_noise_state = np.array(self.white_noise_state)
        self.tone_state = np.array(self.tone_state)
        self.ensemble_state = np.array(self.ensemble_state)
        self.reward_state = np.array(self.reward_state)
        self.trials = np.array(self.trials)
        self.trials_rewards = np.array(self.trials_rewards)

        # print examples from each of the above arrays
        if False:
            print ("threshold_state: ", self.threshold_state[:100])
            print ("white_noise_state: ", self.white_noise_state[:100])
            print ("tone_state: ", self.tone_state[:100])
            print ("ensemble_state: ", self.ensemble_state[:100])
            print ("reward_state: ", self.reward_state[:100])
            print ("trials: ", self.trials[:100])
            print ("trials_rewards: ", self.trials_rewards[:100])


    #
    def compute_warping_functions(self):

        #
        fps = 30
    

        #
        if self.burst_to_threshold_only==False:
            snippets = self.snippets.copy()
            snippets_warped = self.snippets_warped.copy()
            ts2 = self.mean
        else:
            snippets = self.snippets_to_thresh.copy()
            snippets_warped = self.snippets_to_thresh_warped.copy()
            ts2 = self.mean_to_thresh

        #
        t = np.arange(0, len(ts2), 1)/fps

        ##############################################
        ############## SETUP PLOTS ###################
        ##############################################
        if self.plotting:
            plt.figure(figsize=(15, 5))
            ax1 = plt.subplot(131)
            ax1.set_xlabel('Time (s)')
            if self.use_stretched:
                ax1.set_title('Stretched trial traces')
            else:
                ax1.set_title('Original trial traces')

            #
            ax2 = plt.subplot(132)
            ax2.set_xlabel('Time (normalized)')
            ax2.set_title('Time warped trial traces')

            #
            ax3 = plt.subplot(133)
            ax3.set_xlabel('Time (normalized)')
            ax3.set_title('Warping functions')

            if self.use_stretched:
                ax1.plot(t, ts2, color='red', linewidth=5,
                        label="average of stretched functions")# marker='o')
                ax1.legend()
            ax2.plot(t, ts2, color='red', linewidth=5,
                    label="average of stretched functions")# marker='o')
            ax2.legend()

            cmap_r = plt.cm.viridis(np.linspace(0, 1, len(self.snippets)))


        ##############################################
        ############## SETUP PLOTS ###################
        ##############################################
        warping_functions = []
        self.warped_snippets = []

        #
        for k in range(snippets.shape[0]):
            
            # option 1 do dynanmic time warping on the streetched
            if self.use_stretched:
                ts1 = snippets_warped[k]
            else:
                ts1 = snippets[k]


            t2 = np.arange(0, len(ts1), 1)/fps

            # Warp ts1 to best match ts2
            warp_function = warp_ts1_to_ts2(ts1, ts2)

            # Store the warping function
            warping_functions.append(warp_function)

            #
            warped_ts1 = interpolate_time_warp(ts1, ts2, warp_function)
            self.warped_snippets.append(warped_ts1)

            # Plot
            t = np.arange(0, len(warped_ts1), 1)/fps

            # normalize the warping function
            t3 = np.arange(0, len(warp_function), 1)/len(warp_function)

            if self.plotting:
                ax1.plot(t2, ts1, c=cmap_r[k])#, label='Warped Time Series 1', color='red', linestyle='dashed', marker='x')
                ax2.plot(t, warped_ts1, c=cmap_r[k])#, label='Warped Time Series 1', color='red', linestyle='dashed', marker='x')
                ax3.plot(t3, warp_function, c=cmap_r[k])#, label='Warping Function', color='red', linestyle='dashed', marker='x')

        self.warp_functions = warping_functions

        #
        if self.plotting:
            plt.show()

        # save warping functions
        warp_functions = np.array(warping_functions, dtype=object)
        np.save(os.path.join(self.root_dir,
                            self.animal_id,
                            self.session_ids[self.session_id],
                            'warp_functions.npy'), warp_functions, allow_pickle=True)
   
    #
    def plot_warped_cell_ca_trial_and_burst(self):

        if self.burst_sign==1:
            self.plot_warped_cell_ca_trial_and_burst_positive()

        else:
            self.plot_warped_cell_ca_trial_and_burst_negative()


    #
    def plot_warped_cell_ca_trial_and_burst_negative(self):

        #
        fps = 30
        
        #
        n_cells = self.sessions[self.session_id].F_upphase_bin.shape[0]
        print ("n_cells: ", n_cells, "# bursts: ", len(self.ids_snippets))

        #
        ids_snippets_local = self.ids_snippets.copy()

        #

        #
        cell_ids = np.arange(self.sessions[self.session_id].F_upphase_bin.shape[0])
        #line_burst = np.zeros(301)
        line_trial = np.zeros(400)
        burst_raw = []
        pre_burst_raw = []
        post_burst_raw = []
        burst_warped = []
        pre_burst_warped = []
        for cell_id in tqdm(cell_ids):
            
            #
            ts2 = self.mean.copy()

            #
            trace = self.sessions[self.session_id].F_upphase_bin[cell_id].squeeze()

            # get the burst segment and warp it
            stack1 = []
            stack2 = []

            #
            for k in range(len(ids_snippets_local)):

                # get start-ends from the saved snippets for the negative bursts
                se = self.starts_ends[ids_snippets_local[k]]

                try:
                    ts1 = trace[se[0]:se[1]]
                except:
                    continue
    
                #temp = line.copy()*0
                #temp[0:ts1.shape[0]] = ts1

                # get the warp function of current snippet to average mean
                warp_function = warp_ts1_to_ts2(ts1, ts2.copy())

                # apply warp function and resample to fixed points
                warped_ca = interpolate_time_warp(ts1, ts2.copy(), warp_function)
                
                if warped_ca.shape[0]!=30:
                    continue

                stack1.append(ts1)
                stack2.append(warped_ca)


            #
            #burst_raw.append(np.vstack(stack1))
            burst_warped.append(np.vstack(stack2))

            # get the pre-burst segment and warp it
            stack1 = []
            stack2 = []
            for k in range(len(ids_snippets_local)):
                idx = ids_snippets_local[k]
                se = self.starts_ends[idx]

                #
                ts1 = trace[se[0]-fps*10:se[0]]

                #if ts1.shape[0]!=fps*10:
                #    print ("skipping pre-burst segment", k)
                #    continue

                temp = line_trial.copy()*0
                temp[0:ts1.shape[0]] = ts1
                ts1=temp

                #
                stack1.append(ts1)
                
                # split the ts1 array every 3rd frame and compute the mean in each chunk
                ts2 = np.mean(np.split(ts1, 100), axis=1)
                #print ("ts2: ", ts2.shape)
                stack2.append(ts2)            
        
            #
            pre_burst_raw.append(np.vstack(stack1))
            pre_burst_warped.append(np.vstack(stack2))

            # do the same for post-burst segment and warp it
            stack1 = []
            for k in range(len(ids_snippets_local)):
                idx = ids_snippets_local[k]
                se = self.starts_ends[idx]

                # grab the following 10 sec of data after the se ends
                ts1 = trace[se[1]:se[1]+300]

                #
                ts1 = scipy.signal.resample(ts1, 100)
                stack1.append(ts1)

            post_burst_raw.append(np.vstack(stack1))            

        ############################################################
        #################### POST PROCESSING #######################
        ############################################################

        # make traces into arrays
        #burst_raw = np.array(burst_raw).squeeze()
        burst_warped = np.array(burst_warped).squeeze()
        print ("burst_warped: ", burst_warped.shape)

        # same for pre-burst
        pre_burst_raw = np.array(pre_burst_raw)
        pre_burst_warped = np.array(pre_burst_warped)
        print ("pre_burst_raw: ", pre_burst_raw.shape, " pre_burst_warped: ", pre_burst_warped.shape)

        # same for post-burst
        post_burst_raw = np.array(post_burst_raw)
        print ("post_burst_raw: ", post_burst_raw.shape)

        # compute mean in the 0th axis
        # burst_raw_mean = np.mean(burst_raw, axis=1)
        # print ("burst_raw_mean: ", burst_raw_mean.shape)
        # burst_warped_mean = np.mean(burst_warped, axis=1)

        # make stack of all traces
        traces_all = [] 
        for k in range(pre_burst_warped.shape[0]):
            if k==0:
                print (pre_burst_warped[k].shape,
                                            burst_warped[k].shape,
                                            post_burst_raw[k].shape)
            traces_all.append(np.hstack((pre_burst_warped[k],
                                         burst_warped[k],
                                         post_burst_raw[k])))
        traces_all = np.array(traces_all)
        print ("traces_all: ", traces_all.shape)
        
        #
        traces_all_mean = np.mean(traces_all, axis=1)

        # find the max of each cell in the means and reorder the traces
        traces_all_mean_max = np.max(traces_all_mean, axis=1)

        # sort the traces
        traces_all_mean_sorted = np.argsort(traces_all_mean_max)[::-1]

        # reorder the traces
        traces_all = traces_all[traces_all_mean_sorted]


        ############################################################
        ####################### PLOTTING ###########################
        ############################################################
        # plot first 100 of each
        plt.figure(figsize=(10,10))
        plt.suptitle("Animal: "+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(len(self.ids_snippets_rewarded))
                    #+ "\n"+
                    #str(self.unique_tones[tone_id])+"hz" + ", #: "+str(len(starts))
                    )
        ctr=0
        start = self.start_idx
        for k in range(0,36,1):
            plt.subplot(6,6,ctr+1)
            plt.title(str(k),pad=0.1, fontsize=7)
            img = traces_all[start+k]
            plt.imshow(img,
                        aspect='auto')
            plt.xticks([])
            plt.yticks([])
            
            if k==0:
                plt.title("Pre-burst (10s) | Burst (1-5s)| Post (10s)",fontsize=7)

            # plot vertical line at 900 and 1100
            plt.plot([100,100],[-0.5,img.shape[0]-0.5],'--',c='red',linewidth=1)
            plt.plot([130,130],[-0.5,img.shape[0]-0.5],'--',c='red',linewidth=1)


            ctr+=1

        # plot first 100 cells average curve
        plt.figure(figsize=(10,10))
        plt.suptitle("Animal: "+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(len(self.ids_snippets_rewarded))
                    #+ "\n"+
                    #str(self.unique_tones[tone_id])+"hz" + ", #: "+str(len(starts))
                    )
        ctr=0
        for k in range(0,36,1):
            plt.subplot(6,6,ctr+1)

            # plot title with low padding
            plt.title(str(k),pad=0.1, fontsize=7)

            #
            mean = np.mean(traces_all[start+k],axis=0)
            t = np.arange(0, mean.shape[0], 1)/30
            plt.plot(t, mean)
            
            # plot vertical line at 900 and 1100
            plt.plot([t[100],t[100]],[0,1],'--',c='grey',linewidth=1, alpha=0.4)
            plt.plot([t[130],t[130]],[0,1],'--',c='grey',linewidth=1, alpha=0.4)

            # plot horizontal line at 0.3
            plt.plot([t[0],t[-1]],[0.3,0.3],'--',c='grey',linewidth=3, alpha=0.4)

            #
            if k==0:
                plt.title("Pre-burst (10s) | Burst (1-5s)| Post (10s)",fontsize=7)

            if k>0:
                plt.yticks([])

            #if k==90:
            #else:
            plt.xticks([])
            plt.ylim(0,1)

            #
            plt.xlim(t[0],t[-1])

            #
            ctr+=1


    #
    def plot_warped_cell_ca_trial_and_burst_positive(self):

        n_cells = self.sessions[self.session_id].F_upphase_bin.shape[0]
        # if self.burst_sign==1:
        #     print ("n_cells: ", n_cells, "# bursts: ", len(self.ids_snippets_rewarded))
        #     ids_snippets_local = self.ids_snippets_rewarded.copy()
        # else:
        #     print ("n_cells: ", n_cells, "# bursts: ", len(self.ids_snippets))
        #     ids_snippets_local = self.ids_snippets.copy()


 #
        if self.burst_sign==1:
            ids_snippets_local = self.ids_snippets_rewarded.copy()
        else:
            ids_snippets_local = self.ids_snippets.copy()
            
        #
        if self.burst_to_threshold_only==False:
            #snippets = self.snippets.copy()
            #snippets_warped = self.snippets_warped.copy()
            ts2 = self.mean
            starts_ends_local = self.starts_ends.copy()[ids_snippets_local]
        else:
            #snippets = self.snippets_to_thresh.copy()
            #snippets_warped = self.snippets_to_thresh_warped.copy()
            ts2 = self.mean_to_thresh
            starts_ends_local = self.starts_ends_to_threshold.copy()


        #
        ts2 = self.mean

        #
        cell_ids = np.arange(self.sessions[self.session_id].F_upphase_bin.shape[0])
        line_burst = np.zeros(400)
        line_trial = np.zeros(901)
        burst_raw = []
        pre_burst_raw = []
        post_burst_raw = []
        burst_warped = []
        pre_burst_warped = []
        for cell_id in tqdm(cell_ids):
            
            if self.ca_type == 'upphase':
                trace = self.sessions[self.session_id].F_upphase_bin[cell_id].squeeze()
            elif self.ca_type == 'DFF':
                trace = self.sessions[self.session_id].F_filtered[cell_id].squeeze()

            # get the burst segment and warp it
            stack1 = []
            stack2 = []
            for k in range(len(starts_ends_local)):
                #idx = ids_snippets_local[k]
                se = starts_ends_local[k]

                if se[1]-se[0]<2:
                    continue

                try:
                    ts1 = trace[se[0]:se[1]]
                except:
                    continue
                # 
                temp = line_burst.copy()*0
                temp[0:ts1.shape[0]] = ts1
                stack1.append(temp)

                # get the warp function of current snippet to average mean
                warp_function = warp_ts1_to_ts2(ts1, ts2)

                # apply warp function and resample to fixed points
                warped_ca = interpolate_time_warp(ts1, ts2, warp_function)
                #print (cell_id, k, "warped_ca: ", warped_ca.shape, "warp_function: ", warp_function.shape)
                stack2.append(warped_ca)

            #
            burst_raw.append(np.vstack(stack1))
            burst_warped.append(np.vstack(stack2))

            # get the pre-burst segment and warp it
            stack1 = []
            stack2 = []
            for k in range(len(starts_ends_local)):
                #idx = ids_snippets_local[k]
                se = starts_ends_local[k]

                if se[1]-se[0]<2:
                    continue

                # find the burst ID
                burst_seg = self.trials[se[0]:se[1]]
                idx = np.where(burst_seg==1000)[0]
                # delete these indexes
                burst_seg = np.delete(burst_seg, idx)

                # check if all values are identical and grab the value
                if np.all(burst_seg==burst_seg[0]):
                    trial_id = burst_seg[0]
                else:
                    print ("Error: burst not matching to a unique trial")
                
                # grab the calcium during the trial
                idx = np.where(self.trials==trial_id)[0]

                # This is wrong!  It also includes the burst time points so we're counting them twice
                # need to delete the time points from idx that are also in se
                se_idx = np.arange(se[0],se[1])
                idx = np.delete(idx, np.where(np.isin(idx, se_idx))[0])
                #print ("idx: ", idx.shape)

                if idx.shape[0]==0:
                    ts1 = np.zeros(901)    
                else:                      
                    ts1 = trace[idx]

                #print ("[ca] snippet: ", ts1.shape)

                temp = line_trial.copy()*0
                temp[-ts1.shape[0]:] = ts1
                stack1.append(temp)

                # use scipy to resample ts1 to be the size of line_trial.shape[0]
                ts1 = scipy.signal.resample(ts1, 200)
                stack2.append(ts1)
            
                #return
        
            #
            pre_burst_raw.append(np.vstack(stack1))
            pre_burst_warped.append(np.vstack(stack2))

            # do the same for post-burst segment and warp it
            stack1 = []
            for k in range(len(ids_snippets_local)):

                se = starts_ends_local[k]

                if se[1]-se[0]<2:
                    continue
            
                idx = ids_snippets_local[k]
                se = self.starts_ends[idx]

                # grab the following 10 sec of data after the se ends
                ts1 = trace[se[1]:se[1]+300]

                #
                ts1 = scipy.signal.resample(ts1, 100)
                stack1.append(ts1)

            post_burst_raw.append(np.vstack(stack1))            

        ############################################################
        #################### POST PROCESSING #######################
        ############################################################

        # make traces into arrays
        burst_raw = np.array(burst_raw).squeeze()
        burst_warped = np.array(burst_warped).squeeze()
        print ("burst_raw: ", burst_raw.shape)

        # same for pre-burst
        pre_burst_raw = np.array(pre_burst_raw)
        pre_burst_warped = np.array(pre_burst_warped)

        # same for post-burst
        post_burst_raw = np.array(post_burst_raw)
        print ("post_burst_raw: ", post_burst_raw.shape)

        # compute mean in the 0th axis
        burst_raw_mean = np.mean(burst_raw, axis=1)
        print ("burst_raw_mean: ", burst_raw_mean.shape)
        burst_warped_mean = np.mean(burst_warped, axis=1)

        # make stack of all traces
        traces_all = [] 
        for k in range(burst_raw.shape[0]):
            if k==0:
                print (pre_burst_warped[k].shape,
                                            burst_warped[k].shape,
                                            post_burst_raw[k].shape)
            traces_all.append(np.hstack((pre_burst_warped[k],
                                         burst_warped[k],
                                         post_burst_raw[k])))
        traces_all = np.array(traces_all)
        print ("traces_all: ", traces_all.shape)
        
        #
        traces_all_mean = np.mean(traces_all, axis=1)

        # find the max of each cell in the means and reorder the traces
        traces_all_mean_max = np.max(traces_all_mean, axis=1)

        # sort the traces
        traces_all_mean_sorted = np.argsort(traces_all_mean_max)[::-1]

        # reorder the traces
        traces_all = traces_all[traces_all_mean_sorted]


        ############################################################
        ####################### PLOTTING ###########################
        ############################################################
        # plot first 100 of each
        plt.figure(figsize=(10,10))
        plt.suptitle("Animal: "+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(len(self.ids_snippets_rewarded))
                    #+ "\n"+
                    #str(self.unique_tones[tone_id])+"hz" + ", #: "+str(len(starts))
                    )
        ctr=0
        start = self.start_idx
        for k in range(0,36,1):
            plt.subplot(6,6,ctr+1)
            plt.title(str(k),pad=0.1, fontsize=7)
            img = traces_all[start+k]
            plt.imshow(img,
                        aspect='auto')
            plt.xticks([])
            plt.yticks([])
            
            if k==0:
                plt.title("Trial (5-30s) | Burst (1-5s)| Post (10s)",fontsize=7)

            # plot vertical line at 900 and 1100
            plt.plot([200,200],[-0.5,img.shape[0]-0.5],'--',c='red',linewidth=1)
            plt.plot([230,230],[-0.5,img.shape[0]-0.5],'--',c='red',linewidth=1)


            ctr+=1

        # plot first 100 cells average curve
        plt.figure(figsize=(10,10))
        plt.suptitle("Animal: "+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(len(self.ids_snippets_rewarded))
                    #+ "\n"+
                    #str(self.unique_tones[tone_id])+"hz" + ", #: "+str(len(starts))
                    )
        ctr=0
        for k in range(0,36,1):
            plt.subplot(6,6,ctr+1)

            # plot title with low padding
            plt.title(str(k),pad=0.1, fontsize=7)

            #
            mean = np.mean(traces_all[start+k],axis=0)
            t = np.arange(0, mean.shape[0], 1)/30
            plt.plot(t, mean)
            
            # plot vertical line at 900 and 1100
            plt.plot([t[200],t[200]],[0,1],'--',c='grey',linewidth=1, alpha=0.4)
            plt.plot([t[230],t[230]],[0,1],'--',c='grey',linewidth=1, alpha=0.4)

            # plot horizontal line at 0.3
            plt.plot([t[0],t[-1]],[0.3,0.3],'--',c='grey',linewidth=3, alpha=0.4)

            #
            if k==0:
                plt.title("Trial (5-30s) | Burst (1-5s)| Post (10s)",fontsize=7)

            if k>0:
                plt.yticks([])

            #if k==90:
            #else:
            plt.xticks([])
            plt.ylim(0,1)

            #
            plt.xlim(t[0],t[-1])

            #
            ctr+=1

    #
    def plot_warped_cell_ca_burst_only(self):

        n_cells = self.sessions[self.session_id].F_upphase_bin.shape[0]
        print ("n_cells: ", n_cells, "# rewarsds: ", len(self.ids_snippets_rewarded))

        #
        if self.burst_sign==1:
            ids_snippets_local = self.ids_snippets_rewarded.copy()
        else:
            ids_snippets_local = self.ids_snippets.copy()
            
        #
        if self.burst_to_threshold_only==False:
            #snippets = self.snippets.copy()
            #snippets_warped = self.snippets_warped.copy()
            ts2 = self.mean
            starts_ends_local = self.starts_ends.copy()[ids_snippets_local]
        else:
            #snippets = self.snippets_to_thresh.copy()
            #snippets_warped = self.snippets_to_thresh_warped.copy()
            ts2 = self.mean_to_thresh
            starts_ends_local = self.starts_ends_to_threshold.copy()
        #
        #print ("ts2: ", ts2.shape, " starts_ends_local: ", starts_ends_local)

        #
        cell_ids = np.arange(self.sessions[self.session_id].F_upphase_bin.shape[0])
        line = np.zeros(400)
        traces_raw = []
        traces_warped = []
        for ctr, cell_id in enumerate(cell_ids):
            
            trace = self.sessions[self.session_id].F_upphase_bin[cell_id].squeeze()

            #
            stack1 = []
            stack2 = []
            #for k in range(len(self.ids_snippets_rewarded)):
            for k in range(len(starts_ends_local)):

                # get start-ends from the saved snippets for the negative bursts
                #print ("k: ", k, " starts_ends_local[k]: ", starts_ends_local[k])
                se = starts_ends_local[k]

                try:
                    ts1 = trace[se[0]:se[1]]
                except:
                    continue
    
                temp = line.copy()*0
                temp[0:ts1.shape[0]] = ts1
                stack1.append(temp)

                # get the warp function of current snippet to average mean
                warp_function = warp_ts1_to_ts2(ts1, ts2)

                # apply warp function and resample to fixed points
                warped_ca = interpolate_time_warp(ts1, ts2, warp_function)
                stack2.append(warped_ca)

            #
            traces_raw.append(np.vstack(stack1))
            traces_warped.append(np.vstack(stack2))

        # make traces into arrays
        traces_raw = np.array(traces_raw)
        print ("traces-raw: ", traces_raw.shape)
        traces_warped = np.array(traces_warped)

        # compute mean in the 0th axis
        traces_raw_mean = np.mean(traces_raw, axis=1)
        print ("traces mean: ", traces_raw_mean.shape)
        traces_warped_mean = np.mean(traces_warped, axis=1)

        # find the max of each cell in the means and reorder the traces
        traces_raw_mean_max = np.max(traces_raw_mean, axis=1)
        traces_warped_mean_max = np.max(traces_warped_mean, axis=1)

        # sort the traces
        traces_raw_mean_max_sorted = np.argsort(traces_raw_mean_max)[::-1]
        traces_warped_mean_max_sorted = np.argsort(traces_warped_mean_max)[::-1]

        # reorder the traces
        traces_raw = traces_raw[traces_raw_mean_max_sorted]
        traces_warped = traces_warped[traces_warped_mean_max_sorted]

        # plot first 100 of each
        plt.figure(figsize=(10,10))
        ctr=0
        start = self.start_idx
        end = start+100
        for k in range(0,100,1):
            plt.subplot(10,10,ctr+1)
            plt.title(str(k),pad=0.1, fontsize=7)
            plt.imshow(traces_warped[start+k],
                    aspect='auto')
            plt.xticks([])
            plt.yticks([])
            ctr+=1

        # plot first 100 cells average curve
        plt.figure(figsize=(10,10))
        ctr=0
        for k in range(0,100,1):
            plt.subplot(10,10,ctr+1)

            # plot title with low padding
            plt.title(str(k),pad=0.1, fontsize=7)

            #
            temp = traces_warped[start+k]
            #print (temp.shape)
            temp = np.mean(temp, axis=0)
            plt.plot(temp)

            # plot horizontal line at 0.3
            plt.plot([0,temp.shape[0]],[0.3,0.3],'--',c='grey',linewidth=3, alpha=0.4)

            #
            plt.xticks([])
            if k>0:
                plt.yticks([])
            plt.ylim(0,1)
            ctr+=1


    def get_rewarded_trials_align_trial_aligned(self):

        #
        fps=30

        #
        buffer = self.buffer #10*30
        window = self.window + buffer

        #
        img_rewarded = np.zeros((len(self.trial_ids_rewarded),window),dtype=np.float32)
        img_non_rewarded = np.zeros((len(self.trial_ids_not_rewarded),window),dtype=np.float32)

        line = np.zeros(window)+np.nan

        plt.figure(figsize=(12,6))
        plt.suptitle("Ensemble state"+
                     "\n"+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(len(self.trial_ids_rewarded))
                    )
        ax1 = plt.subplot(121)
        ax1.set_title("Rewarded trials")
        ax1.set_xlabel("Time (s)")
        ax1.set_ylabel("Trial #")
        ax2 = plt.subplot(122)
        ax2.set_title("Non rewarded trials")
        ax2.set_xlabel("Time (s)")
        ax2.set_ylabel("Trial #")

        # add the c.ensemble_state triggered on the beggining and ends of each rewarded trial
        for k in range(len(self.trial_ids_rewarded)):
            # get trial id
            i = self.trial_ids_rewarded[k]

            # get times of trial
            idx = np.where(self.trials==i)[0]
            temp = self.ensemble_state[idx[0]:idx[-1]+buffer-1]
            line*=0 + np.nan
            line[:temp.shape[0]] =  temp
            img_rewarded[k,:] = line
 
            # plot reward end for each trial as a red dot at 
            if k ==0:
                ax1.plot([idx.shape[0]/fps, idx.shape[0]/fps],
                     [k-0.5,k+0.5],c='r', linewidth=3, label='Reward time')
            else:
                ax1.plot([idx.shape[0]/fps, idx.shape[0]/fps],
                     [k-0.5,k+0.5],c='r', linewidth=3)

            


        # same for non rewarded trials
        for k in range(len(self.trial_ids_not_rewarded)):
            # get trial id
            i = self.trial_ids_not_rewarded[k]

            # get times of trial
            idx = np.where(self.trials==i)[0]
            temp = self.ensemble_state[idx[0]:idx[-1]+buffer-1]
            line*=0 + np.nan
            line[:temp.shape[0]] =  temp
            img_non_rewarded[k,:] = line

            # 
            # draw a verttical red line of height 1 at idx.shape[0]/fps
            ax2.plot([idx.shape[0]/fps, idx.shape[0]/fps],
                     [k-0.5,k+0.5],c='r', linewidth=3)     
            
            
        # invert the images upside down
        img_rewarded = np.flipud(img_rewarded)
        img_non_rewarded = np.flipud(img_non_rewarded)
        
       
        vmax = np.median(self.threshold_state)
        ax1.imshow(img_rewarded,aspect='auto',
                extent=[0,window/fps,-0.5,len(self.trial_ids_rewarded)-0.5],
                    interpolation='none',
                    cmap = self.cmap,
                    vmin = -vmax,
                    vmax = vmax  
                    )
        ax1.legend()


        clb = ax2.imshow(img_non_rewarded,aspect='auto',
                extent=[0,window/fps,-0.5,len(self.trial_ids_not_rewarded)-0.5],
                    interpolation='none',
                    cmap = self.cmap,
                    vmin = -vmax,
                    vmax = vmax  
                    )

        # show colorbar for ax2 with title "DFF"
        clb = plt.colorbar(clb,ax=ax2,shrink=1).set_label("Ensemble state (DFF units)")
        
        
    def get_rewarded_trials_align_reward_aligned_warped(self):

        fps=30

        #
        buffer = self.buffer #10*30
        window = self.window + buffer

        #
        img_rewarded = np.zeros((len(self.trial_ids_rewarded),window),dtype=np.float32)
        img_non_rewarded = np.zeros((len(self.trial_ids_not_rewarded),window),dtype=np.float32)

        line = np.zeros(window)+np.nan

        plt.figure(figsize=(6,6))
        plt.suptitle("Ensemble state"+
                     "\n"+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(len(self.trial_ids_rewarded))
                    )
        ax1 = plt.subplot(111)
        ax1.set_title("Rewarded trials")
        ax1.set_xlabel("Time (s)")
        ax1.set_ylabel("Trial #")
        # ax2 = plt.subplot(122)
        # ax2.set_title("Non rewarded trials")
        # ax2.set_xlabel("Time (s)")
        # ax2.set_ylabel("Trial #")

        # add the c.ensemble_state triggered on the beggining and ends of each rewarded trial
        for k in range(len(self.trial_ids_rewarded)):
            # get trial id
            i = self.trial_ids_rewarded[k]

            # get times of trial
            idx = np.where(self.trials==i)[0]
            temp_pre = self.ensemble_state[idx[0]:idx[-1]]
            temp_post = self.ensemble_state[idx[-1]:idx[-1]+buffer]
            line*=0 + np.nan

            # resample temp_pre to have the same length as window
            temp_pre = scipy.signal.resample(temp_pre, self.window)
           # print ("temp_pre: ", temp_pre.shape, "line: ", line.shape)
            
            line[:self.window] = temp_pre
            line[self.window:self.window+temp_post.shape[0]] =  temp_post
            
            img_rewarded[k,:] = line
 
            # plot reward end for each trial as a red dot at 
            if k ==0:
                ax1.plot([self.window/fps, self.window/fps],
                     [k-0.5,k+0.5],c='r', linewidth=3, label='reward time')
            else:
                ax1.plot([self.window/fps, self.window/fps],
                     [k-0.5,k+0.5],c='r', linewidth=3)

            
           
        # invert the images upside down
        img_rewarded = np.flipud(img_rewarded)
        # img_non_rewarded = np.flipud(img_non_rewarded)
        
       
        vmax = np.median(self.threshold_state)
        clb = ax1.imshow(img_rewarded,aspect='auto',
                extent=[0,window/fps,-0.5,len(self.trial_ids_rewarded)-0.5],
                    interpolation='none',
                    cmap = self.cmap,
                    vmin = -vmax,
                    vmax = vmax  
                    )
        ax1.legend()


        # clb = ax2.imshow(img_non_rewarded,aspect='auto',
        #         extent=[0,window/fps,-0.5,len(self.trial_ids_not_rewarded)-0.5],
        #             interpolation='none',
        #             cmap = self.cmap,
        #             vmin = -vmax,
        #             vmax = vmax  
        #             )

        # show colorbar for ax2 with title "DFF"
        clb = plt.colorbar(clb,ax=ax1,shrink=1).set_label("Ensemble state (DFF units)")
        
        

    def get_rewarded_trials_align_reward_aligned(self):

        #
        fps=30

        #
        buffer = self.buffer #10*30
        window = self.window + buffer

        #
        img_rewarded = np.zeros((len(self.trial_ids_rewarded),window),dtype=np.float32)
        img_non_rewarded = np.zeros((len(self.trial_ids_not_rewarded),window),dtype=np.float32)

        line = np.zeros(window)+np.nan

        plt.figure(figsize=(6,6))
        plt.suptitle("Ensemble state"+
                     "\n"+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(len(self.trial_ids_rewarded))
                    )
        ax1 = plt.subplot(111)
        ax1.set_title("Rewarded trials")
        ax1.set_xlabel("Time (s)")
        ax1.set_ylabel("Trial #")
        # ax2 = plt.subplot(122)
        # ax2.set_title("Non rewarded trials")
        # ax2.set_xlabel("Time (s)")
        # ax2.set_ylabel("Trial #")

        # add the c.ensemble_state triggered on the beggining and ends of each rewarded trial
        for k in range(len(self.trial_ids_rewarded)):
            # get trial id
            i = self.trial_ids_rewarded[k]

            # get times of trial
            idx = np.where(self.trials==i)[0]
            temp = self.ensemble_state[idx[0]:idx[-1]+buffer-1]
            line*=0 + np.nan

            #
            #print (self.window, buffer, line.shape, idx.shape, temp.shape )
            # 900 90 (990,) (60,) (148,)

            line[self.window-idx.shape[0]:self.window] =  temp[:idx.shape[0]]
            line[self.window:self.window+temp[idx.shape[0]:].shape[0]] =  temp[idx.shape[0]:]
            
            img_rewarded[k,:] = line
 
            # plot reward end for each trial as a red dot at 
            if k ==0:
                ax1.plot([self.window/fps, self.window/fps],
                     [k-0.5,k+0.5],c='r', linewidth=3, label='reward time')
            else:
                ax1.plot([self.window/fps, self.window/fps],
                     [k-0.5,k+0.5],c='r', linewidth=3)

            


        # # same for non rewarded trials
        # for k in range(len(self.trial_ids_not_rewarded)):
        #     # get trial id
        #     i = self.trial_ids_not_rewarded[k]

        #     # get times of trial
        #     idx = np.where(self.trials==i)[0]
        #     temp = self.ensemble_state[idx[0]:idx[-1]+buffer-1]
        #     line*=0 + np.nan
        #     line[:temp.shape[0]] =  temp
        #     img_non_rewarded[k,:] = line

        #     # 
        #     # draw a verttical red line of height 1 at idx.shape[0]/fps
        #     ax2.plot([idx.shape[0]/fps, idx.shape[0]/fps],
        #              [k-0.5,k+0.5],c='r', linewidth=3)     
            
            
        # invert the images upside down
        img_rewarded = np.flipud(img_rewarded)
        # img_non_rewarded = np.flipud(img_non_rewarded)
        
       
        vmax = np.median(self.threshold_state)
        clb = ax1.imshow(img_rewarded,aspect='auto',
                extent=[0,window/fps,-0.5,len(self.trial_ids_rewarded)-0.5],
                    interpolation='none',
                    cmap = self.cmap,
                    vmin = -vmax,
                    vmax = vmax  
                    )
        ax1.legend()


        # clb = ax2.imshow(img_non_rewarded,aspect='auto',
        #         extent=[0,window/fps,-0.5,len(self.trial_ids_not_rewarded)-0.5],
        #             interpolation='none',
        #             cmap = self.cmap,
        #             vmin = -vmax,
        #             vmax = vmax  
        #             )

        # show colorbar for ax2 with title "DFF"
        clb = plt.colorbar(clb,ax=ax1,shrink=1).set_label("Ensemble state (DFF units)")
        
        
        

    def get_rewarded_trials(self):

        trials = self.trials.copy()

        # find trial ids
        trial_ids = np.unique(trials)
        trial_ids = trial_ids[trial_ids < 1000]

        # find which trials are rewarded
        trial_ids_rewarded = []
        trial_ids_not_rewarded = []
        for trial_id in trial_ids:
            idx = np.where(trials == trial_id)[0]
            
            trial_reward = self.trials_rewards[idx].mean()
            if trial_reward == 1:
                trial_ids_rewarded.append(trial_id)
            elif trial_reward == 0:
                trial_ids_not_rewarded.append(trial_id)
            else:
                print ("trial id: ", trial_id, " has issue")

        print ("rewarded trials: ", trial_ids_rewarded)
        print ("not rewarded trials: ", trial_ids_not_rewarded)

        # 
        self.trial_ids_rewarded = trial_ids_rewarded
        self.trial_ids_not_rewarded = trial_ids_not_rewarded
    
    #
    def load_results_npz(self):
        
        #
        print ("...loading session: ", 
               self.session_id, 
               self.session_types[self.session_id])
        
        #
        fname = os.path.join(self.root_dir,
                                self.animal_id,
                                str(self.session_ids[self.session_id]),
                                'results.npz')
        #
        results = np.load(fname, allow_pickle=True)

        # load all fields 
        self.ttl_voltages = results['ttl_voltages']
        self.ttl_n_computed = results['ttl_n_computed']
        self.ttl_n_detected = results['ttl_n_detected']
        self.abs_times_ttl_read = results['abs_times_ttl_read']
        self.abs_times_ca_read = results['abs_times_ca_read']
        self.ttl_times = results['ttl_times']
        self.rois_pixels_ensemble1 = results['rois_pixels_ensemble1']
        self.rois_pixels_ensemble2 = results['rois_pixels_ensemble2']
        self.rois_traces_raw_ensemble1 = results['rois_traces_raw_ensemble1']
        self.rois_traces_raw_ensemble2 = results['rois_traces_raw_ensemble2']
        self.rois_traces_smooth1 = results['rois_traces_smooth1']
        self.rois_traces_smooth2 = results['rois_traces_smooth2']
        self.reward_times = results['reward_times']
        self.rewarded_times_abs = results['rewarded_times_abs']
        self.ensemble_activity = results['ensemble_activity']
        self.ensemble_diff_array = results['ensemble_diff_array']
        self.received_reward_lockout = results['received_reward_lockout']
        self.max_reward_window = results['max_reward_window']
        self.missed_reward_lockout = results['missed_reward_lockout']
        self.trials = results['trials']
        self.high_threshold = results['high_threshold']
        self.sampleRate_NI = results['sampleRate_NI']
        self.ttl_pts = results['ttl_pts']
        self.sampleRate_2P = results['sampleRate_2P']
        self.image_width = results['image_width']
        self.image_length = results['image_length']
        self.max_n_seconds_session = results['max_n_seconds_session']
        self.n_frames = results['n_frames']
        self.n_frames_to_be_acquired = results['n_frames_to_be_acquired']
        self.rois_smooth_window = results['rois_smooth_window']
        self.n_ttl_to_start_applying_dynamic_f0 = results['n_ttl_to_start_applying_dynamic_f0']
        self.n_frames_search_forward = results['n_frames_search_forward']
        self.drift_array = results['drift_array']
        self.lick_detector_abstime = results['lick_detector_abstime']
        self.rotary_encoder1_abstime = results['rotary_encoder1_abstime']
        self.rotary_encoder2_abstime = results['rotary_encoder2_abstime']

        #
        self.reward_times = results['reward_times'].T
        self.reward_times = self.reward_times[self.reward_times[:,1]>0]
        self.n_rewards = self.reward_times.shape[0]
        print ("# of rewards: ", self.n_rewards)

    #
    def fix_spreadsheet(self, plotting=False):

        #
        for session_name in self.session_ids[1:]:

            #
            session_name = str(session_name)

            print ("processing session: ", session_name)
        
            # make an fname_out directory if not already present
            self.fname_out = os.path.join(self.root_dir,
                                            self.animal_id,
                                            session_name,
                                            'results_fixed.xlsx')

            # check if fname_out exists
            if os.path.exists(self.fname_out)==False:
                #session_id = self.session_id

                #load the n_rewards
                fname = os.path.join(self.root_dir,
                                        self.animal_id,
                                        session_name,
                                        'results.npz')
                
                #
                try:
                    results = np.load(fname, allow_pickle=True)
                    rewards = results['reward_times'].T
                    idx = np.where(rewards[:,1]>0)[0]

                    #
                    self.n_rewards = idx.shape[0]
                    reward_times = rewards[idx,1]
                except:
                    print ("couldn't load restuls.npz file...")    

                #print ("# of rewards: ", self.n_rewards, " rewar[:100])


                ##############################################################
                ##############################################################
                ##############################################################
                # Load the workbook
                fname_workbook = os.path.join(self.root_dir,
                                                self.animal_id,
                                                session_name,
                                                'results.xlsx')
                
                #
                wb = load_workbook(fname_workbook, read_only=False)

                # Access the worksheet
                ws = wb[wb.sheetnames[0]]  # assuming you want the first sheet, change as needed

                # check the # of columns in ws
                n_cols = ws.max_column

                #
                if n_cols<7:

                     #
                    self.add_missing_columns(wb,
                                             ws,
                                             fname_workbook)

                    # reload the workbook
                    wb = load_workbook(fname_workbook, read_only=False)
                    ws = wb[wb.sheetnames[0]]  # assuming you want the first sheet, change as needed
                
                # save the 4th and 9th columns
                # get white noise state  
                rows = [6,4,5,7,8,9]          # white_noise, tone_state, ensemble_state, reward_state
                self.reward_lockout_counter = []
                self.white_noise_state = []
                self.tone_state = []
                self.ensemble_state = []
                self.reward_state = []
                self.post_reward_state = []
                for row in ws.iter_rows(min_row=2,values_only=False):
                    self.reward_lockout_counter.append(row[rows[0]-1].value)
                    self.white_noise_state.append(row[rows[1]-1].value)
                    self.post_reward_state.append(row[rows[2]-1].value)
                    self.tone_state.append(row[rows[3]-1].value)
                    self.ensemble_state.append(row[rows[4]-1].value)
                    self.reward_state.append(row[rows[5]-1].value)

                # make all lists into arrays
                self.reward_lockout_counter = np.array(self.reward_lockout_counter)
                self.white_noise_state = np.array(self.white_noise_state)
                self.post_reward_state = np.array(self.post_reward_state)
                self.tone_state = np.array(self.tone_state)
                self.ensemble_state = np.array(self.ensemble_state)
                self.reward_state = np.array(self.reward_state)
            
                #
                trials = get_trials2(self.white_noise_state,
                                     self.post_reward_state,
                                     self.reward_lockout_counter)
                
                #
                idx = np.where(trials[:,2]==1)[0]
                
                # check if 2 arrays are identical: reward_times and trials[idx][:,1]
                # first check if reward_times exists as a variable
                if 'reward_times' in locals():
                    if np.array_equal(reward_times, trials[idx][:,1])==False:
                        print ("reward_times and trials[idx][:,1] are not identical")
                        print ("reward_times: ", reward_times)
                        print ("trials[idx][:,1]: ", trials[idx])
                        print ("assigning reward_times to trials[idx][:,1]")
                        trials = np.zeros((self.reward_state.shape[0],2),dtype=np.int32)
                        trials[:,0] = reward_times
                        trials[:,1] = reward_times+60
                    else:
                        print ("reward_times detected correctly")

                # convert trial structure to an array
                trial = np.zeros((self.reward_state.shape[0]), dtype=np.int32)+ 1000
                reward_flag = np.zeros((self.reward_state.shape[0]), dtype=np.int32) + 1000
                ctr=0
                for k in range(len(trials)):
                    trial[trials[k][0]:trials[k][1]]=ctr
                    reward_flag[trials[k][0]:trials[k][1]]=trials[k][2]   # here keep track of each trial whether it was rewarded
                    ctr+=1

                #print ("trial: ", trial[:300])
                trials = trial.copy()
            
                ############################################################
                ################## SAVE UPDATED SPREADSHEET ################
                ############################################################
                # Access the worksheet
                sheet = wb.active

                # Insert the new column. Here, we insert it at column A (index 1). Adjust as necessary.
                n = 10
                sheet.insert_cols(idx=n)
                sheet.cell(row=1, column=n).value = 'trials'
                for k in range(trials.shape[0]):
                    sheet.cell(row=k+2, column=n).value=trials[k]

                # Insert the new column. Here, we insert it at column A (index 1). Adjust as necessary.
                n = 11
                sheet.insert_cols(idx=n)
                sheet.cell(row=1, column=n).value = 'trial_rewarded'
                for k in range(trials.shape[0]):
                    sheet.cell(row=k+2, column=n).value=reward_flag[k]

                # Save the updated workbook
                wb.save(self.fname_out)

                #
                print ('')

            # exit after processing the 1st session for now
        print ("...done...")
            #break

    def add_missing_columns(self,
                            wb,
                            ws,
                            fname_workbook):
        #print (" session has only # cols", n_cols)
        print (" ... adding BLANK tone_state")
        
        #
        ws.insert_cols(idx=7)
        ws.cell(row=1, column=7).value = 'tone_state'
        # insert 0s in all rows of column 7
        for k in range(2,ws.max_row+1):
            ws.cell(row=k, column=7).value = 0
        
        print (" ... adding BLANK ensemble_state")
        # same for 'ensemble_state'
        ws.insert_cols(idx=8)
        ws.cell(row=1, column=8).value = 'ensemble_state'
        # insert 0s in all rows of column 8
        for k in range(2,ws.max_row+1):
            ws.cell(row=k, column=8).value = 0

        # same for 'reward_state'
        print (" ... adding COPY OF post_reward_state as reward_state")
        ws.insert_cols(idx=9)
        ws.cell(row=1, column=9).value = 'reward_state'

        # the fourth column data is the post_reward_state
        # insert 0s in all rows of column 9
        for k in range(2,ws.max_row+1):
            ws.cell(row=k, column=9).value = ws.cell(row=k, column=4).value

        # overwrite the spreadsheet
        wb.save(fname_workbook)

    #
    def binarize_ensemble_state_multi(self):

        #
        for session_id in range(len(self.sessions))[1:]:

            #
            self.sessions[session_id].percentile_threshold = self.percentile_threshold

            #
            #scale_threshold = 0.2
            temp = self.sessions[session_id].e_state.copy()

            # apply low pass filter
            temp = butter_bandpass_filter(temp, 0.01, 1, 30, order=1)

            #
            trace_upphase_bin = self.sessions[session_id].binarize_traces_single_trace(
                                                                temp,
                                                                self.scale_threshold)
            # find temp values below the dff threshold
            idx = np.where(temp<self.dff_threshold)[0]
            trace_upphase_bin[idx] = 0
            
            # this is the value of the binarzied ensemble from the calcium object method above
            self.sessions[self.session_id].trace_upphase_bin = trace_upphase_bin

            #            
            self.e_state_bin = self.sessions[self.session_id].trace_upphase_bin.copy()

            # get locations of positive and negative values of e_state
            pos = np.where(self.e_state >= 0)[0]
            neg = np.where(self.e_state < 0)[0]

            # make an array of -1 and +1 to represent the sign of the e_state
            e_state_sign = np.zeros(self.e_state.shape)
            e_state_sign[pos] = 1
            e_state_sign[neg] = -1

            #
            self.e_state_sign = e_state_sign.copy()

            # now save both the binarized ensemble state and the sign of the ensemble state
            self.sessions[session_id].e_state_bin = self.e_state_bin
            self.sessions[session_id].e_state_sign = self.e_state_sign
    
    #
    def binarize_ensemble_state(self, plotting):

        #
        self.sessions[self.session_id].percentile_threshold = 0.98

        #
        #scale_threshold = 0.2
        self.sessions[self.session_id].binarize_traces_single_trace(self.e_state,
                                                                    self.scale_threshold)

        # get locations of positive and negative values of e_state
        pos = np.where(self.e_state >= 0)[0]
        neg = np.where(self.e_state < 0)[0]

        # make an array of -1 and +1 to represent the sign of the e_state
        e_state_sign = np.zeros(self.e_state.shape)
        e_state_sign[pos] = 1
        e_state_sign[neg] = -1

        #
        self.e_state_bin = self.sessions[self.session_id].trace_upphase_bin

        #
        self.e_state_sign = e_state_sign.copy()

        #
        if plotting:
            plot_e_state_binarization(self.e_state,
                                        self.e_state_bin,
                                        e_state_sign,
                                        self.thresh,
                                        )
    #        
    def process_snippets(self):

        # This function detects the upphase bursts and finds the snippet locations

        # NOTE Starts and ends is for the ensemble state - not for single cells...
        # show extracted snippets
        starts_ends = np.array(self.sessions[self.session_id].starts_ends).squeeze()
        self.starts_ends = starts_ends

        # get the sign of each extracted snippet
        signs = []
        for k in range(starts_ends.shape[0]):
            signs.append(np.median(self.e_state_sign[starts_ends[k,0]:starts_ends[k,1]]))

        signs = np.int32(signs)
        
        # list total number of snippets and how many are positive
        
        #
        self.bursts = np.zeros(self.sessions[self.session_id].trace_upphase_bin.shape[0], dtype=np.int32)+1000
        self.bursts_to_threshold = np.zeros(self.sessions[self.session_id].trace_upphase_bin.shape[0], dtype=np.int32)+1000
        self.bursts_rewarded = np.zeros(self.sessions[self.session_id].trace_upphase_bin.shape[0], dtype=np.int32)+1000
        
        #
        print ("# of snippets: ", signs.shape[0], " # of detected snippets of sign: ", self.burst_sign, " ", np.where(signs==self.burst_sign)[0].shape[0])

        ctr_total = 0
        ctr_rewarded = 0
        reward_buffer = 15 # no. of frames to search forward for a reward
        fps = 30
        ids_snippets_rewarded = []
        ids_snippets = []
        self.starts_ends_to_threshold = []
        for k in range(starts_ends.shape[0]):

            if signs[k] == self.burst_sign:
                
                #
                ids_snippets.append(k)

                # extract indexes
                temp_starts_ends = starts_ends[k]
                
                # save all burst times
                self.bursts[temp_starts_ends[0]: temp_starts_ends[1]] = ctr_total

                # check if water reard was given during this period
                water_state = self.water_reward[temp_starts_ends[0]:temp_starts_ends[1]+reward_buffer]
                diff = water_state[1:]-water_state[:-1]
                idx = np.where(diff==1)[0]

                # check if the water state changed from a 0 to a 1 during this period
                if idx.shape[0]>0:
                    
                    # add a reward
                    self.bursts_rewarded[temp_starts_ends[0]:temp_starts_ends[1]]= 1

                    # also save the burst to threshold
                    self.bursts_to_threshold[temp_starts_ends[0]:temp_starts_ends[0]+idx[0]]= 1

                    #
                    self.starts_ends_to_threshold.append([temp_starts_ends[0],temp_starts_ends[0]+idx[0]])

                    # save rewarded burst times
                    ctr_rewarded+=1

                    #
                    ids_snippets_rewarded.append(k)
                else:
                    self.bursts_rewarded[temp_starts_ends[0]: temp_starts_ends[1]]= 0

                #
                ctr_total+=1
        
        #
        self.signs = signs
        self.ids_snippets_rewarded = ids_snippets_rewarded
        self.ids_snippets = ids_snippets

        # print how many of each burst types there are
        print ("# of bursts: ", ctr_total, " # of rewarded bursts: ", ctr_rewarded)
        print ("self.update_spreadsheet: ", self.update_spreadsheet)

        if self.update_spreadsheet==True:
            # read the results_fixed.xlsx file 
            fname = os.path.join(self.root_dir,
                                    self.animal_id,
                                    self.session_ids[self.session_id],
                                    'results_fixed.xlsx')
            #print ("updatinging spreadsheet: ", fname)
            df = pd.read_excel(fname)

            # check to see if any columns named "burst" exists and overwrite it
            if 'burst' in df.columns:
                if self.overwrite_spreadsheet:
                    df['burst'] = self.bursts
            else:
                df['burst'] = self.bursts

            # same for bursts_rewarded
            if 'burst_rewarded' in df.columns:
                if self.overwrite_spreadsheet:
                    df['burst_rewarded'] = self.bursts_rewarded
            else:
                df['burst_rewarded'] = self.bursts_rewarded

            # same for bursts_to_threshold
            if 'burst_to_threshold' in df.columns:
                if self.overwrite_spreadsheet:
                    df['burst_to_threshold'] = self.bursts_to_threshold
            else:
                df['burst_to_threshold'] = self.bursts_to_threshold
                
            # save the updated spreadsheet
            df.to_excel(fname, index=False)

    #
    def fix_missing_vals(self, array_in):

        # we repeat the last value for each of the arrays
        temp = np.zeros(90000)
        temp[:array_in.shape[0]] = array_in
        temp[array_in.shape[0]:] = array_in[-1]
        
        return temp

    # 
    def extend_spreadsheet(self, df):
        
        #
        if self.thresh.shape[0] < 90000:
                    
            #
            df = df.reindex(range(90000))

            # 1. Identify the last non-NaN entry in each column
            last_entries = df.apply(lambda col: col[col.last_valid_index()])
            print ("... repeating last_entries: ", last_entries)

            # 2. Fill the NaN values in the extended rows with the last entry
            for column in df.columns:
                df[column].fillna(last_entries[column], inplace=True)

            # get the first column by number (as it doesn't have a name) and set it to go from 1 to 90000
            df.iloc[:,0] = np.arange(0,90000)

            # fix the "n_ttl" colun to go from 1 to 90000
            df['n_ttl'] = np.arange(1,90001)

            self.thresh = self.fix_missing_vals(self.thresh)
            self.white_noise_state = self.fix_missing_vals(self.white_noise_state)
            self.tone_state = self.fix_missing_vals(self.tone_state)
            self.trials = self.fix_missing_vals(self.trials)
            self.water_reward = self.fix_missing_vals(self.water_reward)
            self.trials_rewards = self.fix_missing_vals(self.trials_rewards)
            self.e_state = self.fix_missing_vals(self.e_state)

            # save the updated spreadsheet
            print ("DF: ", df.shape)
            df.to_excel(os.path.join(
                        self.root_dir,
                        self.animal_id,
                        str(self.session_ids[self.session_id]),
                        'results_fixed.xlsx'), 
                    index=False)

            # reload the spreadsheet
            df = pd.read_excel(os.path.join(
                        self.root_dir,
                        self.animal_id,
                        str(self.session_ids[self.session_id]),
                        'results_fixed.xlsx'))

        return df
    
    #
    def load_spreadsheet_entries(self, df):
         #
        self.thresh = df.loc[:, 'current_high_threshold'].values

        # same for white_noise_state
        self.white_noise_state = df.loc[:, 'white_noise_state'].values

        # same for tone_state
        self.tone_state = df.loc[:, 'tone_state'].values

        # same for trials
        self.trials = df.loc[:, 'trials'].values

        #
        try:
            self.water_reward = df.loc[:, 'water_reward'].values
        except:
            print( "Missing water_reward column in spreadsheet... using 'reward_state ' instead ")
            self.water_reward = df.loc[:, 'reward_state'].values


        # same for trial_rewarded
        self.trials_rewards = df.loc[:, 'trial_rewarded'].values

        # same for a columan named "ensemble_state"
        self.e_state = df.loc[:, 'ensemble_state'].values

        # delete first few frames because of some bug in the initilization of the ensemble state
        self.e_state[:100] = 0 

    #   
    def load_spreadsheet_multi(self):

        # make lists for all the data loaded below
        self.thresh_all = []
        self.white_noise_state_all = []
        self.tone_state_all = []
        self.trials_all = []
        self.water_rewards_all = []
        self.trials_rewards_all = []
        self.e_state_all = []

        # load data from spreasheets
        for session_id in trange(len(self.sessions)):
            #
            self.session_id = session_id
            
            # 
            if self.session_types[session_id] == 'day0':
                continue

            #
            fname_pkl = os.path.join(
                                    self.root_dir,
                                    self.animal_id,
                                    str(self.session_ids[session_id]),
                                    'results_fixed.pkl')

            #            
            if os.path.exists(fname_pkl):
                with open(fname_pkl, 'rb') as f:
                    df = pickle.load(f)
            else:
                df = pd.read_excel(os.path.join(
                                    self.root_dir,
                                    self.animal_id,
                                    str(self.session_ids[session_id]),
                                    'results_fixed.xlsx'))
                
            #
            self.load_spreadsheet_entries(df)

            # this function adds rows to the df and updates all other vals, and resaves df to disk
            if self.fix_spreadsheet_missing_vals:
                df = self.extend_spreadsheet(df)      

            # save the dataframe as a pickle file
            with open(fname_pkl, 'wb') as f:
                pickle.dump(df, f)         
            
            # append the loaded data to some lists somewhere
            self.append_data_to_session()

    #               
    def append_data_to_session(self):

        self.sessions[self.session_id].thresh = self.thresh
        self.sessions[self.session_id].white_noise_state = self.white_noise_state
        self.sessions[self.session_id].tone_state = self.tone_state
        self.sessions[self.session_id].trials = self.trials
        self.sessions[self.session_id].water_reward = self.water_reward
        self.sessions[self.session_id].trials_rewards = self.trials_rewards
        self.sessions[self.session_id].e_state = self.e_state

    # #   
    # def load_spreadsheet(self):

    #     # load data from spreassheetds
    #     df = pd.read_excel(os.path.join(
    #                         self.root_dir,
    #                         self.animal_id,
    #                         self.session_ids[self.session_id],
    #                         'results_fixed.xlsx'))

    #     # find a column in the dataframe with the name "current_high_threshold"
    #     # and get its values as a numpy array
    #     self.thresh = df.loc[:, 'current_high_threshold'].values

    #     if self.fix_spreadsheet_missing_vals:
    #         if self.thresh.shape[0] < 90000:
                
    #             # 
    #             print ("Found short spreadsheet: ", self.thresh.shape[0], " entries... fixing it...")
    #             #print ("Filename is: "

    #             #
    #             df = df.reindex(range(90000))

    #             # 1. Identify the last non-NaN entry in each column
    #             last_entries = df.apply(lambda col: col[col.last_valid_index()])
    #             print ("... repeating last_entries: ", last_entries)
    #             # 2. Fill the NaN values in the extended rows with the last entry
    #             for column in df.columns:
    #                 df[column].fillna(last_entries[column], inplace=True)

    #             # get the first column by number (as it doesn't have a name) and set it to go from 1 to 90000
    #             df.iloc[:,0] = np.arange(0,90000)

    #             # fix the "n_ttl" colun to go from 1 to 90000
    #             df['n_ttl'] = np.arange(1,90001)

    #             print ("DF: ", df.shape)
    #             df.to_excel(os.path.join(
    #                         self.root_dir,
    #                         self.animal_id,
    #                         self.session_ids[self.session_id],
    #                         'results_fixed.xlsx'), 
    #                     index=False)

    #             # reload the spreadsheet
    #             df = pd.read_excel(os.path.join(
    #                         self.root_dir,
    #                         self.animal_id,
    #                         self.session_ids[self.session_id],
    #                         'results_fixed.xlsx'))

    #             # and grab the first value again                
    #             self.thresh = df.loc[:, 'current_high_threshold'].values

    #     # same for white_noise_state
    #     self.white_noise_state = df.loc[:, 'white_noise_state'].values

    #     # same for tone_state
    #     self.tone_state = df.loc[:, 'tone_state'].values

    #     # same for trials
    #     self.trials = df.loc[:, 'trials'].values

    #     #
    #     self.water_reward = df.loc[:, 'water_reward'].values

    #     # same for trial_rewarded
    #     self.trials_rewards = df.loc[:, 'trial_rewarded'].values

    #     # same for a columan named "ensemble_state"
    #     e_state = df.loc[:, 'ensemble_state'].values

    #     # delete first few frames
    #     e_state[:100] = 0 

    #     # filter using savgol_filter
    #     self.e_state = savgol_filter(e_state, 31, 3)

    #     # check to see there are <90,000 entries
    #     if self.fix_spreadsheet_missing_vals:
    #         if self.thresh.shape[0] < 90000:

    #             self.thresh = self.fix_missing_vals(self.thresh)
    #             self.white_noise_state = self.fix_missing_vals(self.white_noise_state)
    #             self.tone_state = self.fix_missing_vals(self.tone_state)
    #             self.trials = self.fix_missing_vals(self.trials)
    #             self.water_reward = self.fix_missing_vals(self.water_reward)
    #             self.trials_rewards = self.fix_missing_vals(self.trials_rewards)
    #             self.e_state = self.fix_missing_vals(self.e_state)





    #    
    def get_tones(self, plotting=False):


        # make an fname_out directory if not already present
        self.fname_out = os.path.join(self.root_dir,
                                        self.animal_id,
                                        self.session_ids[self.session_id],
                                        'tone_data.npz')

        # check if fname_out exists
        if os.path.exists(self.fname_out)==False or self.recompute_tones:
            session_id = self.session_id

            # load the n_rewards
            fname = os.path.join(self.root_dir,
                                    self.animal_id,
                                    self.session_ids[session_id],
                                    'results.npz')
            results = np.load(fname, allow_pickle=True)
            rewards = results['reward_times'].T
            idx = np.where(rewards[:,1]>0)[0]
            self.n_rewards = idx.shape[0]
            self.reward_times = rewards[idx,1]
            #print ("n_rewards: ", self.n_rewards)

            # Load the workbook
            wb = load_workbook(os.path.join(self.root_dir,
                                            self.animal_id,
                                            self.session_ids[session_id],
                                        # 'data',
                                            'results.xlsx'), read_only=True)


            # Access the worksheet
            ws = wb[wb.sheetnames[0]]  # assuming you want the first sheet, change as needed

            # Get the n-th column
            n = 7  # change this to your desired column (Python uses 0-based indexing)

            # Create a numpy array from the n-th column, skipping the header row
            self.tones = np.array([row[n-1].value for row in ws.iter_rows(min_row=2, values_only=False)])

            #
            self.tones = np.int32(self.tones)

            #
            n = 8
            self.ensemble_state = np.array([row[n-1].value for row in ws.iter_rows(min_row=2, values_only=False)])

            # save fname_out
            np.savez(self.fname_out, 
                     tones = self.tones,
                     reward_times = self.reward_times,
                     n_rewards = self.n_rewards,
                     ensemble_state = self.ensemble_state)


            # plot the mode as a reference
            if plotting:
                plt.figure()
                # find the most frequent value in tone
                from scipy.stats import mode
                mode_tone = mode(self.tones)

                plt.plot(self.tones)
                plt.plot(np.ones(self.tones.shape)*mode_tone[0],'r--')

                #
                plt.xlabel("Frames")
                plt.ylabel("Tone")

                #
                plt.suptitle(c.animal_id+ " - " + c.session_ids[1])

                #
                plt.show()
        else:
            #print ("loading tones from: ", self.fname_out)
            data = np.load(self.fname_out, allow_pickle=True)
            self.tones = data['tones']
            self.n_rewards = data['n_rewards']
            self.reward_times = data['reward_times']

    # load single data [ca] data
    def load_single_data(self):

        #          
        data_dir = os.path.join(
                        self.root_dir,
                        self.animal_id,
                        str(self.session_ids[self.session_id]),
                        'plane0'
                        )

        C = binca.Binarize(self.root_dir,
                           self.animal_id,
                           str(self.session_id)
                           )
        C.data_dir = data_dir

        #
        C.load_suite2p()

        #
        C.load_footprints()

        # load binarization
        C.detrend_model_type = 'polynomial'
        C.detrend_model_order = 2
        C.percentile_threshold = 0.999999
        
        # 
        print ("loading binarization")
        C.load_binarization()
        
        return C

    #
    def compute_bmi_to_suite2p_matches(self):

        # load the ROIs from the session and match them to the ROIS from suite2p loaded data
        session_ids = np.arange(1,len(self.session_types),1)

        #
        for session_id in session_ids:
            # load the suite2p ROIS as F_filtered
            try:
                F_filtered = self.sessions[session_id].F_detrended
            except:
                F_filtered = self.F_detrended

            # 
            # load the BMI live ROIS from the results.npz file
            if self.session_types[session_id]=='day0':
                fname_rois = os.path.join(self.root_dir,
                                        self.animal_id,
                                        str(self.session_ids[session_id]),
                                        'rois_pixels_and_thresholds_day0.npz')
                d = np.load(fname_rois, allow_pickle=True)
                cell_ids = d['cell_ids']

                # grab the [ca] data from the suite2p data
                # TODO

            # we load 
            else:
                #fname = c.sessions[session_id].session_dir + '/plane0/results.npz'

                # grab the .npz results file
                fname_results = os.path.join(self.root_dir,
                                            self.animal_id,
                                            str(self.session_ids[session_id]),
                                            'data',
                                            'results.npz')
                
                #
                # save the best match ids 
                dir_ensembles = os.path.join(self.root_dir,
                                            self.animal_id,
                                            str(self.session_ids[session_id]),
                                            'ensembles')
                
                if os.path.exists(dir_ensembles)==False:
                    os.mkdir(dir_ensembles)

                fname_match = os.path.join(dir_ensembles, 
                                        'ensembles_matches_bmi_to_suite2p.npz')
                
                if os.path.exists(fname_match)==True:
                    print (".... already computed...")
                    return

                
                # load the results file
                d = np.load(fname_results, allow_pickle=True)

                #
                rois1 = d['rois_traces_raw_ensemble1']
                rois2 = d['rois_traces_raw_ensemble2']
                # print ("rois1: ", rois1.shape)
                # print ("rois2: ", rois2.shape)

                # compute pearson correlation between each roi in the two ensembles and all the F_filtered rois
                #   and find the best match
                corr1 = np.zeros((2,F_filtered.shape[0]))
                corr2 = np.zeros((2,F_filtered.shape[0]))


                
                #
                corrs = parmap.map(compute_corr2, 
                                    [rois1[0], rois1[1], rois2[0], rois2[1]], 
                                    F_filtered, 
                                    pm_processes=4, 
                                    pm_pbar=True)

                ##########################################################
                ##########################################################
                ##########################################################


                
                rois = [rois1[0], rois1[1], rois2[0], rois2[1]]

                #
                np.savez(fname_match,
                        idx_text = "These are the cells in the suite2p data that match the BMI live ROIs.",
                        idx_ensemble1_0 = np.argmax(corrs[0]),
                        idx_ensemble1_1 = np.argmax(corrs[1]),
                        idx_ensemble2_0 = np.argmax(corrs[2]),
                        idx_ensemble2_1 = np.argmax(corrs[3]),
                        
                        # save the raw 
                        ensemble_bmi_text = "These are the raw ROIs from the BMI live data.",
                        ensemble1_0_bmi = rois[0],
                        ensemble1_1_bmi = rois[1],
                        ensemble2_0_bmi = rois[2],
                        ensemble2_1_bmi = rois[3],

                        # save the matching suite2p data
                        ensemble_suite2p_text = "These are the matching ROIs from the suite2p data.",
                        ensemble1_0_suite2p = F_filtered[np.argmax(corrs[0])],
                        ensemble1_1_suite2p = F_filtered[np.argmax(corrs[1])],
                        ensemble2_0_suite2p = F_filtered[np.argmax(corrs[2])],
                        ensemble2_1_suite2p = F_filtered[np.argmax(corrs[3])],
                        )
                
                # autosave the matches to disk also
                for k in range(4):
                    from matplotlib.figure import Figure
                    from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvasAgg

                    # Create a figure
                    fig = plt.figure(figsize=(16,4))

                    # Associate the figure with the 'Agg' backend by creating a FigureCanvasAgg instance
                    #canvas = FigureCanvasAgg(fig)

                    # Now you can add axes, plot data, and set labels and titles as usual
                    ax = plt.subplot(111)  # Add an axes to the figure
                    temp = rois[k]
                    f0 = np.median(temp)
                    temp = (temp-f0)/f0

                    # filter temp using box filter
                    # Define the box filter kernel (window)
                    window_size = 151  # Size of the box filter window
                    box_filter = np.ones(window_size) / window_size

                    # Apply the box filter using numpy's convolution function
                    temp = np.convolve(temp, box_filter, mode='same')
                    ax.plot(temp, label='ensemble roi: '+str(k),
                             alpha=.8)

                    temp = np.convolve(F_filtered[np.argmax(corrs[k])], box_filter, mode='same')
                    ax.plot(temp, 
                            label='matching suite2p cell: '+str(np.argmax(corrs[k])),
                            alpha=.8,
                            linewidth=1)
                    
                    #
                    ax.set_xlabel("Frames")
                    ax.legend()
                    plt.savefig(fname_match[:-4]+str(k)+'.png', dpi=300)
                    
                    #
                    plt.close()

    #
    def visualize_ensemble_to_reward_average(self):


        roi_names = ['pos1','pos2','neg1','neg2']

        # make viridis colormap for len(c.session_ids) sessions
        cmap = matplotlib.cm.get_cmap('viridis')
        colors = cmap(np.linspace(0.1, .9, len(self.session_ids)))

        #
        plt.figure(figsize=(10,10))
        for session_id in np.arange(1,len(self.session_ids),1):

            #
            psths_avg, psths_shuffled_avg, n_bursts  = get_reward_triggered_bmi_ensembles(session_id, 
                                                                                        self,
                                                                                        self.window)
            
            # print ("psths_avg: ", psths_avg.shape)
            # print ("psths_shuffled_avg: ", psths_shuffled_avg.shape)

            for k in range(4):
                ax = plt.subplot(2,2,k+1)

                temp = np.mean(psths_avg[k],0)*100

                #
                t = np.arange(-self.window,self.window,1)/30

                plt.plot(t,temp,
                        c=colors[session_id],
                        label = self.session_types[session_id] if k==0 else '')
                
                plt.title(roi_names[k])

                # plot vertical line at t=0
                plt.plot([0,0],[0,100],'--',c='red')

                # plot horizontal line at y = 0
                plt.plot([-self.window/30.,
                        self.window/30.],[0,0],'--',
                        c='black')

                #
                plt.xlabel('Time (sec)')
                if self.use_DFF:
                    plt.ylabel("Mean DFF (%)")
                else:
                    plt.ylabel("Mean burst (upphase)")

                #
                plt.xlim(-self.window/30.,self.window/30.)

                #
                if k==0:
                    plt.legend()

        #  
        plt.suptitle(self.animal_id+ " " + str(self.rec_type))


        # 
        plt.show()
        
    #
    def load_sessions(self):

        #for animal_id in self.animal_ids:
        self.sessions = []
        self.session_types = []
        ctrs = 0
        self.reward_times = []
        for session_ in tqdm(self.session_ids):

            ################################################
            ################################################
            ################################################
            fname_yaml = os.path.join(self.root_dir,
                                        self.animal_id,
                                        str(session_),
                                        str(session_)+'.yaml')
            with open(fname_yaml) as file:
                doc = yaml.load(file, Loader=yaml.FullLoader)

            session_type = doc['session_type']
            self.session_types.append(session_type)

            # check if there's a flag to use non-merged binarized data 
            #self.use_non_merged = self.use_non_merged_rois

            # load rewards
            if ctrs>0:
                try:
                    _, _, reward_times = load_results_npz_standalone(self.root_dir,
                                                            self.animal_id,
                                                            ctrs,
                                                            self.session_ids)
                    reward_times = reward_times[:,1]
                except:
                    print ("couldn't reward results.npz file...")
                    reward_times = []

            else:
                reward_times = []

            self.reward_times.append(reward_times)

            ################################################
            ################################################
            ################################################
            #
            if self.use_non_merged:
                print (" ***** USING non-merged binarization")
                data_dir = os.path.join(
                                        self.root_dir,
                                        self.animal_id,
                                        str(session_),
                                        'plane0',
                                        )

            else:
                data_dir = os.path.join(
                                        self.root_dir,
                                        self.animal_id,
                                        str(session_),
                                        'plane0',
                                        'merged'
                                        )
            
            #
            if os.path.exists(data_dir)==False:
                print ("couldn't find binarization...")
                return

            #
            C = binca.Binarize(root_dir = self.root_dir,
                               animal_id = self.animal_id,
                               session_name = str(session_),
                                    )
            
            C.data_dir = data_dir
            C.data_type = '2p'
            C.set_default_parameters_2p()

            # #
            # paramers for binarization
            C.dff_min = self.dff_min                  # min %DFF for [ca] burst to considered a spike (default 5%) overwrites percentile threshold parameter
            C.percentile_threshold = self.percentile_threshold   # this is pretty fixed, we don't change it; we want [ca] bursts that are well outside the "physics-caused"noise
            C.maximum_std_of_signal = self.maximum_std_of_signal     # if std of signal is greater than this, then we have a noisy signal and we don't want to binarize it

            #
            C.save_figures = self.save_figures

            # overwrite some of the defaults here.
            C.remove_bad_cells = self.remove_bad_cells
            C.recompute_binarization = self.recompute_binarization
            
            # 
            C.load_binarization()            

            #
            self.sessions.append(C)

            #
            ctrs+=1
            print ('')

    #
    def plot_ROIs_contours(self):
        clrs = ['blue','lightblue','red','pink']
        centres = []
        for k in range(len(self.contours_ROIs)):
            temp2 = self.contours_ROIs[k][0]
            temp= temp2.copy()
            temp[:,0] = temp2[:,1]
            temp[:,1] = temp2[:,0]
            #print (temp.shape)
            plt.plot(temp[:,0],
                    temp[:,1],
                    c=clrs[k],
                    alpha=.5,
                    linewidth=6,
                    label="ROI: "+str(k))
            centre = np.mean(temp,0)
            centres.append(centre)

        # make vstack array
        centres = np.vstack(centres)

        return centres


    #
    def plot_matching_contours(self):
        
        #
        plt.figure()
        self.session_contours = []
        for k in range(len(self.sessions)):

            #
            shift = self.shifts[k]

            #
            plt.subplot(2,5,k+1)

            #
            self.session_contours.append([])

            #
            session_contour = self.sessions[k].contours

            #
            centres = self.plot_ROIs_contours()

            # loop through each session and plot contours individually
            found=False
            for p in range(len(session_contour)):
                
                #
                temp = session_contour[p]
                centre = np.median(temp,0)

                # check to see if temp centre is close to any of the centres
                dist = np.linalg.norm(centres-centre, axis=1)
                idx = np.argmin(dist)

                if dist[idx]<self.contour_ROI_max_dist:
                #print ("session: ", k, "  cell: ", p, "  closest cell: ", idx, "  dist: ", dist[idx])
                    self.session_contours[k].append([idx,p])
                    if found==False:
                        plt.plot(temp[:,0]+shift[0],
                            temp[:,1]+shift[1],
                            #c=colors[k],
                            c='black',
                            alpha=1,
                            linewidth=2,
                            label='Suite2p nearest cell'
                            )
                        found=True
                    else:
                        plt.plot(temp[:,0]+shift[0],
                            temp[:,1]+  shift[1],
                            #c=colors[k],
                            c='black',
                            alpha=1,
                            linewidth=2,
                            )
            plt.xlim(0,512)
            plt.ylim(0,512)

            if k==0:
                plt.legend()
            plt.title(self.session_ids[k])

            # 
            temp = np.vstack(self.session_contours[k])
            # sort by first column
            temp = temp[temp[:,0].argsort()]
            self.session_contours[k] = temp



        plt.suptitle(self.animal_id)
        plt.show()

#
def get_reward_centered_traces(session_id,
                                root_dir,
                                animal_id,
                                reward_window,
                                filter = True,
                                recompute=True):

    if session_id=='day0':
        return

    #   
    fname_out = os.path.join(root_dir,
                        animal_id,
                        session_id,
                        'reward_centered_traces_rewardwindow_'+str(reward_window)+'.npy')
    
    if os.path.exists(fname_out)==False or recompute==True:

        # load reward times            
        fname = os.path.join(root_dir,
                            animal_id,
                            session_id,
                            'results.npz')
        
        #
        data = np.load(fname, allow_pickle=True)
        rewards = data['reward_times'].T[:,::-1][:,0]
        idx = np.where(rewards>0)[0]
        reward_switches = rewards[idx]

        # load calcium traces raw from suite2p
        if False: 
            fname = os.path.join(root_dir,
                                animal_id,
                                session_id,
                                'plane0',
                                'F.npy')

            F = np.load(fname, allow_pickle=True)

            #
            # filter the trace
            if filter:
                F = scipy.signal.savgol_filter(F, 31, 3)

            ########################################################

            # get f0 values first as the median of axis 1
            #print ("F shape: ", F.shape)
            f0s = np.nanmedian(F, axis=1)

            #
            F = (F-f0s[:,np.newaxis])/f0s[:,np.newaxis]

        # use the F_upphase instead
        else:
            fname = os.path.join(root_dir,
                    animal_id,
                    session_id,
                    'plane0',
                    'binarized_traces.npz')
            d = np.load(fname, allow_pickle=True)
                        
            F = d["F_upphase"]
            F = d["F_onphase"]
            F = d["F_detrended"]
            F = d["F_filtered"]

            #


        # split c.sessions[0].F around the reward_witches time with a window of 10 seconds
        temp = []
        temp_shuffled = []
        for k in range(reward_switches.shape[0]):
            
            #
            temp1 = F[:,reward_switches[k]-reward_window:reward_switches[k]+reward_window]
            
            # make temp2 with the index shuffled
            # make a random value from 500 to -500 in F
            idx = np.random.randint(500, F.shape[1]-500)
            temp2 = F[:,reward_switches[k]-reward_window+idx:reward_switches[k]+reward_window+idx]

            # check if the window is correct 
            if temp1.shape[1]==reward_window*2:
                temp.append(temp1)

            # check if the window is correct 
            if temp2.shape[1]==reward_window*2:
                temp_shuffled.append(temp2)


        #
        reward_centered_traces = np.array(temp)
        reward_centered_traces_shuffled = np.array(temp_shuffled)

        #
        print ("# of cells, times: ", F.shape, ", output: ", reward_centered_traces.shape)

        # save the traces
        np.save(fname_out, reward_centered_traces)
        np.save(fname_out[:-4]+'_shuffled.npy', reward_centered_traces_shuffled)


def get_img_ca(cell_id, 
            F_filtered,
            window,
            tone_starts,
            shuffle,
            smooth):

    #
    img=[]
    #img_shuffled=[]
    for starts in tone_starts:
        traces = [] 
        for s in starts:
            #temp = c.sessions[session_id].F_upphase_bin[cell_id,s-window:s+window]
            if shuffle==False:
                temp = F_filtered[cell_id,s-window:s+window]
                
            # same but shuffle the tone times
            else:
                s_shuffled = np.random.randint(2*window,F_filtered.shape[1]-2*window)
                temp = F_filtered[cell_id,s_shuffled-window:s_shuffled+window]

            if temp.shape[0]!=2*window:
                continue
            #
            if smooth:
                temp = scipy.signal.savgol_filter(temp, 31, 3)

            traces.append(temp)

        # plot the average of traces
        if len(traces)==0:
            img.append(np.zeros((2*window)))
        
        else:
            img.append(np.mean(traces,axis=0))

    #
    return np.array(img)

def get_img_tone(
                cell_id,
                F_filtered,
                window,
                tones,
                shuffle,
                smooth,
                tone_type,
                remove_base_tone
                ):


    # find 1s in the [ca] vector
    idx_starts_only = np.where(F_filtered[cell_id]==1)[0]

    # find the beginning and ends of idx_starts
    from scipy.signal import chirp, find_peaks, peak_widths
    peaks, _ = find_peaks(F_filtered[cell_id])  # middle of the pluse/peak
    widths, heights, starts, ends = peak_widths(F_filtered[cell_id], peaks)
    starts = np.int32(starts)
    
    if tone_type=='starts':
        idx_starts = starts
    elif tone_type=='all':
        idx_starts = idx_starts_only

    #
    unique_tones, counts = np.unique(tones, return_counts=True)
    #print ("unique_tones: ", unique_tones)
    #
    traces = []
    for ctr, s in enumerate(idx_starts):
    
        #temp = c.sessions[session_id].F_upphase_bin[cell_id,s-window:s+window]
        if shuffle==False:
            temp = tones[s-window:s+window]
            
        # same but shuffle the tone times
        else:
            s_shuffled = np.random.randint(2*window,F_filtered.shape[1]-2*window)
            temp = tones[s_shuffled-window:s_shuffled+window]

        if temp.shape[0]!=2*window:
            continue
        #
        if smooth:
            temp = scipy.signal.savgol_filter(temp, 31, 3)

        # remove 100 vals
        temp = np.float32(temp)
        idx = np.where(temp==100)[0]
        temp[idx]=np.nan

        # same for 0 values
        idx = np.where(temp==0)[0]
        temp[idx]=np.nan

        traces.append(temp)

        #
        #if ctr%1000==0:
        #    print ("cell_id: ", cell_id, "  s: ", s, "  temp: ", temp.shape)
        #    print (temp)

    # plot the average of traces
    if len(traces)==0:
        img = np.zeros((unique_tones.shape[0], 2*window)).T
    
    else:
        #img = np.nanmean(traces,axis=0)

        # make a histogram for every time point in the window
        img=[]
        for k in range(len(traces)):
            temp = traces[k]

            if remove_base_tone:
                idx = np.where(temp==4756)[0]
                temp = np.delete(temp, idx)

            #
            y = np.histogram(temp, bins=unique_tones)

            #
            img.append(y[0])

        img = np.vstack(img).T

        
    #
    return img

  

def plot_e_state_binarization(e_state,
                            e_state_bin,
                            e_state_sign,
                            thresh):
#
    t = np.arange(0, 90000, 1)/30.

    # plot ensemble state graph
    plt.figure(figsize=(10, 10))

    plt.plot(t,e_state, label = 'ensemble state')
    plt.plot(t,e_state_bin*e_state_sign, label = 'ensemble state bin')

    # plot horizontal line from 0 to 90000 at thresh
    plt.plot(t, thresh, 'k-', lw=2, label = 'thresh')

    #
    plt.xlabel('time (s)')

    # set xlims
    plt.xlim(t[0], t[-1])

    plt.legend()
    plt.show()

    #
def resample_snippet(snippet):

    #
    time_original = np.arange(snippet.shape[0])
    data_original = snippet

    #
    f = scipy.interpolate.interp1d(time_original, data_original, kind='linear')

    # Use this function to obtain new data samples
    time_new = np.linspace(0, time_original[-1], 10000)
    #idx = np.where(time_new < time_original[-1])[0]
    #time_new = time_new[idx]

    data_new = f(time_new)
    #print (data_new.shape)
    
    # take a total of 100 samples from the snippet
    idx = np.linspace(0, data_new.shape[0]-1, 30).astype(int)
    snippet_downsampled = data_new[idx]

    return snippet_downsampled



def warp_ts1_to_ts2(ts1, ts2):
    """
    Warps ts1 to best match ts2 using DTW.
    
    Args:
    - ts1 (array-like): First time series
    - ts2 (array-like): Second time series
    
    Returns:
    - numpy.ndarray: Warped version of ts1
    """

    #
    #print (ts1)
    #print (ts2)

    #
    alignment = dtw(ts1, ts2, keep_internals=True)
    
    # Re-sample ts1 based on the warping path's indices for ts1
    #basic_warped_ts1 = ts1[alignment.index1]

    #
    return alignment.index1

def interpolate_time_warp(ts1, ts2, warp_function):

    ts1 = ts1[warp_function]
    
    # Interpolate the basic_warped_ts1 to match the length of ts2
    x_new = np.linspace(0, len(ts1) - 1, len(ts2))
    f = interp1d(np.arange(len(ts1)), ts1, kind='linear')
    resampled_warped_ts1 = f(x_new)

    return resampled_warped_ts1

#

def load_results_npz_standalone(root_dir,
                                animal_id,
                                session_id,
                                session_ids,
                                ):
    
    #
    fname = os.path.join(root_dir,
                        animal_id,
                        str(session_ids[session_id]),
                        'data',
                        'results.npz')
    #
    results = np.load(fname, allow_pickle=True)

    # load all fields 
    # self.ttl_voltages = results['ttl_voltages']
    # self.ttl_n_computed = results['ttl_n_computed']
    # self.ttl_n_detected = results['ttl_n_detected']
    # self.abs_times_ttl_read = results['abs_times_ttl_read']
    # self.abs_times_ca_read = results['abs_times_ca_read']
    # self.ttl_times = results['ttl_times']
    # self.rois_pixels_ensemble1 = results['rois_pixels_ensemble1']
    # self.rois_pixels_ensemble2 = results['rois_pixels_ensemble2']
    # self.rois_traces_raw_ensemble1 = results['rois_traces_raw_ensemble1']
    # self.rois_traces_raw_ensemble2 = results['rois_traces_raw_ensemble2']
    rois_traces_smooth1 = results['rois_traces_smooth1']
    rois_traces_smooth2 = results['rois_traces_smooth2']
    reward_times = results['reward_times']
    # self.rewarded_times_abs = results['rewarded_times_abs']
    # self.ensemble_activity = results['ensemble_activity']
    # self.ensemble_diff_array = results['ensemble_diff_array']
    # self.received_reward_lockout = results['received_reward_lockout']
    # self.max_reward_window = results['max_reward_window']
    # self.missed_reward_lockout = results['missed_reward_lockout']
    # self.trials = results['trials']
    # self.high_threshold = results['high_threshold']
    # self.sampleRate_NI = results['sampleRate_NI']
    # self.ttl_pts = results['ttl_pts']
    # self.sampleRate_2P = results['sampleRate_2P']
    # self.image_width = results['image_width']
    # self.image_length = results['image_length']
    # self.max_n_seconds_session = results['max_n_seconds_session']
    # self.n_frames = results['n_frames']
    # self.n_frames_to_be_acquired = results['n_frames_to_be_acquired']
    # self.rois_smooth_window = results['rois_smooth_window']
    # self.n_ttl_to_start_applying_dynamic_f0 = results['n_ttl_to_start_applying_dynamic_f0']
    # self.n_frames_search_forward = results['n_frames_search_forward']
    # self.drift_array = results['drift_array']
    # self.lick_detector_abstime = results['lick_detector_abstime']
    # self.rotary_encoder1_abstime = results['rotary_encoder1_abstime']
    # self.rotary_encoder2_abstime = results['rotary_encoder2_abstime']

    #
    reward_times = results['reward_times'].T
    reward_times = reward_times[reward_times[:,1]>0]
    n_rewards = reward_times.shape[0]

    #
    return rois_traces_smooth1, rois_traces_smooth2, reward_times


def butter_bandpass(lowcut, highcut, fs, order=5):
    nyq = 0.5 * fs
    low = lowcut / nyq
    high = highcut / nyq
    sos = butter(order, [low, high],analog=False, btype='band', output='sos')
    #b, a = scipy.signal.cheby1(order, [low, high], btype='band')
    return sos


def butter_bandpass_filter(data, lowcut, highcut, fs, order=5):
    sos = butter_bandpass(lowcut, highcut, fs, order=order)
    y = sosfilt(sos, data)
    return y

#
def get_reward_triggered_bmi_ensembles(session_id, 
                                       c,
                                       window,
                                       n_tests = 100
                                       ):

    # get reward times for session
    reward_times = c.reward_times[session_id]
   # print ("session: ", session_id, " of", len(c.session_ids)-1, ",  reward times: ", reward_times.shape)

    # get cells
    if c.use_DFF:
        cells = c.sessions[session_id].F_detrended
    else:
        cells = c.sessions[session_id].F_upphase_bin

    # load the esnembel matching info
    fname_ensemble_matching = os.path.join(c.root_dir,
                                            c.animal_id,
                                            str(c.session_ids[session_id]),
                                            'ensembles',
                                            'ensembles_matches_bmi_to_suite2p.npz')

    #
    d = np.load(fname_ensemble_matching, allow_pickle=True)
    idxs = []
    idxs.append(d['idx_ensemble1_0'])
    idxs.append(d['idx_ensemble1_1'])
    idxs.append(d['idx_ensemble2_0'])
    idxs.append(d['idx_ensemble2_1'])
    
    #
    psth = []
    psth_shuffled = []
    n_bursts = []
    ctr =0
    for idx in idxs:
        psth.append([])
        psth_shuffled.append([])
        temp = cells[idx]

        # count the number of bursts in temp
        diff = temp[1:]-temp[:-1]
        idx = np.where(diff==1)[0]
        n_bursts.append(idx.shape[0])

        #
        for r in reward_times:
            temp2 = temp[r-window:r+window]
            if temp2.shape[0]==window*2:
                psth[ctr].append(temp2)

            # same for random times
            temp3 = []
            for n in range(n_tests):
                idx = np.random.choice(np.arange(window,cells.shape[1]-window,1))
                temp2 = temp[idx-window:idx+window]
                if temp2.shape[0]==window*2:
                    temp3.append(temp2)

            psth_shuffled[ctr].append(np.mean(temp3,0))

        #
        ctr+=1
    #
    psths = np.array(psth)
    psths_shuffled = np.array(psth_shuffled)

    #

    return psths, psths_shuffled, n_bursts

def get_reward_triggered_psth(session_id, 
                              c,
                              window,
                              idx_cells,
                              global_order,
                              n_tests = 100):
    
    fname_out = os.path.join(c.root_dir,
                            c.animal_id,
                            str(c.session_ids[session_id]),
                            'results',
                            'reward_triggered_psth.npz')
    
    if os.path.exists(fname_out)==False:

        # get reward times for session
        reward_times = c.reward_times[session_id]
        print ("session: ", session_id, " of", 
            len(c.session_ids)-1, 
            ",  reward times: ", len(reward_times))

        # get cells
        cells = c.sessions[session_id].F_upphase_bin
        print ("[ca] matrix: ", cells.shape)

        #
        psth = []
        psth_shuffled = []
        n_bursts = []
        for k in range(cells.shape[0]):
            psth.append([])
            psth_shuffled.append([])
            temp = cells[k]

            # count the number of bursts in temp
            diff = temp[1:]-temp[:-1]
            idx = np.where(diff==1)[0]
            n_bursts.append(idx.shape[0])

            #
            for r in reward_times:
                temp2 = temp[r-window:r+window]
                if temp2.shape[0]==window*2:
                    psth[k].append(temp2)

                # same for random times
                temp3 = []
                for n in range(n_tests):
                    idx = np.random.choice(np.arange(window,cells.shape[1]-window,1))
                    temp2 = temp[idx-window:idx+window]
                    if temp2.shape[0]==window*2:
                        temp3.append(temp2)

                psth_shuffled[k].append(np.mean(temp3,0))

        #
        psths = np.array(psth)
        psths_shuffled = np.array(psth_shuffled)

        #
        np.savez(fname_out, 
                 psths= psths,
                psths_shuffled = psths_shuffled,
                n_bursts = n_bursts)
    
    else:
        d = np.load(fname_out, allow_pickle=True)
        psths = d['psths']
        psths_shuffled = d['psths_shuffled']
        n_bursts = d['n_bursts']


    # average across trials
    if True:
        psths_avg = np.nanmean(psths,1)
        psths_shuffled_avg = np.nanmean(psths_shuffled,1)
    else:
        psths_avg = np.nanmedian(psths,1)
        psths_shuffled_avg = np.nanmedian(psths_shuffled,1)

    # find cells with all nans
    try:
        psth_sums = np.nansum(psths_avg,1)
        idx = np.where(psth_sums==0)[0]
        psths_avg[idx]=0
    except:
        pass

    #
    if global_order==False or idx_cells==None:
        try:
            # order the cells by ptp in axis 1
            ptps = np.ptp(psths_avg,1)
            idx_cells = np.argsort(ptps)
        except:
            pass

    return psths_avg, psths_shuffled_avg, idx_cells, n_bursts

##########################################
from scipy.signal import butter, sosfilt

def butter_bandpass(lowcut, highcut, fs, order=5):
    nyq = 0.5 * fs
    low = lowcut / nyq
    high = highcut / nyq
    sos = butter(order, [low, high],analog=False, btype='band', output='sos')
    #b, a = scipy.signal.cheby1(order, [low, high], btype='band')
    return sos


def butter_bandpass_filter(data, lowcut, highcut, fs, order=5):
    sos = butter_bandpass(lowcut, highcut, fs, order=order)
    y = sosfilt(sos, data)
    return y

#
def plot_multi_session_psth_imshow(c,
                                   session_ids,
                                   psth_array,
                                   window,
                                   n_bursts_array,
                                   vmax):
    
    #
    #plt.figure()
    from matplotlib import gridspec
    # make a gridspec with 7 rows and 7 columns
    gs = gridspec.GridSpec(8,len(c.session_ids)-1)

    #
    suas = []
    for k in range(len(psth_array)):
        ax = plt.subplot(gs[:6, k])

        #
        ax.imshow(psth_array[k],
                  aspect='auto',
                  interpolation='none',
                  vmin=0,vmax=vmax)

        # also make a sua by summing over axis 0
        sua = np.sum(psth_array[k],0)

        # plot sua as a while line
        ax.plot(sua,'white',linewidth=2)

        ##############################################
        # plot vertical line at reward time 
        ax.plot([window,window],[0,len(psth_array[k])],'r--')

        # change xticks to range from -3 to 3
        #tt1 = np.arange(0,psth_array.shape[2]+1,30)
        tt1 = np.linspace(0,len(psth_array[k][0]),5)
        tt2 = np.round(np.linspace(-window/30,window/30,tt1.shape[0]),1)
        ax.set_xticks(tt1,tt2)

        #
        if k==0:
            ax.set_ylabel("Cell #")
            ax.set_yticks(np.arange(0,len(psth_array[k]),5),
                          np.arange(0,len(psth_array[k]),5))
        else:
            ax.set_yticks([])

        #
        ax.set_ylim(0,len(psth_array[k]))

        #
        ax.set_title("Session: "+str(k+1)+"\n # rew: "
                     + str(len(c.reward_times[session_ids[k]])),fontsize=7 )

        #
        ax.set_xlabel("Time (sec)",fontsize=10)

        ##################################################
        ##################################################
        ##################################################
        # add the distribution of bursts at the bottom
        ax = plt.subplot(gs[-1,k])
        temp = n_bursts_array[k]

        y = np.histogram(temp, bins=np.arange(0,300,10))
        temp = y[1][1:]/30/60

        suas.append([temp,y[0]])

        ax.plot(temp,y[0],'black',linewidth=2)
        
        #
        ax.set_xlabel("# burst/min",fontsize=10)
        
        #
        if k==0:
            ax.set_ylabel("# cells")

        # plot vertical lines at 0.05, 0.1, 0.15
        ymin, ymax = ax.get_ylim()
        ax.plot([0.05,0.05],[0,ymax],
                '--',
                c='black', 
                alpha=0.5)
        ax.plot([0.1,0.1],[0,ymax],
                '--',
                c='black', 
                alpha=0.5)
        ax.plot([0.15,0.15],[0,ymax],
                '--',
                c='black', 
                alpha=0.5)

    #
    plt.suptitle(c.animal_id + " "+c.rec_type+ " (Vmax = "+str(vmax)+")",fontsize=14)

    plt.show()

    return suas

#
def plot_psth(c,
              psths_avg,
              psths_shuffled_avg,
              session_id,
              axes,
              start_cell,
              idx_cells,
              show_random=False,
              smoothing=False,
              window=100
              ):
    

    # make viridis colormap for line plotting up to 7 sessions
    cmap = matplotlib.cm.get_cmap('viridis')
    clrs = cmap(np.linspace(0.25, 0.85, 8))
    clrs_b = ['blue','lightblue','red','pink']
    #
    t = np.arange(-window,window,1)/30.

    #
    cell_ids = np.arange(start_cell,start_cell+100,1)
    for ctr, k in enumerate(cell_ids):
        ax=axes[ctr]
        
        #
        temp = psths_avg[idx_cells[k]]
        temp2 = psths_shuffled_avg[idx_cells[k]]
        # use savgol filter to smooth from scipy
        if smoothing:
            temp = butter_bandpass_filter(temp, 0.01, 1, 30, order=1)
            temp2 = butter_bandpass_filter(temp2, 0.01, 1, 30, order=1)
        
        if k==0:
            ax.plot(t,temp,
                c=clrs[session_id-1],
                label="session: "+str(session_id))
            ax.legend(fontsize=5)
        else:
            ax.plot(t,temp,
                c=clrs[session_id-1])

        #
        if show_random:
            ax.plot(t,temp2,'black', alpha=0.5)

        ###############################################
        ###############################################
        ###############################################
        # plot vertical line at reward time
        # get ylimits or range from axes[k]
        ax.relim()
        ax.autoscale_view()
        
        ymin, ymax = ax.get_ylim()
        #ax.set_ylim(bottom=0)

        ax.plot([t[window],t[window]],
                 [0,ymax],'b--')
        
        # 
        idx_b = np.where(idx_cells[k]==c.best_matches)[0]
        if idx_b.shape[0]>0:
            print ("found roi: ", idx_b[0], ", matches cell: ", idx_cells[k],
                   "ctr : ", ctr)
            ax.set_title(str(idx_cells[k]),
                         fontsize=10,
                         color = clrs_b[idx_b[0]],
                         pad=0.1)
        else:
            ax.set_title(str(idx_cells[k]),
                         fontsize=10,
                         pad=0.1)


        #
        #if ctr%10==0:
        #ax.set_ylabel("DF/F")
        #else:
        #    ax.set_yticks([])

        # set yticks fontsize
        ax.tick_params(axis='y', labelsize=5, pad=0.1)

        # 
        if ctr<90:
            ax.set_xticks([])
        else:
            ax.set_xlabel("Time (sec)",fontsize=10)

    return idx_cells



#
def plot_psth_rois_only(c,
                        psths_avg,
                        psths_shuffled_avg,
                        session_id,
                        axes,
                        start_cell,
                        idx_cells,
                        show_random=False,
                        smoothing=False,
                        window=100
                        ):

    # make viridis colormap for line plotting up to 7 sessions
    cmap = matplotlib.cm.get_cmap('viridis')
    clrs = cmap(np.linspace(0.25, 0.85, 8))
    clrs_b = ['blue','lightblue','red','pink']
    #
    t = np.arange(-window,window,1)/30.

    #
    cell_ids = np.arange(c.sessions[session_id].F_upphase_bin.shape[0])
    ctr=0
    for k in cell_ids:
    #for ctr, k, in enumerate(c.best_matches):
        
        #
        temp = psths_avg[idx_cells[k]]
        temp2 = psths_shuffled_avg[idx_cells[k]]
        # use savgol filter to smooth from scipy
        if smoothing:
            temp = butter_bandpass_filter(temp, 0.01, 1, 30, order=1)
            temp2 = butter_bandpass_filter(temp2, 0.01, 1, 30, order=1)
        
       

        ###############################################
        ###############################################
        ###############################################
        # plot vertical line at reward time
        # get ylimits or range from axes[k]
        
        # 
        idx_b = np.where(idx_cells[k]==c.cells_to_plot)[0]
        if idx_b.shape[0]>0:
            ax=axes[ctr]
            ax.set_title(str(idx_cells[k]),
                         fontsize=10,
                         color = clrs_b[idx_b[0]],
                         pad=0.1)
            
            #
            if ctr==0:
                ax.plot(t,temp,
                    c=clrs[session_id-1],
                    label="session: "+str(session_id))
                ax.legend(fontsize=5)
            else:
                ax.plot(t,temp,
                    c=clrs[session_id-1])

            #
            if show_random:
                ax.plot(t,temp2,'black', alpha=0.5)
                
            ax.relim()
            ax.autoscale_view()
            
            ymin, ymax = ax.get_ylim()
            #ax.set_ylim(bottom=0)

            ax.plot([t[window],t[window]],
                    [0,ymax],'b--')


            # set yticks fontsize
            #ax.tick_params(axis='y', labelsize=5, pad=0.1)

            # 
            ax.set_xlabel("Time (sec)",fontsize=10)

            #
            ctr+=1


    return idx_cells

#
def preprocess_trace(temp):

    # smooth with a sliding window
    temp = savgol_filter(temp, 101, 3)

    #
    low_cutoff = 0.1
    high_cutoff = 3
    sample_rate = 30
    temp = butter_bandpass_filter(temp,
                                low_cutoff,
                                high_cutoff,
                                sample_rate,
                                order=1)
    
    return temp

#
def find_best_match_bmi_vs_mastermask(cell_name,
                                      root_dir,
                                      animal_id,
                                      sessions_F_filtered,
                                      session_ids):

    # loop over all sessions
    corrs_array = []
    best_cells = []
    for session_id in range(1,len(sessions_F_filtered),1):

        # these are already loaded, so can just index into them
        cells = sessions_F_filtered[session_id]

        # # load results.npz file
        rois1, rois2, _ = load_results_npz_standalone(root_dir,
                                                      animal_id,
                                                      session_id,
                                                      session_ids)
        #
        if cell_name=='roi_pos1':
            roi = rois1[0]
        elif cell_name=='roi_pos2':
            roi = rois1[1]
        elif cell_name=='roi_neg1':
            roi = rois2[0]
        elif cell_name=='roi_neg2':
            roi = rois2[1]
        else:
            print ("Cell does not exist: ", cell_name)
            return
        
        #
        roi[:100] = 0
        roi = preprocess_trace(roi)

        #
        corrs = []
        for cell in cells:

            # smooth cell using savgol filter
            cell = preprocess_trace(cell)

            # Compute the Pearson correlation
            try:
                correlation, _ = pearsonr(cell, roi)
            except:
                correlation = 0
            corrs.append(correlation)

        # find armgax as best match
        idx = np.argmax(corrs)

        # plt.figure()
        # cell = preprocess_trace(cells[380])
        # plt.plot(cell)
        # plt.plot(roi)
        # plt.show()

        #
        corrs_array.append(corrs)
        best_cells.append(idx)

    #
    return best_cells


def compute_correlations_parallel(data_dir,
                                  rasters,
                                  rasters_DFF,
                                  n_cores,
                                  #method='all',
                                  binning_window=30,
                                  subsample=5,
                                  scale_by_DFF=False,
                                  corr_parallel_flag=True,
                                  zscore=False,
                                  n_tests_zscore=1000,
                                  recompute_correlation=False,
                                  min_number_bursts=0):

    """
    This function computes pairwise Pearson correlations between different rasters in a parallelized manner.

    Parameters:
    data_dir (str): The directory where the data is stored.
    rasters (np.array): The rasters to be processed.
    rasters_DFF (np.array): The rasters to be processed after applying DeltaF/F.
    n_cores (int): The number of cores to be used for parallel processing.
    binning_window (int, optional): The size of the binning window. Default is 30.
    subsample (int, optional): The subsampling rate. Default is 5.
    scale_by_DFF (bool, optional): A flag indicating whether to scale by DeltaF/F. Default is False.
    corr_parallel_flag (bool, optional): A flag indicating whether to run the correlation computation in parallel. Default is True.
    zscore (bool, optional): A flag indicating whether to compute the z-score. Default is False.
    n_tests_zscore (int, optional): The number of tests to be performed for z-score computation. Default is 1000.
    recompute_correlation (bool, optional): A flag indicating whether to recompute the correlation. Default is False.
    min_number_bursts (int, optional): The minimum number of bursts required. Default is 0.

    Returns:
    None. The function saves the computed correlations to a .npz file in 'data_dir'.
    
    Note: If a file with the same name already exists in 'data_dir' and 'recompute_correlation' is False,
          the function will return without doing anything.
    """
    # make a small class to hold all the input variables
    class C:
        pass

    c1 = C()
    c1.n_cores = n_cores
    c1.n_tests = n_tests_zscore
    #c1.correlation_method = method
    c1.binning_window = binning_window
    c1.subsample = subsample
    c1.scale_by_DFF = scale_by_DFF
    c1.corr_parallel_flag = corr_parallel_flag
    c1.zscore = zscore
    c1.rasters = rasters
    c1.rasters_DFF = rasters_DFF
    c1.recompute_correlation = recompute_correlation
    c1.min_number_bursts = min_number_bursts
    
    #
    print ("... computing pairwise pearson correlation ...")
    print (" RASTERS IN: ", rasters.shape)
    print (" BINNING WINDOW: ", binning_window)

    # split data
    #ids = np.array_split(np.arange(rasters.shape[0]),100)
    # for correlations_parallel2 function we don't need to split ids anymore

    # make output directory 'correlations'
    # check to see if data_dir exists:
    # if os.path.exists(data_dir)==False:
    #     os.mkdir(data_dir)

    # add dynamic data_dir
    # if zscore:
    #     data_dir = os.path.join(data_dir,'zscore')
    #     if os.path.exists(data_dir)==False:
    #         os.mkdir(data_dir)
    # else:
    #     data_dir = os.path.join(data_dir,'threshold')
    #     if os.path.exists(data_dir)==False:
    #         os.mkdir(data_dir)

    # # finally add the 'correlations' directory
    # data_dir = os.path.join(data_dir,'correlations')
    # if os.path.exists(data_dir)==False:
    #     os.mkdir(data_dir)

    #
    c1.data_dir = data_dir

    # convert object into a dictionary
    c1 = c1.__dict__

    #############################################################
    #############################################################
    #############################################################
    # run parallel
    ids = np.arange(rasters.shape[0])
    if corr_parallel_flag:
        parmap.map(correlations_parallel2,
                    ids,
                    c1,
                    pm_processes=n_cores,
                    pm_pbar = True
                    )
    else:
        for k in tqdm(ids, desc='computing intercell correlation'):
            correlations_parallel2(k,
                                   c1)
            


# this computes the correlation for a single cell against all others and then saves it to disk
def correlations_parallel2(id, 
                           c1):
    """
    This function computes the correlation between different rasters in a parallelized manner.

    Parameters:
    id (int): The ID of the raster to be processed.
    c1 (dict): A dictionary containing various parameters and data needed for the computation.

    The dictionary 'c1' has the following keys:
    - data_dir (str): The directory where the data is stored.
    - rasters (np.array): The rasters to be processed.
    - rasters_DFF (np.array): The rasters to be processed after applying DeltaF/F.
    - binning_window (int): The size of the binning window.
    - subsample (int): The subsampling rate.
    - scale_by_DFF (bool): A flag indicating whether to scale by DeltaF/F.
    - zscore (bool): A flag indicating whether to compute the z-score.
    - n_tests (int): The number of tests to be performed.
    - recompute_correlation (bool): A flag indicating whether to recompute the correlation.
    - min_number_bursts (int): The minimum number of bursts required.

    Returns:
    None. The function saves the computed correlations to a .npz file in 'data_dir'.
    
    Note: If a file with the same name already exists in 'data_dir' and 'recompute_correlation' is False, 
          the function will return without doing anything.
    """
    # extract values from dicionary c1
    data_dir = c1["data_dir"]
    rasters = c1["rasters"]
    rasters_DFF = c1["rasters_DFF"]
    binning_window = c1["binning_window"]
    subsample = c1["subsample"]
    scale_by_DFF = c1["scale_by_DFF"]
    zscore = c1["zscore"]
    n_tests = c1["n_tests"]
    recompute_correlation = c1["recompute_correlation"]
    min_number_bursts = c1["min_number_bursts"]

    # 
    fname_out = os.path.join(data_dir,
                                str(id)+ '.npz'
                                )

    # not used for now, but may wish to skip computation if file already exists
    if os.path.exists(fname_out) and recompute_correlation==False:
        return

    #        
    temp1 = rasters[id][::subsample]

    # scale by rasters_DFF
    if scale_by_DFF:
        temp1 = temp1*rasters_DFF[id][::subsample]

    # bin data in chunks of size binning_window
    if binning_window!=1:
        tt = []
        for q in range(0, temp1.shape[0], binning_window):
            temp = np.sum(temp1[q:q + binning_window])
            tt.append(temp)
        temp1 = np.array(tt)

    #
    corrs = []
    for p in range(rasters.shape[0]):
        temp2 = rasters[p][::subsample]
        
        # scale by rasters_DFF
        if scale_by_DFF:
            temp2 = temp2*rasters_DFF[p][::subsample]
        
        # 
        if binning_window!=1:
            tt = []
            for q in range(0, temp2.shape[0], binning_window):
                temp = np.sum(temp2[q:q + binning_window])
                tt.append(temp)
            temp2 = np.array(tt)
        
        #
        corr, corr_z = get_corr2(temp1, temp2, zscore, n_tests, min_number_bursts)

        # 
        corrs.append([id, p, corr[0], corr[1], corr_z[0]])

    #
    corrs = np.vstack(corrs)
    #
    np.savez(fname_out, 
             binning_window = binning_window,
             subsample = subsample,
             scale_by_DFF = scale_by_DFF,
             zscore_flag = zscore,
             id = id,
             compared_cells = corrs[:,1],
             pearson_corr = corrs[:,2],
             pvalue_pearson_corr = corrs[:,3],
             z_score_pearson_corr = corrs[:,4],
             n_tests = n_tests,
            )

    #return corrs

#
def get_corr3_parallel(p,
                       id,
                       rasters,
                       rasters_DFF, 
                       subsample,
                       zscore, 
                       n_tests=1000, 
                       min_number_bursts=0,
                       scale_by_DFF=True,
                       binning_window=30,
                       ):


    #
    temp1 = rasters[id][::subsample]
    temp2 = rasters[p][::subsample]
    
    # scale by rasters_DFF
    if scale_by_DFF:
        temp2 = temp2*rasters_DFF[p][::subsample]
    
    # 
    if binning_window!=1:
        
        tt = []
        for q in range(0, temp1.shape[0], binning_window):
            temp = np.sum(temp1[q:q + binning_window])
            tt.append(temp)
        temp1 = np.array(tt)
    
        tt = []
        for q in range(0, temp2.shape[0], binning_window):
            temp = np.sum(temp2[q:q + binning_window])
            tt.append(temp)
        temp2 = np.array(tt)

       

    # check if all values are the same 
    if len(np.unique(temp1))==1 or len(np.unique(temp2))==1:
        corr = [np.nan,1]
        return corr, [np.nan]
    
    # check if number bursts will be below self.min_num_bursts
    if np.sum(temp1!=0)<min_number_bursts or np.sum(temp2!=0)<min_number_bursts:
        corr = [np.nan,1]
        return corr, [np.nan]

    # if using dynamic correlation we need to compute the correlation for 1000 shuffles
    corr_original = scipy.stats.pearsonr(temp1, temp2)

    # make array and keep track
    corr_array = []
    corr_array.append(corr_original[0])
                                
    #
    if zscore:
        corr_s = []
        for k in range(n_tests):
            # choose a random index ranging from 0 to the length of the array minus 1
            idx = np.random.randint(-temp2.shape[0], temp2.shape[0],1)
            #idx = np.random.randint(temp2.shape[0])
            temp2_shuffled = np.roll(temp2, idx)
            corr_s = scipy.stats.pearsonr(temp1, temp2_shuffled)
            corr_array.append(corr_s[0])

        # compute z-score
        corr_z = stats.zscore(corr_array)

    else:
        corr_z = [np.nan]

    return corr_original, corr_z



def get_corr2(temp1, temp2, zscore, n_tests=1000, min_number_bursts=0):
    """
    This function calculates the Pearson correlation coefficient between two arrays of data, temp1 and temp2. 
    If zscore is True, the function also calculates the z-score of the correlation coefficient based on n_tests random shuffles of temp2.
    
    :param temp1: 1D numpy array of data
    :param temp2: 1D numpy array of data
    :param zscore: boolean, if True calculate z-score of correlation coefficient
    :param n_tests: int, number of random shuffles to perform when calculating z-score (default=1000)
    :param min_number_bursts: int, minimum number of bursts
    :return: tuple containing the Pearson correlation coefficient between temp1 and temp2, and the z-score of the correlation coefficient (if zscore=True) or np.nan (if zscore=False)
    """
    # check if all values are the same 
    if len(np.unique(temp1))==1 or len(np.unique(temp2))==1:
        corr = [np.nan,1]
        return corr, [np.nan]
    
    # check if number bursts will be below self.min_num_bursts
    if np.sum(temp1!=0)<min_number_bursts or np.sum(temp2!=0)<min_number_bursts:
        corr = [np.nan,1]
        return corr, [np.nan]

    # if using dynamic correlation we need to compute the correlation for 1000 shuffles
    corr_original = scipy.stats.pearsonr(temp1, temp2)

    # make array and keep track
    corr_array = []
    corr_array.append(corr_original[0])
                                
    #
    if zscore:
        corr_s = []
        for k in range(n_tests):
            # choose a random index ranging from 0 to the length of the array minus 1
            idx = np.random.randint(-temp2.shape[0], temp2.shape[0],1)
            #idx = np.random.randint(temp2.shape[0])
            temp2_shuffled = np.roll(temp2, idx)
            corr_s = scipy.stats.pearsonr(temp1, temp2_shuffled)
            corr_array.append(corr_s[0])

        # compute z-score
        corr_z = stats.zscore(corr_array)

    else:
        corr_z = [np.nan]

    return corr_original, corr_z

#
def get_si2(rate_map,
            time_map):
    
    #
    sparsity = None
    selectivity = None
    
    #
    duration = np.ma.sum(time_map)
    position_PDF = time_map / (duration + np.spacing(1))

    mean_rate = np.ma.sum(rate_map * position_PDF)
    mean_rate_sq = np.ma.sum(np.ma.power(rate_map, 2) * position_PDF)

    max_rate = np.max(rate_map)

    if mean_rate_sq != 0:
        sparsity = mean_rate * mean_rate / mean_rate_sq

    if mean_rate != 0:
        selectivity = max_rate / mean_rate

    log_argument = rate_map / (mean_rate+0.00001)
    log_argument[log_argument < 1] = 1

    #
    inf_rate = np.ma.sum(position_PDF * rate_map * np.ma.log2(log_argument))
    inf_content = inf_rate / (mean_rate+0.00001)

    return inf_rate, inf_content, sparsity, selectivity

#
def compute_si_1D(cell_ids,
                  cells_upphase,           # this is the calcium activivty in the session
                  reward_times,
                  root_dir,
                  animal_id,
                  session_name,
                  window,
                  n_tests,
                  recompute=False,
                  plotting=False):

    #
    for cell_id in cell_ids:
        #
        fname_out = os.path.join(root_dir,
                                animal_id,
                                session_name,
                                'spatial_info',
                                str(cell_id)+'.npz')

        #                             
        if os.path.exists(fname_out)==False or recompute:

            #
            psth = []
            psth_shuffled = []
            n_bursts = []
            #for k in trange(cells.shape[0]):
            k = cell_id
            psth.append([])
            psth_shuffled.append([])

            # [ca] activity
            cell_upphase = cells_upphase[k]

            # count the number of bursts in temp
            diff = cell_upphase[1:]-cell_upphase[:-1]
            idx = np.where(diff==1)[0]
            n_bursts.append(idx.shape[0])

            # So time map should just count over and over again 
            time_map = []
            rate_map = []
            try:
                time_map = np.arange(window*2)*reward_times.shape[0]
            except:
                # most likely the .npz file with the reward times is corrupt
                # so try to recover it from the excel file
                fname = os.path.join(root_dir,
                                     animal_id,
                                     session_name,
                                    'results_fixed.xlsx')

                reward_times = find_reward_times_corrupt_npz(fname)

                #
                time_map = np.arange(window*2)*reward_times.shape[0]


            for r in reward_times:
                temp2 = cell_upphase[r-window:r+window]
                if temp2.shape[0]==window*2:
                    rate_map.append(temp2)
                    #time_map.append(temp_time_map)

            #
            sum_map = np.sum(rate_map,0)
            ave_map = np.mean(rate_map,0)
            rate_map = np.sum(rate_map,0)
            #time_map = temp

            # get information rate and content
            #print ("rate map: ", rate_map.shape, " time map: ", time_map.shape)
            inf_rate, inf_content, sparsity, selectivity = get_si2(rate_map, time_map)
            si_rate = inf_rate
            si = inf_content

            # same but now we shuffle the time map
            si_shuffle = []
            ave_map_shuffle = []
            for _ in range(n_tests):
                
                rate_map_shuffled = []
                time_map_shuffled = []
                for _ in reward_times:
                    # get random r value
                    r = np.random.choice(np.arange(window, cell_upphase.shape[0]-window,1))

                    temp2 = cell_upphase[r-window:r+window]
                    if temp2.shape[0]==window*2:
                        rate_map_shuffled.append(temp2)
                        #time_map_shuffled.append(temp_time_map)
                    
                #
                ave_map_shuffle.append(np.mean(rate_map_shuffled,0))
                rate_map_shuffled = np.sum(rate_map_shuffled,0)
                #time_map_shuffled = np.hstack(time_map_shuffled)
                
                #
                inf_rate, _, _, _= get_si2(rate_map_shuffled, time_map)
                si_shuffle.append(inf_rate)

            #
            stack = np.hstack([si_rate, si_shuffle])
            zscore = stats.zscore(stack)[0]

            #
            # print ("cell: ", k, " of ", cells.shape[0], 
            #         " si: ", si[k], " si rate: ", si_rate[k], " zscore: ", zscore[k])

            #
            if plotting:
                ave_map_shuffle = np.vstack(ave_map_shuffle)
                plt.figure()
                plt.plot(ave_map)
                #
                temp = np.mean(ave_map_shuffle,0)
                plt.plot(temp,
                        linewidth=5,)
                plt.show()

            # save each cell results
            np.savez(fname_out,
                     reward_times = reward_times,
                     sum_map = sum_map,
                     ave_map = ave_map,
                     ave_map_shuffle = ave_map_shuffle,
                     rate_map = rate_map,
                     time_map = time_map,
                     si = si,
                     si_rate = si_rate,
                     sparsity = sparsity,
                     selectivity = selectivity,
                     si_shuffle = si_shuffle,
                     zscore = zscore,
                     n_bursts = n_bursts)
            

def compute_corr2(roi, F_filtered):

    corr_array = np.zeros((F_filtered.shape[0]))
    temp = roi
    f0 = np.median(temp)
    temp = (temp-f0)/f0
    for k in range(F_filtered.shape[0]):
        #
        res = scipy.stats.pearsonr(F_filtered[k],temp)
        corr_array[k]=res[0]
    
    return corr_array


def find_reward_times_corrupt_npz(fname):


    fname_out = os.path.split(fname)[0] + '/reward_times_recovered.npz'

    # check if the file already exists
    if os.path.exists(fname_out):
        d = np.load(fname_out, allow_pickle=True)
        return d['reward_times']

    # load the excel file
    import pandas as pd
    df = pd.read_excel(fname)

    # grab column named "water_reward"
    water_reward = df['water_reward'].values

    # find reward onsets from 0 to 1 in the boolean time series
    idxs = np.where(water_reward==1)[0]
    diffs = idxs[1:]-idxs[:-1]
    idx2 = np.where(diffs>1)[0]


    idx3 = idxs[idx2+1]
    # add idxs[0] to idx3
    idx3 = np.append(idxs[0],idx3)

    #
    plt.figure()
    plt.plot(water_reward)
    # and plot vertical lines at idx3
    for k in range(len(idx3)):
        plt.plot([idx3[k],idx3[k]],[0,1],'r--')


    plt.show()

    # save the idx3 as an npz file
    np.savez(fname_out, 
             reward_times=idx3)
    

    return idx3

#
def get_trials2(
                white_noise_state,
                post_reward_state,
                reward_lockout_counter
                ):

    # simpler version of the above function 
    # we loop over reward_lockout_counter 

    trials = []
    start = 0
    in_trial = True

    #print ("reward_lockout_counter: ", reward_lockout_counter)
    #print (np.where(reward_lockout_counter>=0)[0].shape[0])
    for k in trange(reward_lockout_counter.shape[0]):
        # check if we're hit a reward lockout
        if reward_lockout_counter[k]>=0:
            # check to see if in trial that should be ended
            if in_trial:
                # end trial
                in_trial = False
                end = k
                
                # also check to see if the trial ended into a post-reward state
                if post_reward_state[k+1]==1:
                    # add trial
                    trials.append([start,end,1])
                else: 
                    # it was a non-rewarded trial
                    trials.append([start,end,0])
        
        # else we are in a rewardable period    
        elif reward_lockout_counter[k]<0: 
            # check to see if we're in a trial otherwise start one
            if in_trial==False:
                # start trial
                in_trial = True
                start = k
    
    return np.array(trials)


def process_populations(c, plotting=False):
    window = 30*6
    psth_array = []
    idx_cells = None
    global_order = False
    session_ids = np.arange(1,len(c.session_ids),1)
    n_tests = 1

    #
    n_bursts_array = []
    for session_id in session_ids:
    #for session_id in [1]:

        #
        psths_avg, psths_shuffled_avg, idx_cells, n_bursts  = get_reward_triggered_psth(session_id, 
                                                                                        c,
                                                                                        window,
                                                                                        idx_cells,
                                                                                        global_order,
                                                                                        n_tests)
        
        #
        n_bursts_array.append(n_bursts)
        
        #
        psth_array.append(psths_avg[idx_cells])

        if global_order==False:
            idx_cells = None

    #
    if plotting:
        vmax = 0.3
        plt.figure(figsize=(20,12))
        sua = plot_multi_session_psth_imshow(c,
                                        session_ids,
                                        psth_array,
                                        window,
                                        n_bursts_array,
                                        vmax=vmax)


def generate_learning_profiles(root_dir, 
                               animal_ids,
                               group_name, 
                               CA3_groups, 
                               M1_groups, 
                               threshold, 
                               norm_percentage, 
                               normalize):

    #
    plt.figure()
    pops = [[],[],[]]
    for animal_id in tqdm(animal_ids):

        #session_ids = np.arange(1,len(c.session_ids),1)
        if group_name=='M1':
            idx = find_animal_group(M1_groups, animal_id)
        else:
            idx = find_animal_group(CA3_groups, animal_id)
        
        # if we can't find the group, skip
        if idx ==-1:
            continue

        #
        plt.subplot(1,3,idx+1)

        #       
        c = ProcessCalcium(root_dir,
                        animal_id)
        
        #
        n_cells = []
        if animal_id in ['DON-011733']:
            thresh = 0.1
        else:
            thresh = threshold

        #
        for session_id in range(1,len(c.session_ids),1):

            #    
            fname_out = os.path.join(c.root_dir,
                                    c.animal_id,
                                    str(c.session_ids[session_id]),
                                    'results',
                                    'reward_triggered_psth.npz')
            
            if os.path.exists(fname_out)==False:
                n_cells.append(np.nan)
                continue

            # get # of cells by looking into the binarized file
            fname_binarized = os.path.join(c.root_dir,
                                            c.animal_id,
                                            str(c.session_ids[session_id]),
                                            'plane0',
                                            'binarized_traces.npz')
            d = np.load(fname_binarized, mmap_mode='r', allow_pickle=True)
            F_upphase = d['F_upphase']
            total_cells = F_upphase.shape[0]

            #
            d = np.load(fname_out, allow_pickle=True)
            psths = d['psths']
            psths_shuffled = d['psths_shuffled']
            n_bursts = d['n_bursts']

            # check how many pshs have a peak > thrsh
            psth_mean = np.mean(psths, axis=1)

            #
            try:
                temp = np.where(psth_mean.max(1)>thresh)[0]
            except:
                print (fname_out)
                #

            if norm_percentage:
                n_cells.append(temp.shape[0]/total_cells*100)
            else:
                n_cells.append(temp.shape[0])

        #
        if normalize:
            n_cells = np.array(n_cells)/np.max(n_cells)
        else:
            n_cells = np.array(n_cells)

        
        #
        if n_cells.shape[0]<8:
            n_cells = np.concatenate((n_cells, np.zeros(8-n_cells.shape[0])+np.nan))
        
        #
        engagement = n_cells[3]/n_cells[0]

        #
        plt.plot(n_cells,
                linewidth = 4, 
                label=animal_id + " " + str(np.round(engagement*100,2)) + "%")
        #
        pops[idx].append(n_cells)
        
        #
        plt.legend()

    # now compute average pops and std also
    for k in range(len(pops)):
        ax = plt.subplot(1,3,k+1)
        temp = np.array(pops[k])
        mean = np.nanmean(temp, axis=0)
        std = np.nanstd(temp, axis=0)
        try:
            plt.plot(mean, c='black')
            plt.fill_between(np.arange(mean.shape[0]), mean-std, mean+std, alpha=.2, color='black')
            plt.ylabel("% of cells with reward triggered response")
            plt.xlabel("Sessions")
        except:
            pass

    plt.show()




def find_animal_group(groups, animal_id):
    for k in range(len(groups)):
        if animal_id in groups[k]:
            return k
    return -1



def run_pearson_corr_single_cell_vs_all_cells_parallel(idxs, 
                                                       dir_corrs,
                                                       rasters,
                                                       rasters_DFF, 
                                                       n_tests=1000, 
                                                       min_number_bursts=0
                                                      ):
    #
    for id in idxs:

        fname_out = os.path.join(dir_corrs, 
                                str(id)+'.npz')
        
        if os.path.exists(fname_out):
            return

        #
        subsample = 1
        scale_by_DFF = True
        n_tests = 1000
        zscore = True
        binning_window = 30
        n_cores = 16

        # not used for now, but may wish to skip computation if file already exists
        if os.path.exists(fname_out): # and recompute_correlation==False:
            return

        #        
        temp1 = rasters[id][::subsample]

        # scale by rasters_DFF
        if scale_by_DFF:
            temp1 = temp1*rasters_DFF[id][::subsample]

        # bin data in chunks of size binning_window
        if binning_window!=1:
            tt = []
            for q in range(0, temp1.shape[0], binning_window):
                temp = np.sum(temp1[q:q + binning_window])
                tt.append(temp)
            temp1 = np.array(tt)

        #
        corrs = []
        ps = np.arange(rasters.shape[0])
          
        #
        if True:
            res = parmap.map(get_corr3_parallel,
                        ps,
                        id,
                        rasters,
                        rasters_DFF,
                        subsample,
                        zscore, 
                        n_tests, 
                        min_number_bursts,
                        pm_processes=n_cores,
                        pm_pbar = True)
        else:
            # single core version
            res = []
            for p in ps:
                res.append(get_corr3_parallel(p,
                                              id,
                                              rasters,
                                              rasters_DFF,
                                              subsample,
                                              zscore, 
                                              n_tests, 
                                              min_number_bursts))
                
        # extract corr and corr_z from res
        corrs
        for p in range(len(res)):
            corr = res[p][0]
            corr_z = res[p][1]

            # 
            corrs.append([id, p, corr[0], corr[1], corr_z[0]])

        #############
        corrs = np.vstack(corrs)

        #
        np.savez(fname_out, 
                binning_window = binning_window,
                subsample = subsample,
                scale_by_DFF = scale_by_DFF,
                zscore_flag = zscore,
                id = id,
                compared_cells = corrs[:,1],
                pearson_corr = corrs[:,2],
                pvalue_pearson_corr = corrs[:,3],
                z_score_pearson_corr = corrs[:,4],
                n_tests = n_tests,
                )

#
def compute_ensemble_only_correlations(c):  

    #
    for session_id in trange(1,len(c.session_ids), desc='computing ensemble only correlations'):

        #
        fname_ensemble_matching = os.path.join(c.root_dir, 
                                            c.animal_id, 
                                            str(c.session_ids[session_id]), 
                                            'ensembles',
                                            'ensembles_matches_bmi_to_suite2p.npz')

        #        
        data = np.load(fname_ensemble_matching, allow_pickle=True)

        #
        idxs = []
        idxs.append(data['idx_ensemble1_0'])
        idxs.append(data['idx_ensemble1_1'])
        idxs.append(data['idx_ensemble2_0'])
        idxs.append(data['idx_ensemble2_1'])
        idxs = np.array(idxs)

        # load the cell correlation vals 
        dir_corrs = os.path.join(c.root_dir,
                                    c.animal_id,
                                    str(c.session_ids[session_id]),
                                    'plane0',
                                    'correlations')
        
        # chekc if dir exists
        if os.path.exists(dir_corrs)==False:
            os.mkdir(dir_corrs)

        # compute the correlation
        run_pearson_corr_single_cell_vs_all_cells_parallel(
                                                        idxs, 
                                                        dir_corrs,
                                                        c.sessions[session_id].F_upphase_bin.copy(),
                                                        c.sessions[session_id].F_detrended.copy(), 
                                                        n_tests=1000, 
                                                        min_number_bursts=0
                                                        )
            