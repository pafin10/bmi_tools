import numpy as np
# Visualisation
import matplotlib.pyplot as plt
import numpy as np
import os
import pandas as pd
from tqdm import tqdm, trange
import parmap
import scipy
from openpyxl import load_workbook
from scipy.signal import savgol_filter
from scipy.interpolate import interp1d
from dtw import dtw
from matplotlib.widgets import Button

#
from calcium import Calcium

#
def smooth_ca_time_series4(diff):
	#
	''' This returns the last value. i.e. no filter

	'''

	temp = (diff[-1]*0.4+
			diff[-2] * 0.25 +
			diff[-3] * 0.15 +
			diff[-4] * 0.10 +
			diff[-5] * 0.10)

	return temp

#
class ProcessCalcium():

    def __init__(self, root_dir, 
                 animal_id 
                 ):

        #
        self.root_dir = root_dir
        self.animal_id = animal_id
        
        # load yaml file
        fname = os.path.join(self.root_dir,
                            self.animal_id,
                            animal_id+'.yaml')
        
        # load yaml file
        import yaml
        with open(fname) as file:
            doc = yaml.load(file, Loader=yaml.FullLoader)

        # print all the keys in doc
        print ("doc.keys(): ", doc.keys())

        self.session_ids = np.array(doc['session_ids'],dtype='str')

        #
        try:
            self.shifts = doc['shifts']
        except:
            print ("Could not find shifts in yaml file ... ")


        # load session types from each session yaml file
        self.session_types = []
        for session_id in self.session_ids:
            fname = os.path.join(self.root_dir,
                                self.animal_id,
                                session_id,
                                session_id+'.yaml')
            
            # load yaml file
            import yaml
            with open(fname) as file:
                doc = yaml.load(file, Loader=yaml.FullLoader)

            #
            self.session_types.append(doc['session_type'])

        #
        self.verbose = True

    # #
    def load_day0_mask(self):

    
        # fname = os.path.join(
        #                 self.root_dir,
        #                 self.animal_id,
        #                 'day0',
        #                 'rois_pixels_and_thresholds_day0.npz')

        # find session id with day0 session_type

        # find the session type that has 'day0' in it
        idx = np.where(np.array(self.session_types)=='day0')[0]

        fname = os.path.join(
                        self.root_dir,
                        self.animal_id,
                        self.session_ids[idx[0]],                        
                        'rois_pixels_and_thresholds_day0.npz')
        
        #
        try:
            data = np.load(fname, allow_pickle=True)
        except:
            print ("Could not Day0 masks ... ")
            return
           
        #
        contours_all_cells = data['contours_all_cells']
        self.cell_ids = data['cell_ids']

        self.contours_ROIs = contours_all_cells[self.cell_ids]

        
    #
    def compute_reward_centered_traces(self):

        
        if True:
            parmap.map(get_reward_centered_traces,
                        self.session_ids,
                        self.root_dir,
                        self.animal_id,
                        self.reward_window,
                        self.filter,
                        pm_processes=10,
                        pm_pbar=True)

    ###################
    def plot_network_graph(self):

        print (self.reward_centered_traces.shape)

        #
        temp = self.reward_centered_traces.copy().transpose(1,0,2)
        print (temp.shape)

        temp = temp.reshape(temp.shape[0],-1)
        print (temp.shape)

        # remove means from each cell
        temp = temp - np.median(temp,1)[:,np.newaxis]

        # compute the correlation matrix
        corr_matrix = np.corrcoef(temp)
        print (corr_matrix.shape)

        # set threshold for correlation matrix at 0.3
        corr_matrix[corr_matrix<0.3]=0

        ###################################################
        # use the corr_matrix as an adjacency matrix and compute the graph
        import networkx as nx
        G = nx.from_numpy_array(corr_matrix)

        # remove self-loops
        G.remove_edges_from(nx.selfloop_edges(G))

        # from pylab import rcParams
        # rcParams['figure.figsize'] = 14, 10
        pos = nx.spring_layout(G, scale=20, k=3/np.sqrt(G.order()))
        d = dict(G.degree)
        nx.draw(G, pos, 
                node_color='lightblue', 
                with_labels=False, 
                node_size = 2,
                width = 0.15,
                nodelist=d, 
                alpha=1,
                #node_size=[d[k]*300 for k in d]
                ax=self.ax
                )


        # title should show graph density and also the number of nodes
        density = 2*G.number_of_edges() / (G.number_of_nodes() *(G.number_of_nodes() - 1))

        self.ax.set_title("\n\n\n"+self.animal_id+",  "+self.session_ids[self.session_id]+'\n' 
                    +"# of cells: "+ str(corr_matrix.shape[0])+'\n'
                        +'Graph density: '+str(np.round(density,3))+'\n'
                        +'Number of nodes: '+str(G.number_of_nodes())+'\n'
                        +'Number of edges: '+str(G.number_of_edges())+'\n'
                        +'Number of isolates: '+str(nx.number_of_isolates(G)),
                        fontsize=14)


    def plot_reward_centered_traces_ROIs_all_days(self):


        #
        clrs = ['blue','lightblue','red','pink']
        clrs2 = ['black','green','magenta','orange']

        # make a viridis based discrete colormap
        import matplotlib.colors as colors
        cmap = plt.cm.viridis(np.linspace(0,1,len(self.session_ids)-1))
        
        
        #
        plt.figure()
        for cell_id in range(4):
            #
            ax=plt.subplot(2,2,cell_id+1)

            # plot title about cell_id
            ax.set_title('Cells matching ROI: '+str(cell_id), fontsize=14)

            
            # loop over all the session_ids
            ctr_session = 0
            for session_id in range(1,len(self.session_ids),1):

                # get the roi for the session      
                closest_ROI_cells = self.session_contours[session_id]
                #print ("closest_ROI_cells: ", closest_ROI_cells)

                # load reward centered traces for the session
                fname = os.path.join(
                                    self.root_dir,
                                    self.animal_id,
                                    self.session_ids[session_id],
                                    'reward_centered_traces_rewardwindow_'+str(self.reward_window)+'.npy')

                reward_centered_traces = np.load(fname, allow_pickle=True)
        
                # plot the mean and std for all the cells spacing them out by 3
                t = np.arange(reward_centered_traces.shape[2])/30-(self.reward_window/30)

                #
                cell_traces = []
                for k in range(reward_centered_traces.shape[1]):


                    # loop over all the closest_ROI_cells and check if k is matching
                    for i in range(len(closest_ROI_cells)):
                        matching_mask = closest_ROI_cells[i][0]
                        matching_cell = closest_ROI_cells[i][1]
                        
                        #print ("cell_id: ", cell_id, 'k: ', k, ' matching_cell: ', matching_cell, ' matching_mask: ', matching_mask)
                        if k == matching_cell and matching_mask == cell_id:

                            # grab the data                            #
                            temp = reward_centered_traces[:,k]

                            # subtract the median from each trace
                            # compute f0 as the median of the stacked trace
                            #f0 = np.median(temp.flatten())
                            #print ("f0: ", f0)
                            #temp = (temp-f0)/f0*100

                            #
                            #import scipy
                            #temp_mean = scipy.stats.mode(temp,0)
                            
                            #print ("temp_mean: ", temp_mean.shape)

                            # compute the std of temp in the first axis
                            temp_mean = np.mean(temp,0)
                            
                            cell_traces.append(temp_mean)
                            
                            


                # 
                cell_traces = np.array(cell_traces)

                # pick the cell trace with the highest ptp peak to peak
                #print ("cell_traces: ", cell_traces.shape)

                if cell_traces.shape[0]==0:
                    continue
               
                elif cell_traces.shape[0]==1:
                    cell_trace = cell_traces[0]

                elif cell_traces.shape[0]>1:
                    ptps = np.ptp(cell_traces,1)

                    idx = np.argmax(ptps)
                    cell_trace = cell_traces[idx]


                # plot legend only if cell_id ==0:
                if cell_id==0:
                    plt.plot(t,cell_trace,
                            linewidth=3,
                            c=cmap[ctr_session],
                            label=self.session_ids[session_id]+ ", # rew: "+str(reward_centered_traces[:,k].shape[0]))
                else:
                    plt.plot(t,cell_trace,
                            linewidth=3,
                            c=cmap[ctr_session],
                            )


                ctr_session+=1


            # plot horizontal line at 0
            plt.plot(t, temp_mean*0,'--',
                    c='grey')    
            
            # plot vertical line at 0
            #plt.plot([0,0],[-50,100],'--',
            #        c='grey',
            #        linewidth=3)

            # 
            #plt.plot([0,0],[-.50,100],'--',
            #         c='grey',
            #         linewidth=3)
            #
            plt.ylabel("DFF")
            #
            plt.legend()

        # label y axis using ylabels 
        #plt.yticks(ylabels, np.arange(temp.shape[0]))
        
        #
        plt.suptitle(self.animal_id)
        plt.show()

    #
    def plot_reward_centered_traces_ROIs_only(self):

        #
        clrs = ['blue','lightblue','red','pink']
        clrs2 = ['black','green','magenta','orange','yellow','cyan','purple','brown']


        # plot the mean and std for all the cells spacing them out by 3
        t = np.arange(self.reward_centered_traces.shape[2])/30-(self.reward_window/30)

        #
        closest_ROI_cells = self.session_contours[self.session_id]
        print ("closest_ROI_cells: ", closest_ROI_cells)
    
        #
        linetypes = ['solid',':','--','dashdot']

        plt.figure()
        scale = 100
        for cell_id in range(4):
            #
            ax=plt.subplot(2,2,cell_id+1)
            
            # plot title about cell_id
            ax.set_title('Cells matching ROI: '+str(cell_id), fontsize=14)


            #
            ctr = 0
            for k in range(self.reward_centered_traces.shape[1]):

                #
                temp = self.reward_centered_traces[:,k]

                # subtract the median from each trace
                #temp = temp - np.median(temp,1)[:,np.newaxis]
                #print ("temp: ", temp.shape)
                temp_mean = np.mean(temp,0)

                # compute the std of temp in the first axis
                #temp_mean = np.mean(temp,0)

                # loop over all the closest_ROI_cells and check if k is matching
                for i in range(len(closest_ROI_cells)):
                    matching_mask = closest_ROI_cells[i][0]
                    matching_cell = closest_ROI_cells[i][1]
                    
                    #print ("cell_id: ", cell_id, 'k: ', k, ' matching_cell: ', matching_cell, ' matching_mask: ', matching_mask)
                    if k == matching_cell and matching_mask == cell_id:
                    
                        # if k is the first two values in cell_ids plot using blue
                        plt.plot(t,temp_mean,
                                linewidth=3,
                                #linestyle=linetypes[ctr],
                               # alpha=0.5,
                                c=clrs2[ctr],
                                label=str(matching_cell))

                        ctr+=1
                
            # plot horizontal line at 0
            plt.plot(t, temp_mean*0,'--',
                    c='grey')    
            
            # plot vertical line at 0
            #plt.plot([0,0],[-100,500],'--',
            #        c='grey',
            #        linewidth=3)

            # 
            #plt.plot([0,0],[-100,500],'--',
            #         c='grey',
            #         linewidth=3)

            #
            plt.legend()

        # label y axis using ylabels 
        #plt.yticks(ylabels, np.arange(temp.shape[0]))
        plt.xlabel('Time (s)',fontsize=20)

        #
        plt.suptitle(self.session_ids[self.session_id])
        plt.show()


    #
    def plot_reward_centered_traces(self):
        
        #
        clrs = ['blue','lightblue','red','pink']

        # load reward centered traces for the session
        fname = os.path.join(
                            self.root_dir,
                            self.animal_id,
                            self.session_ids[self.session_id],
                            'reward_centered_traces_rewardwindow_'+str(self.reward_window)+'.npy')
        #
        fname_shuffled = os.path.join(
                            self.root_dir,
                            self.animal_id,
                            self.session_ids[self.session_id],
                            'reward_centered_traces_rewardwindow_'+str(self.reward_window)+'_shuffled.npy')

        #
        self.reward_centered_traces = np.load(fname, allow_pickle=True)
        self.reward_centered_traces_shuffled = np.load(fname_shuffled, allow_pickle=True)

        # compute the mean of temp in the first axis
        temp = np.mean(self.reward_centered_traces,0)
        temp_shuffled = np.mean(self.reward_centered_traces_shuffled,0)

        # compute the std of temp in the first axis
        #temp_std = np.std(self.reward_centered_traces,0)

        # plot the mean and std for all the cells spacing them out by 3
        t = np.arange(temp.shape[1])/30-(self.reward_window/30)

        #
        closest_ROI_cells = self.session_contours[self.session_id]
        print ("closest_ROI_cells: ", closest_ROI_cells)
    
        #
        # looping over all suite2p cells
        traces = []
        traces_shuffled = []
        clrs2 = []
        print ("loopin gover all suite2p cells: ", temp.shape[0])
        for k in range(temp.shape[0]):
            
            #
            temp2 = temp[k]

            #
            temp2 = temp2 - np.median(temp2)

            # loop over all the closest_ROI_cells and check if k is matching
            traces.append(temp2)

            #
            temp2_shuffled = temp_shuffled[k]
            temp2_shuffled = temp2_shuffled - np.median(temp2_shuffled)
            traces_shuffled.append(temp2_shuffled)
            
            #
            found_match = False
            for i in range(len(closest_ROI_cells)):
                matching_mask = closest_ROI_cells[i][0]
                matching_cell = closest_ROI_cells[i][1]

                if k == matching_cell:
                    print ('k: ', k, ' matching_cell: ', matching_cell, ' matching_mask: ', matching_mask)
                    clrs2.append(clrs[matching_mask])
                    found_match = True
                    break
 
            if found_match == False:
                clrs2.append('black')

        #
        traces = np.vstack(traces)
        traces_shuffled = np.vstack(traces_shuffled)
        clrs2 = np.vstack(clrs2)
        cell_ids = np.arange(traces.shape[0])

        #
        if self.sort_by_peak:
            peak_idxs = np.argmax(traces,1)
            
            # sort by location of ptp
            sorted_idx = np.argsort(peak_idxs)[::-1]

            # sort traces
            traces = traces[sorted_idx]
            clrs2 = clrs2[sorted_idx]
            cell_ids = cell_ids[sorted_idx]

            # same for traces_shuffled
            peak_idxs = np.argmax(traces_shuffled,1)
            sorted_idx = np.argsort(peak_idxs)
            traces_shuffled = traces_shuffled[sorted_idx]
            

        ###################################
        plt.figure()
        ax=plt.subplot(131)
        plt.ylabel("Neuron ID")
        # plotthe traces
        ctr = 0
        labels = []
        img = []
        img_shuffled = []
        for k in range(traces.shape[0]):
            
            # grab the shuffled traces first 
            temp4 = traces_shuffled[k]
            max = np.max(np.abs(temp4))
            if max >= self.min_ca_peak:
                img_shuffled.append(temp4)

            #
            temp3 = traces[k]
            max = np.max(np.abs(temp3))

            #
            if max < self.min_ca_peak:
                continue

            #
            img.append(temp3)

            # 
            plt.plot(t,temp3+ctr*self.scale,
                    linewidth=3,
                    c=clrs2[k][0])
            
            # plot horizontal line at temp3 = 0
            plt.plot(t, temp3*0+ctr*self.scale,'--',
                    c='grey',alpha=0.5)
            

            # print colors
            if clrs2[k][0]!='black':
                print ('k: ', k, ' clrs2[k][0]: ', clrs2[k][0])
            
            # plot the peak using peak_idx
            peak_idx = np.argmax(temp3)
            plt.plot(t[peak_idx],temp3[peak_idx]+ctr*self.scale,'o',
                    c='red')   

            #
            labels.append(cell_ids[k])

            #            
            ctr+=1
        
        # plot a vertical line at 0
        plt.plot([0,0],[0,ctr*self.scale],'--',c='grey',linewidth=3, alpha=0.2)


        #
        img = np.vstack(img)[::-1]

        #
        img_shuffled = np.vstack(img_shuffled)


        # 
        #plt.plot([0,0],
        #         [0,(ctr+1)*self.scale],
        #         'r--',linewidth=3)
        
        plt.title("# of cells: "+str(img.shape[0]))

        # label y axis using ylabels 
        plt.yticks(np.arange(ctr)*self.scale, labels)
        
        plt.xlabel('Time (s)',fontsize=20)
        plt.suptitle(self.animal_id+' '+self.session_ids[self.session_id]+"\n # rewards: "
                     +str(self.reward_centered_traces.shape[0]),fontsize=20)
        # limit plot to size of image
        plt.xlim(t[0],t[-1])
        plt.ylim(0,ctr*self.scale)

        ##########################################
        ##########################################
        ##########################################
        ax=plt.subplot(132)
        plt.imshow(img,
                       aspect='auto',
                       cmap='jet',
                       interpolation='none',
                       extent=[t[0],t[-1],0,img.shape[0]],
                       vmin=-self.vmax,
                       vmax=self.vmax
                       )
        # plot a vertical line at 0
        plt.plot([0,0],[0,img.shape[0]],'--',c='grey',linewidth=3, alpha=0.2)

        #plt.colorbar()
        plt.xlabel('Time (s)',fontsize=20)

        ##########################################
        ##########################################
        ##########################################
        ax=plt.subplot(133)
        plt.title('Shuffled')
        # plotthe traces shuffled
        if img_shuffled.shape[0]<=img.shape[0]:
            img_c = img.copy()

            img_c[:] = np.nan
            img_c[-img_shuffled.shape[0]:,] = img_shuffled
        else:
            img_c = img_shuffled.copy()
        #
        print (img_c.shape)
        plt.imshow(img_c,
                       aspect='auto',
                       cmap='jet',
                       interpolation='none',
                       extent=[t[0],t[-1],0,img_c.shape[0]],
                       vmin=-self.vmax,
                       vmax=self.vmax
                       )
        # plot a vertical line at 0
        plt.plot([0,0],[0,img_c.shape[0]],'--',c='grey',linewidth=3, alpha=0.2)

        # plot colorbar with label "DFF"
        plt.colorbar(label="DFF")


        plt.xlabel('Time (s)',fontsize=20)


        plt.show()

    #
    def plot_tone_triggered_ca(self):

        tone_ids = np.arange(2,self.unique_tones.shape[0])
        print ("tone_ids: ", tone_ids)

        #
        window = 10*30
        min_space_starts = window
        vmax = self.vmax

        #
        plt.figure()

        for tone_id in tone_ids:

            #
            idx = np.where(np.int32(self.tones)==self.unique_tones[tone_id])[0]
            n_tones = idx.shape[0]
            print ("tone: ", self.unique_tones[tone_id], " # of tones: ", n_tones)

            # find the starts in each sequence of idx values
            starts = []
            for k in range(idx.shape[0]-1):
                if idx[k+1]-idx[k]>1:
                    starts.append(idx[k+1])
            starts = np.array(starts) #+3000

            # # plot the calcium events for the first tone
            t = np.arange(-window,window)/30

            # remove starts that are too close to each other
            ss = []
            ss.append(starts[0])
            for k in range(1,starts.shape[0]):
                # skip k if too close to previous
                if (starts[k]-starts[k-1])>min_space_starts:
                    ss.append(starts[k])

            starts = np.array(ss)

            ######################################
            # make image stack
            img = []
            img_shuffled = []
            for cell_id in range(self.sessions[self.session_id].F.shape[0]):
                traces = []
                traces_shuffled = []
                for s in starts:
                    #temp = c.sessions[session_id].F_upphase_bin[cell_id,s-window:s+window]
                    temp = self.sessions[self.session_id].F_filtered[cell_id,s-window:s+window]
                    if temp.shape[0]==2*window:
                        traces.append(temp)
                    
                    # same but shuffle the tone times
                    s_shuffled = np.random.randint(300,self.sessions[self.session_id].F.shape[1]-2*window)
                    temp = self.sessions[self.session_id].F_filtered[cell_id,s_shuffled-window:s_shuffled+window]
                    if temp.shape[0]==2*window:
                        traces_shuffled.append(temp)

                # plot the average of traces
                img.append(np.mean(traces,axis=0))
                img_shuffled.append(np.mean(traces_shuffled,axis=0))

            # Reorder by peak
            img = np.array(img)

            # order img by ptp
            if False:
                ptp = np.max(img,axis=1)
                idx = np.argsort(ptp)[::-1]
                img = img[idx,:]
            #
            good_cell_ids = np.where(np.max(np.abs(img),axis=1)>self.min_ca_val)[0]
            img = img[good_cell_ids]

            # also order them by location of their max/min
            #max_val = np.argmax(img,axis=1)
            max_val = np.max(img,axis=1)
            idx = np.argsort(max_val)[::-1]
            img = img[idx,:]
            good_cell_ids = good_cell_ids[idx]


            # same for img shuffled
            img_shuffled = np.array(img_shuffled)
            if False:
                ptp = np.max(img_shuffled,axis=1)
                idx = np.argsort(ptp)[::-1]
                img_shuffled = img_shuffled[idx,:]
            
            #
            idx = np.where(np.max(np.abs(img_shuffled),axis=1)>self.min_ca_val)[0]
            img_shuffled = img_shuffled[idx]

            max_val = np.max(img_shuffled,axis=1)
            idx = np.argsort(max_val)[::-1]
            img_shuffled = img_shuffled[idx,:]


            ######################################
            ######################################
            ######################################
            top_n = 1000

            #
            ax=plt.subplot(2,15,tone_id-1)
            img = img[:top_n]

            # arrange img by location of max value argument
            if False:
                idx = np.argmax(img,axis=1)
                idx = np.argsort(idx)
                img = img[idx,:]

            plt.imshow(img,
                    aspect='auto',
                    extent=[t[0],t[-1],0,img.shape[0]],
                    interpolation='none',cmap=self.cmap,
                    vmin=self.vmin,
                    vmax=self.vmax
                    )
            
            # plot yticks as good_cell_ids
            plt.yticks(np.arange(img.shape[0])+0.5,good_cell_ids)
            

            if tone_id == 2:
                plt.ylabel("Cell #")
            #else:
            #    plt.yticks([])

            plt.title(str(self.unique_tones[tone_id])+"hz" + ", #: "+str(len(starts)))

            # draw verical line at tone onset
            plt.plot([0,0],[0,img.shape[0]],'k--')
            plt.xlabel("Time (s)")

            #####################################
            # sem for shuffled
            img_shuffled = img_shuffled[:top_n]
            ax=plt.subplot(2,15,tone_id-1+15)
            if tone_id == 2:
                plt.ylabel("Cell # (SHUFFLED)")
            else:
                plt.yticks([])

            #
            plt.imshow(img_shuffled,
                       aspect='auto'
                       ,extent=[t[0],t[-1],0,img_shuffled.shape[0]],
                        interpolation='none',cmap=self.cmap,
                        vmin=self.vmin,
                        vmax=self.vmax)
            


            # draw verical line at tone onset
            plt.plot([0,0],[0,img_shuffled.shape[0]],'k--')


            plt.xlabel("Time (s)")

        plt.colorbar()
        #plt.plot(t,np.mean(traces,axis=0))
        #plt.plot(c.sessions[session_id].F[:,idx])
        #plt.title("Tone: "+str(tone))

        # plot suptitle with animal id, session id and number of rewards
        plt.suptitle("Animal: "+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(self.n_rewards))

        plt.show()


    #
    def get_tone_triggered_ca_single_cell(self):

        tone_ids = np.arange(2,self.unique_tones.shape[0])
        print ("tone_ids: ", tone_ids)

        #
        window = self.window
        min_space_starts = self.window
        vmax = self.vmax

        #
        plt.figure()
        # make image stack
        img = []
        img_shuffled = []
        traces = []
        traces_shuffled = []
        cell_id = self.cell_id
        for tone_id in tone_ids:

            #
            idx = np.where(np.int32(self.tones)==self.unique_tones[tone_id])[0]
            n_tones = idx.shape[0]
            #print ("tone: ", self.unique_tones[tone_id], " # of tones: ", n_tones)

            # find the starts in each sequence of idx values
            starts = []
            for k in range(idx.shape[0]-1):
                if idx[k+1]-idx[k]>1:
                    starts.append(idx[k+1])
            starts = np.array(starts) #+3000

            # # plot the calcium events for the first tone
            t = np.arange(-window,window)/30

            # remove starts that are too close to each other
            ss = []
            ss.append(starts[0])
            for k in range(1,starts.shape[0]):
                # skip k if too close to previous
                if (starts[k]-starts[k-1])>min_space_starts:
                    ss.append(starts[k])

            starts = np.array(ss)

            ######################################

            for s in starts:
                #temp = c.sessions[session_id].F_upphase_bin[cell_id,s-window:s+window]
                temp = self.sessions[self.session_id].F_filtered[cell_id,s-window:s+window]
                if temp.shape[0]==2*window:
                    traces.append(temp)
                
                # same but shuffle the tone times
                s_shuffled = np.random.randint(300,self.sessions[self.session_id].F.shape[1]-2*window)
                temp = self.sessions[self.session_id].F_filtered[cell_id,s_shuffled-window:s_shuffled+window]
                if temp.shape[0]==2*window:
                    traces_shuffled.append(temp)

            # plot the average of traces
            img.append(np.mean(traces,axis=0))
            img_shuffled.append(np.mean(traces_shuffled,axis=0))

        # Reorder by peak
        img = np.array(img)

        # # order img by ptp
        # if False:
        #     ptp = np.max(img,axis=1)
        #     idx = np.argsort(ptp)[::-1]
        #     img = img[idx,:]
        # #
        # good_cell_ids = np.where(np.max(np.abs(img),axis=1)>self.min_ca_val)[0]
        # img = img[good_cell_ids]

        # # also order them by location of their max/min
        # #max_val = np.argmax(img,axis=1)
        # max_val = np.max(img,axis=1)
        # idx = np.argsort(max_val)[::-1]
        # img = img[idx,:]
        # good_cell_ids = good_cell_ids[idx]


        # same for img shuffled
        img_shuffled = np.array(img_shuffled)
        if False:
            ptp = np.max(img_shuffled,axis=1)
            idx = np.argsort(ptp)[::-1]
            img_shuffled = img_shuffled[idx,:]
        
        # #
        # idx = np.where(np.max(np.abs(img_shuffled),axis=1)>self.min_ca_val)[0]
        # img_shuffled = img_shuffled[idx]

        # max_val = np.max(img_shuffled,axis=1)
        # idx = np.argsort(max_val)[::-1]
        # img_shuffled = img_shuffled[idx,:]


        ######################################
        ######################################
        ######################################
        top_n = 1000

        #
        ax=plt.subplot(1,2,1)
        img = img[:top_n]

        # arrange img by location of max value argument
        if False:
            idx = np.argmax(img,axis=1)
            idx = np.argsort(idx)
            img = img[idx,:]

        plt.imshow(img,
                aspect='auto',
                extent=[t[0],t[-1],0,img.shape[0]],
                interpolation='none',cmap=self.cmap,
                vmin=self.vmin,
                vmax=self.vmax
                )
        
        # plot yticks as unique tones
        plt.yticks(np.arange(img.shape[0])+0.5,self.unique_tones[tone_ids])
        #plt.yticks(np.arange(img.shape[0])+0.5,good_cell_ids)
        

        # if tone_id == 2:
        #     plt.ylabel("Cell #")
        # #else:
        # #    plt.yticks([])

        # draw verical line at tone onset
        plt.plot([0,0],[0,img.shape[0]],'k--')
        plt.xlabel("Time (s)")

        #####################################
        # sem for shuffled
        img_shuffled = img_shuffled[:top_n]
        ax=plt.subplot(1,2,2)
        if tone_id == 2:
            plt.ylabel("Cell # (SHUFFLED)")
        else:
            plt.yticks([])

        #
        plt.imshow(img_shuffled,
                    aspect='auto'
                    ,extent=[t[0],t[-1],0,img_shuffled.shape[0]],
                    interpolation='none',cmap=self.cmap,
                    vmin=self.vmin,
                    vmax=self.vmax)
        
        # draw verical line at tone onset
        plt.plot([0,0],[0,img_shuffled.shape[0]],'k--')

        #
        plt.xlabel("Time (s)")

        #
        plt.colorbar()

        # plot suptitle with animal id, session id and number of rewards
        plt.suptitle("Cell_id" + str(self.cell_id )+ "\nAnimal: "+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(self.n_rewards)
                    + "\n"+
                    str(self.unique_tones[tone_id])+"hz" + ", #: "+str(len(starts)))

        plt.show()

        plt.savefig(os.path.join(self.root_dir,
                                 self.animal_id,
                                 self.session_ids[self.session_id],
                                 'cells',
                                 str(cell_id)+'.png'))
        plt.close()

    #
    def get_mean_warped_trace(self):    

        #
        if self.burst_sign == 1:
            self.compute_mean_warped_trace_positive()
        else:
            self.compute_mean_warped_trace_negative()

    #
    def compute_mean_warped_trace_negative(self):
        
        #
        fps = 30

        # plot the ensemble state for each of the starts_ends separating into 2 plots by sign
        if self.plotting:
            plt.figure(figsize=(10,5))

            ax1 = plt.subplot(1,2,1)
            ax1.set_title('negative')
            plt.xlabel('time (s)')
            # plot the thresh value
            ax1.axhline(y=np.mean(self.thresh), color='grey', linestyle='--')

            #
            ax2 = plt.subplot(1,2,2)
            ax2.set_title('warped')
            plt.xlabel('time (s)')

            # make a colormap of fixed colors in viridis for 10 values
            # count # of positive and negative vals in signs
            neg = np.sum(self.signs == -1)            
            cmap_r = plt.cm.viridis(np.linspace(0, 1, neg))

        #
        ctr_neg = 0
        self.snippets = []
        self.snippets_warped = []
        print ("TODO: remove the worst negative bursts from the list...")
        for k in range(self.starts_ends.shape[0]):
            temp = self.e_state[self.starts_ends[k,0]:self.starts_ends[k,1]]

            # if time warp
            #if self.time_warp == True:           # this just stretches the snippet to be the same length - 30 timesteps
            #    temp = resample_snippet(temp)

            if self.signs[k] == -1:
        
                #
                self.snippets.append(temp)

                # 
                temp2 = resample_snippet(temp.copy())

                #                         
                self.snippets_warped.append(temp2)

            #
                if self.plotting:
                    ax1.plot(np.arange(temp.shape[0])/fps, temp,
                        c=cmap_r[ctr_neg])

                    #
                    ax2.plot(np.arange(temp2.shape[0])/fps, 
                             temp2, c=cmap_r[ctr_neg])
                ctr_neg += 1

        #
        # plot average of temps
        self.snippets_warped = np.array(self.snippets_warped)
        mean = np.mean(self.snippets_warped,axis=0)
        self.mean = mean

        #
        self.snippets = np.array(self.snippets, dtype=object)
        if self.plotting:
            ax2.plot(np.arange(mean.shape[0])/30, mean ,
                    c='red', linewidth=3)

        #
        self.cmap_r = cmap_r

        # plot the thresh value
        if self.plotting:   
            plt.show()


    #
    def compute_mean_warped_trace_positive(self):

        # plot the ensemble state for each of the starts_ends separating into 2 plots by sign
        if self.plotting:
            plt.figure(figsize=(20,5))

            ax1 = plt.subplot(1,4,1)
            ax1.set_title('positive')
            plt.xlabel('time (s)')
            # plot the thresh value
            ax1.axhline(y=np.mean(self.thresh), color='grey', linestyle='--')

            #
            ax2 = plt.subplot(1,4,2)
            ax2.set_title('negative')
            plt.xlabel('time (s)')

            #
            ax3 = plt.subplot(1,4,3)
            ax3.set_title('good snippets (> threshold; start from 0.5 max)')
            ax3.axhline(y=np.mean(self.thresh), color='grey', linestyle='--')
            plt.xlabel('Time (warped)')

            #
            ax4 = plt.subplot(1,4,4)
            ax4.set_title('snippets only to threshold')
            ax4.axhline(y=np.mean(self.thresh), color='grey', linestyle='--')
            plt.xlabel('Time (warped)')
            ax4.set_ylim(top=1)

            #
            cmap_r = plt.cm.viridis(np.linspace(0, 1, len(self.ids_snippets_rewarded)))


        # count # of positive and negative vals in signs
        pos = np.sum(self.signs == 1)
        neg = np.sum(self.signs == -1)

        # make a colormap of fixed colors in viridis for 10 values
        cmap_pos = plt.cm.viridis(np.linspace(0, 1, pos))
        cmap_neg = plt.cm.viridis(np.linspace(0, 1, neg))

        #
        ctr_pos = 0
        ctr_neg = 0
        # this just counts for purposes of making a nice list below
        for k in range(self.starts_ends.shape[0]):
            temp = self.e_state[self.starts_ends[k,0]:self.starts_ends[k,1]]

            # if time warp
            #if self.time_warp == True:
            #    temp = resample_snippet(temp)

            if self.signs[k] == 1:
                if self.plotting:
                    ax1.plot(np.arange(temp.shape[0])/30, temp,
                        c=cmap_pos[ctr_pos])
                ctr_pos += 1
            else:
                if self.plotting:
                    ax2.plot(np.arange(temp.shape[0])/30, temp,
                        c=cmap_neg[ctr_neg])
                ctr_neg += 1

        
        #
        self.snippets = []
        self.snippets_warped = []
        self.snippets_to_thresh = []
        self.snippets_to_thresh_warped = []
        for k in range(len(self.ids_snippets_rewarded)):
            idx = self.ids_snippets_rewarded[k]
            temp1 = self.e_state[self.starts_ends[idx,0]:self.starts_ends[idx,1]]

            #
            self.snippets.append(temp1)

            #
            temp2 = resample_snippet(temp1.copy())
            #print ("temp: ", temp1.shape, " temp2.shape: ", temp2.shape)
                                      
            self.snippets_warped.append(temp2)

            # find the index up to where temp is below thresh
            temp_thresh = np.mean(self.thresh)
            idx = np.where(temp1<temp_thresh)[0]
            snippet_to_threshold = temp1[idx]
            #print ("snippet_to_threshold: ", snippet_to_threshold.shape)
            if snippet_to_threshold.shape[0]==0:
                continue

            self.snippets_to_thresh.append(snippet_to_threshold)

            #
            snippet_to_threshold_warped = resample_snippet(snippet_to_threshold.copy())
            self.snippets_to_thresh_warped.append(snippet_to_threshold_warped)

            #
            if self.plotting:
                ax3.plot(np.arange(temp2.shape[0])/30, temp2, c=cmap_r[k])

                #
                ax4.plot(np.arange(snippet_to_threshold_warped.shape[0])/30, 
                         snippet_to_threshold_warped, c=cmap_r[k])

        # plot average of temps
        self.snippets_warped = np.array(self.snippets_warped)
        print ("snippets_warped: ", self.snippets_warped.shape)
        mean = np.mean(self.snippets_warped,axis=0)
        self.mean = mean

        #
        self.snippets_to_thresh = np.array(self.snippets_to_thresh)
        self.snippets_to_thresh_warped = np.array(self.snippets_to_thresh_warped)
        self.mean_to_thresh = np.mean(self.snippets_to_thresh_warped,axis=0)

        #
        self.snippets = np.array(self.snippets, dtype=object)
        #print ("temps: ", temps.shape)
        if self.plotting:
            ax3.plot(np.arange(mean.shape[0])/30, mean ,
                    c='red', linewidth=3)

            #
            ax4.plot(np.arange(self.mean_to_thresh.shape[0])/30, 
                     self.mean_to_thresh ,
                    c='red', linewidth=3)

        # plot the thresh value
        if self.plotting:   
            plt.show()

    #
    def get_tone_triggered_ca_single_cell_static(self):

        tone_ids = np.arange(2,self.unique_tones.shape[0])
        #print ("tone_ids: ", tone_ids)

        #
        if self.F_type=='upphase':
            F = self.sessions[self.session_id].F_upphase_bin
        elif self.F_type == 'F_filtered':
            F = self.sessions[self.session_id].F_filtered

        # 
        vector = []
        for tone_id in tone_ids:
            
            #
            idx = np.where(np.int32(self.tones)==self.unique_tones[tone_id])[0]

            temp = np.zeros((self.tones.shape[0]))
            temp[idx] = 1

            # FIND BEGINNIGN AND ENDS OF tones
            from scipy.signal import chirp, find_peaks, peak_widths
            peaks, _ = find_peaks(temp)  # middle of the pluse/peak
            widths, heights, starts, ends = peak_widths(temp, peaks)
            starts = np.int32(starts)
            ends = np.int32(ends)

            #print ("duration of tones: ", widths/30, "sec")

            # index into temp between tone starts and ends
            cell_id_ca = []
            for k in range(len(starts)):
                temp = F[self.cell_id][starts[k]:ends[k]]
                cell_id_ca.append(temp)
            cell_id_ca = np.hstack(cell_id_ca)

            #
            #print ("cell_id_ca: ", cell_id_ca.shape)
            vector.append(np.mean(cell_id_ca,0))
       
        ######################################
        ######################################
        ######################################
        #
        vector = np.vstack(vector)
        #plt.figure()
        #plt.imshow(vector,
        #           aspect='auto',
        #           extent = [0,1,0,len(tone_ids)])
        plt.plot(vector,
                  )
        
        # plot xticks as unique tones and rotated 45 degrees
        plt.xticks(np.arange(len(tone_ids)),self.unique_tones[tone_ids],rotation=45)

        
        #
       # plt.show()
#


    #
    def get_ca_triggered_tone_all_cells(self):
        
        #
        
        #
        window = self.window
        vmax = self.vmax
        shuffle = self.shuffle

        #
        #reward_times = self.reward_times

        #


        # submit a particular type of trigger: tone_starts vs. all
        if self.tone_type == 'starts':
            tone_starts = self.tone_starts
        elif self.tone_type == 'all':
            tone_starts = self.tone_all

            #print ("selected all: ", tone_starts)

        ##
        if self.F_type=='upphase':
            F = self.sessions[self.session_id].F_upphase_bin
        elif self.F_type == 'F_filtered':
            F = self.sessions[self.session_id].F_filtered

        ######################################
        cell_ids = np.arange(self.sessions[self.session_id].F.shape[0])
        #cell_ids = np.arange(4)
        imgs = parmap.map(get_img_tone,
                          cell_ids,
                          F,
                          window,
                          self.tones,
                          shuffle,
                          self.smooth,
                          self.tone_type,
                          self.remove_base_tone,
                          pm_processes=16,
                          pm_pbar=True)
       
        ######################################
        ######################################
        ######################################

        #imgs = np.vstack(imgs)

        ctr=0
        k=0
        plt.figure(figsize=(12,12))
        while True:
            if ctr%100==0 and ctr>0:

                self.savefig_tones(ctr)

                plt.figure(figsize=(12,12))
                k=0

            ###################################### 
            ax=plt.subplot(10,10,k+1)
            if self.vmax==None:
                vmax = np.max(imgs[ctr])
                vmin = -vmax
                plt.imshow(imgs[ctr],
                        aspect='auto',
                        interpolation='none',cmap=self.cmap,
                        vmin=vmin,
                        vmax=vmax
                        #vmax=self.vmax
                        )
    
            else:
                #if self.vmax < np.max(imgs[ctr]):
                #    vmax = np.max(imgs[ctr])
                #else:
                vmax = self.vmax
                vmin = -vmax

                plt.imshow(imgs[ctr],
                        aspect='auto',
                        interpolation='none',cmap=self.cmap,
                        vmin=-vmax,
                        vmax=vmax
                        )
                
            plt.xticks([])
            plt.yticks([])
            #plt.ylim(-0.5,imgs[k].shape[0]-0.5)
            # plot y ticks as tones values
            if k == 0:
                plt.ylabel("Tone value (hz)")
                
                # make a list of strings that contains the tone values and the number of times they appear
                tone_labels = np.zeros((self.unique_tones.shape[0]-2,2))
                tone_labels[:,0] = self.unique_tones[2:]
                tone_labels[:,1] = self.counts[2:]

                ls = []
                for q in range(2, self.unique_tones.shape[0],1):
                    x = self.unique_tones[q]
                    y = tone_starts[q-2].shape[0]
                    #print ("x: ", x, " y: ", y  )
                    ls.append([str(x)+"hz (# "+str(y)+")"])

                ls = np.array(ls)
                #
                #print (np.arange(imgs[k].shape[0]))
                #print ("tone_labels: ", tone_labels)
                #plt.yticks(np.arange(imgs[k].shape[0]),
                #           ls[:,0],
                #            fontsize=5)
                

                # plot also time ticks on x axis but only the first and last and zero
                #plt.xticks([0,imgs[k].shape[1]/2,imgs[k].shape[1]-1],
                #            [str(-window/30),0,str(window/30)],
                #            fontsize=5)
                
                # print xlabel also but with padding so it's close to the image
                plt.xlabel("Time (s)",fontsize=5,labelpad=0.1)

                # plot colorbar with label "DFF" and without shrinking it and with fontsize 5 and with only 3 ticks
                clb = plt.colorbar(shrink=1, ticks=[-vmin,0,vmax])
                clb.ax.tick_params(labelsize=5)

                # put the colorbar at the top of the plot
                clb.ax.set_title('DFF')

            # plot title with padding to be close to the image
            if k!=10:
                plt.title(str(ctr),pad=0.1)

            # plot a vertical line at 0
            #plt.plot([window,window],[0,imgs[k].shape[0]],'--',c='grey',linewidth=3, alpha=0.4)

            #####################
            ctr+=1
            k+=1

            # 
            if ctr == len(imgs):

                self.savefig_tones(ctr)

                break

  

    #
    def get_tone_triggered_ca_all_cells(self):
        
        #
        
        #
        window = self.window
        vmax = self.vmax
        shuffle = self.shuffle

        #
        #reward_times = self.reward_times

        #


        # submit a particular type of trigger: tone_starts vs. all
        if self.tone_type == 'starts':
            tone_starts = self.tone_starts
        elif self.tone_type == 'all':
            tone_starts = self.tone_all

            #print ("selected all: ", tone_starts)

        ##
        if self.F_type=='upphase':
            F = self.sessions[self.session_id].F_upphase_bin
        elif self.F_type == 'F_filtered':
            F = self.sessions[self.session_id].F_filtered

        ######################################
        cell_ids = np.arange(self.sessions[self.session_id].F.shape[0])
        imgs = parmap.map(get_img_ca,
                          cell_ids,
                          F,
                          window,
                          tone_starts,
                          shuffle,
                          self.smooth,
                          pm_processes=16,
                          pm_pbar=True)
       
        ######################################
        ######################################
        ######################################

        # do a 10 x 10 grid showing each element of imgs using imshow
        ctr=0
        k=0
        plt.figure(figsize=(12,12))
        while True:
            if ctr%100==0 and ctr>0:

                self.savefig(ctr)

                plt.figure(figsize=(12,12))
                k=0

            ###################################### 
            ax=plt.subplot(10,10,k+1)
            if self.vmax==None:
                vmax = np.max(imgs[ctr])
                vmin = -vmax
                plt.imshow(imgs[ctr],
                        aspect='auto',
                        interpolation='none',cmap=self.cmap,
                        vmin=vmin,
                        vmax=vmax
                        #vmax=self.vmax
                        )
    
            else:
                #if self.vmax < np.max(imgs[ctr]):
                #    vmax = np.max(imgs[ctr])
                #else:
                vmax = self.vmax
                vmin = -vmax

                plt.imshow(imgs[ctr],
                        aspect='auto',
                        interpolation='none',cmap=self.cmap,
                        vmin=-vmax,
                        vmax=vmax
                        )
                
            plt.xticks([])
            plt.yticks([])
            plt.ylim(-0.5,imgs[k].shape[0]-0.5)
            # plot y ticks as tones values
            if k == 0:
                plt.ylabel("Tone value (hz)")
                
                # make a list of strings that contains the tone values and the number of times they appear
                tone_labels = np.zeros((self.unique_tones.shape[0]-2,2))
                tone_labels[:,0] = self.unique_tones[2:]
                tone_labels[:,1] = self.counts[2:]

                ls = []
                for q in range(2, self.unique_tones.shape[0],1):
                    x = self.unique_tones[q]
                    y = tone_starts[q-2].shape[0]
                    #print ("x: ", x, " y: ", y  )
                    ls.append([str(x)+"hz (# "+str(y)+")"])

                ls = np.array(ls)
                #
                #print (np.arange(imgs[k].shape[0]))
                #print ("tone_labels: ", tone_labels)
                plt.yticks(np.arange(imgs[k].shape[0]),
                           ls[:,0],
                            fontsize=5)
                

                # plot also time ticks on x axis but only the first and last and zero
                plt.xticks([0,imgs[k].shape[1]/2,imgs[k].shape[1]-1],
                            [str(-window/30),0,str(window/30)],
                            fontsize=5)
                
                # print xlabel also but with padding so it's close to the image
                plt.xlabel("Time (s)",fontsize=5,labelpad=0.1)

                # plot colorbar with label "DFF" and without shrinking it and with fontsize 5 and with only 3 ticks
                clb = plt.colorbar(shrink=1, ticks=[-vmin,0,vmax])
                clb.ax.tick_params(labelsize=5)

                # put the colorbar at the top of the plot
                clb.ax.set_title('DFF')

            # plot title with padding to be close to the image
            if k!=10:
                plt.title(str(ctr),pad=0.1)

            # plot a vertical line at 0
            plt.plot([window,window],[0,imgs[k].shape[0]],'--',c='grey',linewidth=3, alpha=0.4)

            #####################
            ctr+=1
            k+=1

            # 
            if ctr == len(imgs):

                self.savefig(ctr)

                break


    def savefig_tones(self, ctr):   
        
        # plot suptitle with animal id, session id and number of rewards
        plt.suptitle("Animal: "+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(self.n_rewards)
                    #+ "\n"+
                    #str(self.unique_tones[tone_id])+"hz" + ", #: "+str(len(starts))
                    )

        plt.show()
        
        #
        if self.shuffle == False:
            fname_out = os.path.join(self.root_dir,
                                self.animal_id,
                                self.session_ids[self.session_id],
                                'all_tones_'+str(ctr)+"_"+
                                str(self.F_type)+
                                "_reward_lockout_"+str(self.remove_rewards)+    
                                "_window_"+str(self.window//30)+
                                "_tone_type_"+self.tone_type+
                                '.png')
        else:
            fname_out = os.path.join(self.root_dir,
                                self.animal_id,
                                self.session_ids[self.session_id],
                                'all_tones_shuffled_'+str(ctr)+
                                "_"+str(self.F_type)+
                                "_reward_lockout_"+str(self.remove_rewards)+                                
                                "_window_"+str(self.window//30)+
                                "_tone_type_"+self.tone_type+
                                '.png')
        #
        plt.savefig(fname_out, dpi=300)
        plt.close()
        

    def savefig(self, ctr):   
        
        # plot suptitle with animal id, session id and number of rewards
        plt.suptitle("Animal: "+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(self.n_rewards)
                    #+ "\n"+
                    #str(self.unique_tones[tone_id])+"hz" + ", #: "+str(len(starts))
                    )

        plt.show()
        
        #
        if self.shuffle == False:
            fname_out = os.path.join(self.root_dir,
                                self.animal_id,
                                self.session_ids[self.session_id],
                                'all_cells_'+str(ctr)+"_"+
                                str(self.F_type)+
                                "_reward_lockout_"+str(self.remove_rewards)+    
                                "_window_"+str(self.window//30)+
                                "_tone_type_"+self.tone_type+
                                '.png')
        else:
            fname_out = os.path.join(self.root_dir,
                                self.animal_id,
                                self.session_ids[self.session_id],
                                'all_cells_shuffled_'+str(ctr)+
                                "_"+str(self.F_type)+
                                "_reward_lockout_"+str(self.remove_rewards)+                                
                                "_window_"+str(self.window//30)+
                                "_tone_type_"+self.tone_type+
                                '.png')
        #
        plt.savefig(fname_out, dpi=300)
        plt.close()
        
    #
    def get_unique_tones(self, plotting=False):

        #       
        # get all unique values in tone and the number of times they appear
        self.unique_tones, self.counts = np.unique(np.int32(self.tones), return_counts=True)

        # make a list of all unique tone indices
        unique_tones_idx = []
        for k in range(self.unique_tones.shape[0]):
            idx = np.where(self.tones==self.unique_tones[k])[0]
            unique_tones_idx.append(idx)

        # delete values from tones that occurs less than 200 times
        for k in range(len(unique_tones_idx)):
            if unique_tones_idx[k].shape[0]<200:
                self.tones[unique_tones_idx[k]]=0

        # recompute unique tones based on what's left
        self.unique_tones, self.counts = np.unique(self.tones, return_counts=True)

        #
        self.tone_all = []
        tone_starts = []
        reward_times = self.reward_times
        tone_ids = np.arange(2,self.unique_tones.shape[0])
        min_space_starts = self.window
        for tone_id in tone_ids:

            #
            ##############################################################
            idx = np.where(np.int32(self.tones)==self.unique_tones[tone_id])[0]
            
            # exclude tone timesthat are too close to rewards
            if self.remove_rewards:
                ss = []
                for k in range(idx.shape[0]):
                    if np.min(np.abs(idx[k]-reward_times))>min_space_starts:
                        ss.append(idx[k])
                temp = np.array(ss)
                #print ("tone: ", self.unique_tones[tone_id], " # of tones: ", temp.shape[0])
                self.tone_all.append(temp)

            ##############################################################
            # find the starts in each sequence of idx values
            starts = []
            for k in range(idx.shape[0]-1):
                if idx[k+1]-idx[k]>1:
                    starts.append(idx[k+1])
            starts = np.array(starts) #+3000

            # # plot the calcium events for the first tone
            #t = np.arange(-window,window)/30

            # remove starts that are too close to each other
            ss = []
            ss.append(starts[0])
            for k in range(1,starts.shape[0]):
                # skip k if too close to previous
                if (starts[k]-starts[k-1])>min_space_starts:
                    ss.append(starts[k])

            starts = np.array(ss)

            # remove starts that are too close to rewards
            if self.remove_rewards:
                ss = []
                for k in range(starts.shape[0]):
                    if np.min(np.abs(starts[k]-reward_times))>min_space_starts:
                        ss.append(starts[k])
                starts = np.array(ss)

            # 
            tone_starts.append(starts)

        #
        self.tone_starts = tone_starts


        # make a histogram of the tone values
        if plotting:
            plt.figure()
            y = np.histogram(self.tones, bins=np.arange(1,16500,100))
            plt.bar(y[1][:-1],y[0], width=100)

            #
            plt.xlabel("Tone value (hz)")
            plt.ylabel("# of frames")

            #
            plt.suptitle(self.animal_id+ " - " + self.session_ids[1])

            plt.show()

        # load the calcium events for all cells
        #print ("# cells: ", self.sessions[self.session_id].F.shape)
       # print ("# cells: ", self.sessions[self.session_id].F_upphase_bin.shape)


    #
    def load_trials(self):
        # make an fname_out directory if not already present
        session_name = self.session_ids[self.session_id]

        # #load the n_rewards
        # fname = os.path.join(self.root_dir,
        #                         self.animal_id,
        #                         session_name,
        #                         'results.npz')
        
        # #
        # results = np.load(fname, allow_pickle=True)
        # rewards = results['reward_times'].T
        # idx = np.where(rewards[:,1]>0)[0]
        # self.n_rewards = idx.shape[0]
        # reward_times = rewards[idx,1]
        #print ("# of rewards: ", self.n_rewards, " rewar[:100])

        # Load the workbook
        wb = load_workbook(os.path.join(self.root_dir,
                                        self.animal_id,
                                        session_name,
                                        'results_fixed.xlsx'), read_only=False)

        # Access the worksheet
        ws = wb[wb.sheetnames[0]]  # assuming you want the first sheet, change as needed

        # get white noise state  
        rows = [3,4,7,8,9,10,11]          # white_noise, tone_state, ensemble_state, reward_state
        self.threshold_state = []       # 3rd column
        self.white_noise_state = []    # 4th column
        self.tone_state = []           # 7th
        self.ensemble_state = []       # 8th
        self.reward_state = []          # 9th
        self.trials = []               # 10th
        self.trials_rewards = []       # 11th  this is the flag which indicates if trial was rewarded
        for row in ws.iter_rows(min_row=2,values_only=False):
            self.threshold_state.append(row[rows[0]-1].value)
            self.white_noise_state.append(row[rows[1]-1].value)
            self.tone_state.append(row[rows[2]-1].value)
            self.ensemble_state.append(row[rows[3]-1].value)
            self.reward_state.append(row[rows[4]-1].value)
            self.trials.append(row[rows[5]-1].value)
            self.trials_rewards.append(row[rows[6]-1].value)

        # make all lists into arrays
        self.threshold_state = np.array(self.threshold_state)
        self.white_noise_state = np.array(self.white_noise_state)
        self.tone_state = np.array(self.tone_state)
        self.ensemble_state = np.array(self.ensemble_state)
        self.reward_state = np.array(self.reward_state)
        self.trials = np.array(self.trials)
        self.trials_rewards = np.array(self.trials_rewards)

        # print examples from each of the above arrays
        if False:
            print ("threshold_state: ", self.threshold_state[:100])
            print ("white_noise_state: ", self.white_noise_state[:100])
            print ("tone_state: ", self.tone_state[:100])
            print ("ensemble_state: ", self.ensemble_state[:100])
            print ("reward_state: ", self.reward_state[:100])
            print ("trials: ", self.trials[:100])
            print ("trials_rewards: ", self.trials_rewards[:100])


    #
    def compute_warping_functions(self):

        #
        fps = 30
    

        #
        if self.burst_to_threshold_only==False:
            snippets = self.snippets.copy()
            snippets_warped = self.snippets_warped.copy()
            ts2 = self.mean
        else:
            snippets = self.snippets_to_thresh.copy()
            snippets_warped = self.snippets_to_thresh_warped.copy()
            ts2 = self.mean_to_thresh

        #
        t = np.arange(0, len(ts2), 1)/fps

        ##############################################
        ############## SETUP PLOTS ###################
        ##############################################
        if self.plotting:
            plt.figure(figsize=(15, 5))
            ax1 = plt.subplot(131)
            ax1.set_xlabel('Time (s)')
            if self.use_stretched:
                ax1.set_title('Stretched trial traces')
            else:
                ax1.set_title('Original trial traces')

            #
            ax2 = plt.subplot(132)
            ax2.set_xlabel('Time (normalized)')
            ax2.set_title('Time warped trial traces')

            #
            ax3 = plt.subplot(133)
            ax3.set_xlabel('Time (normalized)')
            ax3.set_title('Warping functions')

            if self.use_stretched:
                ax1.plot(t, ts2, color='red', linewidth=5,
                        label="average of stretched functions")# marker='o')
                ax1.legend()
            ax2.plot(t, ts2, color='red', linewidth=5,
                    label="average of stretched functions")# marker='o')
            ax2.legend()

            cmap_r = plt.cm.viridis(np.linspace(0, 1, len(self.snippets)))


        ##############################################
        ############## SETUP PLOTS ###################
        ##############################################
        warping_functions = []
        self.warped_snippets = []

        #
        for k in range(snippets.shape[0]):
            
            # option 1 do dynanmic time warping on the streetched
            if self.use_stretched:
                ts1 = snippets_warped[k]
            else:
                ts1 = snippets[k]


            t2 = np.arange(0, len(ts1), 1)/fps

            # Warp ts1 to best match ts2
            warp_function = warp_ts1_to_ts2(ts1, ts2)

            # Store the warping function
            warping_functions.append(warp_function)

            #
            warped_ts1 = interpolate_time_warp(ts1, ts2, warp_function)
            self.warped_snippets.append(warped_ts1)

            # Plot
            t = np.arange(0, len(warped_ts1), 1)/fps

            # normalize the warping function
            t3 = np.arange(0, len(warp_function), 1)/len(warp_function)

            if self.plotting:
                ax1.plot(t2, ts1, c=cmap_r[k])#, label='Warped Time Series 1', color='red', linestyle='dashed', marker='x')
                ax2.plot(t, warped_ts1, c=cmap_r[k])#, label='Warped Time Series 1', color='red', linestyle='dashed', marker='x')
                ax3.plot(t3, warp_function, c=cmap_r[k])#, label='Warping Function', color='red', linestyle='dashed', marker='x')

        self.warp_functions = warping_functions

        #
        if self.plotting:
            plt.show()

        # save warping functions
        warp_functions = np.array(warping_functions, dtype=object)
        np.save(os.path.join(self.root_dir,
                            self.animal_id,
                            self.session_ids[self.session_id],
                            'warp_functions.npy'), warp_functions, allow_pickle=True)
   
    #
    def plot_warped_cell_ca_trial_and_burst(self):

        if self.burst_sign==1:
            self.plot_warped_cell_ca_trial_and_burst_positive()

        else:
            self.plot_warped_cell_ca_trial_and_burst_negative()


    #
    def plot_warped_cell_ca_trial_and_burst_negative(self):

        #
        fps = 30
        
        #
        n_cells = self.sessions[self.session_id].F_upphase_bin.shape[0]
        print ("n_cells: ", n_cells, "# bursts: ", len(self.ids_snippets))

        #
        ids_snippets_local = self.ids_snippets.copy()

        #

        #
        cell_ids = np.arange(self.sessions[self.session_id].F_upphase_bin.shape[0])
        #line_burst = np.zeros(301)
        line_trial = np.zeros(400)
        burst_raw = []
        pre_burst_raw = []
        post_burst_raw = []
        burst_warped = []
        pre_burst_warped = []
        for cell_id in tqdm(cell_ids):
            
            #
            ts2 = self.mean.copy()

            #
            trace = self.sessions[self.session_id].F_upphase_bin[cell_id].squeeze()

            # get the burst segment and warp it
            stack1 = []
            stack2 = []

            #
            for k in range(len(ids_snippets_local)):

                # get start-ends from the saved snippets for the negative bursts
                se = self.starts_ends[ids_snippets_local[k]]

                try:
                    ts1 = trace[se[0]:se[1]]
                except:
                    continue
    
                #temp = line.copy()*0
                #temp[0:ts1.shape[0]] = ts1

                # get the warp function of current snippet to average mean
                warp_function = warp_ts1_to_ts2(ts1, ts2.copy())

                # apply warp function and resample to fixed points
                warped_ca = interpolate_time_warp(ts1, ts2.copy(), warp_function)
                
                if warped_ca.shape[0]!=30:
                    continue

                stack1.append(ts1)
                stack2.append(warped_ca)


            #
            #burst_raw.append(np.vstack(stack1))
            burst_warped.append(np.vstack(stack2))

            # get the pre-burst segment and warp it
            stack1 = []
            stack2 = []
            for k in range(len(ids_snippets_local)):
                idx = ids_snippets_local[k]
                se = self.starts_ends[idx]

                #
                ts1 = trace[se[0]-fps*10:se[0]]

                #if ts1.shape[0]!=fps*10:
                #    print ("skipping pre-burst segment", k)
                #    continue

                temp = line_trial.copy()*0
                temp[0:ts1.shape[0]] = ts1
                ts1=temp

                #
                stack1.append(ts1)
                
                # split the ts1 array every 3rd frame and compute the mean in each chunk
                ts2 = np.mean(np.split(ts1, 100), axis=1)
                #print ("ts2: ", ts2.shape)
                stack2.append(ts2)            
        
            #
            pre_burst_raw.append(np.vstack(stack1))
            pre_burst_warped.append(np.vstack(stack2))

            # do the same for post-burst segment and warp it
            stack1 = []
            for k in range(len(ids_snippets_local)):
                idx = ids_snippets_local[k]
                se = self.starts_ends[idx]

                # grab the following 10 sec of data after the se ends
                ts1 = trace[se[1]:se[1]+300]

                #
                ts1 = scipy.signal.resample(ts1, 100)
                stack1.append(ts1)

            post_burst_raw.append(np.vstack(stack1))            

        ############################################################
        #################### POST PROCESSING #######################
        ############################################################

        # make traces into arrays
        #burst_raw = np.array(burst_raw).squeeze()
        burst_warped = np.array(burst_warped).squeeze()
        print ("burst_warped: ", burst_warped.shape)

        # same for pre-burst
        pre_burst_raw = np.array(pre_burst_raw)
        pre_burst_warped = np.array(pre_burst_warped)
        print ("pre_burst_raw: ", pre_burst_raw.shape, " pre_burst_warped: ", pre_burst_warped.shape)

        # same for post-burst
        post_burst_raw = np.array(post_burst_raw)
        print ("post_burst_raw: ", post_burst_raw.shape)

        # compute mean in the 0th axis
        # burst_raw_mean = np.mean(burst_raw, axis=1)
        # print ("burst_raw_mean: ", burst_raw_mean.shape)
        # burst_warped_mean = np.mean(burst_warped, axis=1)

        # make stack of all traces
        traces_all = [] 
        for k in range(pre_burst_warped.shape[0]):
            if k==0:
                print (pre_burst_warped[k].shape,
                                            burst_warped[k].shape,
                                            post_burst_raw[k].shape)
            traces_all.append(np.hstack((pre_burst_warped[k],
                                         burst_warped[k],
                                         post_burst_raw[k])))
        traces_all = np.array(traces_all)
        print ("traces_all: ", traces_all.shape)
        
        #
        traces_all_mean = np.mean(traces_all, axis=1)

        # find the max of each cell in the means and reorder the traces
        traces_all_mean_max = np.max(traces_all_mean, axis=1)

        # sort the traces
        traces_all_mean_sorted = np.argsort(traces_all_mean_max)[::-1]

        # reorder the traces
        traces_all = traces_all[traces_all_mean_sorted]


        ############################################################
        ####################### PLOTTING ###########################
        ############################################################
        # plot first 100 of each
        plt.figure(figsize=(10,10))
        plt.suptitle("Animal: "+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(len(self.ids_snippets_rewarded))
                    #+ "\n"+
                    #str(self.unique_tones[tone_id])+"hz" + ", #: "+str(len(starts))
                    )
        ctr=0
        start = self.start_idx
        for k in range(0,36,1):
            plt.subplot(6,6,ctr+1)
            plt.title(str(k),pad=0.1, fontsize=7)
            img = traces_all[start+k]
            plt.imshow(img,
                        aspect='auto')
            plt.xticks([])
            plt.yticks([])
            
            if k==0:
                plt.title("Pre-burst (10s) | Burst (1-5s)| Post (10s)",fontsize=7)

            # plot vertical line at 900 and 1100
            plt.plot([100,100],[-0.5,img.shape[0]-0.5],'--',c='red',linewidth=1)
            plt.plot([130,130],[-0.5,img.shape[0]-0.5],'--',c='red',linewidth=1)


            ctr+=1

        # plot first 100 cells average curve
        plt.figure(figsize=(10,10))
        plt.suptitle("Animal: "+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(len(self.ids_snippets_rewarded))
                    #+ "\n"+
                    #str(self.unique_tones[tone_id])+"hz" + ", #: "+str(len(starts))
                    )
        ctr=0
        for k in range(0,36,1):
            plt.subplot(6,6,ctr+1)

            # plot title with low padding
            plt.title(str(k),pad=0.1, fontsize=7)

            #
            mean = np.mean(traces_all[start+k],axis=0)
            t = np.arange(0, mean.shape[0], 1)/30
            plt.plot(t, mean)
            
            # plot vertical line at 900 and 1100
            plt.plot([t[100],t[100]],[0,1],'--',c='grey',linewidth=1, alpha=0.4)
            plt.plot([t[130],t[130]],[0,1],'--',c='grey',linewidth=1, alpha=0.4)

            # plot horizontal line at 0.3
            plt.plot([t[0],t[-1]],[0.3,0.3],'--',c='grey',linewidth=3, alpha=0.4)

            #
            if k==0:
                plt.title("Pre-burst (10s) | Burst (1-5s)| Post (10s)",fontsize=7)

            if k>0:
                plt.yticks([])

            #if k==90:
            #else:
            plt.xticks([])
            plt.ylim(0,1)

            #
            plt.xlim(t[0],t[-1])

            #
            ctr+=1


    #
    def plot_warped_cell_ca_trial_and_burst_positive(self):

        n_cells = self.sessions[self.session_id].F_upphase_bin.shape[0]
        # if self.burst_sign==1:
        #     print ("n_cells: ", n_cells, "# bursts: ", len(self.ids_snippets_rewarded))
        #     ids_snippets_local = self.ids_snippets_rewarded.copy()
        # else:
        #     print ("n_cells: ", n_cells, "# bursts: ", len(self.ids_snippets))
        #     ids_snippets_local = self.ids_snippets.copy()


 #
        if self.burst_sign==1:
            ids_snippets_local = self.ids_snippets_rewarded.copy()
        else:
            ids_snippets_local = self.ids_snippets.copy()
            
        #
        if self.burst_to_threshold_only==False:
            #snippets = self.snippets.copy()
            #snippets_warped = self.snippets_warped.copy()
            ts2 = self.mean
            starts_ends_local = self.starts_ends.copy()[ids_snippets_local]
        else:
            #snippets = self.snippets_to_thresh.copy()
            #snippets_warped = self.snippets_to_thresh_warped.copy()
            ts2 = self.mean_to_thresh
            starts_ends_local = self.starts_ends_to_threshold.copy()


        #
        ts2 = self.mean

        #
        cell_ids = np.arange(self.sessions[self.session_id].F_upphase_bin.shape[0])
        line_burst = np.zeros(400)
        line_trial = np.zeros(901)
        burst_raw = []
        pre_burst_raw = []
        post_burst_raw = []
        burst_warped = []
        pre_burst_warped = []
        for cell_id in tqdm(cell_ids):
            
            if self.ca_type == 'upphase':
                trace = self.sessions[self.session_id].F_upphase_bin[cell_id].squeeze()
            elif self.ca_type == 'DFF':
                trace = self.sessions[self.session_id].F_filtered[cell_id].squeeze()

            # get the burst segment and warp it
            stack1 = []
            stack2 = []
            for k in range(len(starts_ends_local)):
                #idx = ids_snippets_local[k]
                se = starts_ends_local[k]

                if se[1]-se[0]<2:
                    continue

                try:
                    ts1 = trace[se[0]:se[1]]
                except:
                    continue
                # 
                temp = line_burst.copy()*0
                temp[0:ts1.shape[0]] = ts1
                stack1.append(temp)

                # get the warp function of current snippet to average mean
                warp_function = warp_ts1_to_ts2(ts1, ts2)

                # apply warp function and resample to fixed points
                warped_ca = interpolate_time_warp(ts1, ts2, warp_function)
                #print (cell_id, k, "warped_ca: ", warped_ca.shape, "warp_function: ", warp_function.shape)
                stack2.append(warped_ca)

            #
            burst_raw.append(np.vstack(stack1))
            burst_warped.append(np.vstack(stack2))

            # get the pre-burst segment and warp it
            stack1 = []
            stack2 = []
            for k in range(len(starts_ends_local)):
                #idx = ids_snippets_local[k]
                se = starts_ends_local[k]

                if se[1]-se[0]<2:
                    continue

                # find the burst ID
                burst_seg = self.trials[se[0]:se[1]]
                idx = np.where(burst_seg==1000)[0]
                # delete these indexes
                burst_seg = np.delete(burst_seg, idx)

                # check if all values are identical and grab the value
                if np.all(burst_seg==burst_seg[0]):
                    trial_id = burst_seg[0]
                else:
                    print ("Error: burst not matching to a unique trial")
                
                # grab the calcium during the trial
                idx = np.where(self.trials==trial_id)[0]

                # This is wrong!  It also includes the burst time points so we're counting them twice
                # need to delete the time points from idx that are also in se
                se_idx = np.arange(se[0],se[1])
                idx = np.delete(idx, np.where(np.isin(idx, se_idx))[0])
                #print ("idx: ", idx.shape)

                if idx.shape[0]==0:
                    ts1 = np.zeros(901)    
                else:                      
                    ts1 = trace[idx]

                #print ("[ca] snippet: ", ts1.shape)

                temp = line_trial.copy()*0
                temp[-ts1.shape[0]:] = ts1
                stack1.append(temp)

                # use scipy to resample ts1 to be the size of line_trial.shape[0]
                ts1 = scipy.signal.resample(ts1, 200)
                stack2.append(ts1)
            
                #return
        
            #
            pre_burst_raw.append(np.vstack(stack1))
            pre_burst_warped.append(np.vstack(stack2))

            # do the same for post-burst segment and warp it
            stack1 = []
            for k in range(len(ids_snippets_local)):

                se = starts_ends_local[k]

                if se[1]-se[0]<2:
                    continue
            
                idx = ids_snippets_local[k]
                se = self.starts_ends[idx]

                # grab the following 10 sec of data after the se ends
                ts1 = trace[se[1]:se[1]+300]

                #
                ts1 = scipy.signal.resample(ts1, 100)
                stack1.append(ts1)

            post_burst_raw.append(np.vstack(stack1))            

        ############################################################
        #################### POST PROCESSING #######################
        ############################################################

        # make traces into arrays
        burst_raw = np.array(burst_raw).squeeze()
        burst_warped = np.array(burst_warped).squeeze()
        print ("burst_raw: ", burst_raw.shape)

        # same for pre-burst
        pre_burst_raw = np.array(pre_burst_raw)
        pre_burst_warped = np.array(pre_burst_warped)

        # same for post-burst
        post_burst_raw = np.array(post_burst_raw)
        print ("post_burst_raw: ", post_burst_raw.shape)

        # compute mean in the 0th axis
        burst_raw_mean = np.mean(burst_raw, axis=1)
        print ("burst_raw_mean: ", burst_raw_mean.shape)
        burst_warped_mean = np.mean(burst_warped, axis=1)

        # make stack of all traces
        traces_all = [] 
        for k in range(burst_raw.shape[0]):
            if k==0:
                print (pre_burst_warped[k].shape,
                                            burst_warped[k].shape,
                                            post_burst_raw[k].shape)
            traces_all.append(np.hstack((pre_burst_warped[k],
                                         burst_warped[k],
                                         post_burst_raw[k])))
        traces_all = np.array(traces_all)
        print ("traces_all: ", traces_all.shape)
        
        #
        traces_all_mean = np.mean(traces_all, axis=1)

        # find the max of each cell in the means and reorder the traces
        traces_all_mean_max = np.max(traces_all_mean, axis=1)

        # sort the traces
        traces_all_mean_sorted = np.argsort(traces_all_mean_max)[::-1]

        # reorder the traces
        traces_all = traces_all[traces_all_mean_sorted]


        ############################################################
        ####################### PLOTTING ###########################
        ############################################################
        # plot first 100 of each
        plt.figure(figsize=(10,10))
        plt.suptitle("Animal: "+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(len(self.ids_snippets_rewarded))
                    #+ "\n"+
                    #str(self.unique_tones[tone_id])+"hz" + ", #: "+str(len(starts))
                    )
        ctr=0
        start = self.start_idx
        for k in range(0,36,1):
            plt.subplot(6,6,ctr+1)
            plt.title(str(k),pad=0.1, fontsize=7)
            img = traces_all[start+k]
            plt.imshow(img,
                        aspect='auto')
            plt.xticks([])
            plt.yticks([])
            
            if k==0:
                plt.title("Trial (5-30s) | Burst (1-5s)| Post (10s)",fontsize=7)

            # plot vertical line at 900 and 1100
            plt.plot([200,200],[-0.5,img.shape[0]-0.5],'--',c='red',linewidth=1)
            plt.plot([230,230],[-0.5,img.shape[0]-0.5],'--',c='red',linewidth=1)


            ctr+=1

        # plot first 100 cells average curve
        plt.figure(figsize=(10,10))
        plt.suptitle("Animal: "+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(len(self.ids_snippets_rewarded))
                    #+ "\n"+
                    #str(self.unique_tones[tone_id])+"hz" + ", #: "+str(len(starts))
                    )
        ctr=0
        for k in range(0,36,1):
            plt.subplot(6,6,ctr+1)

            # plot title with low padding
            plt.title(str(k),pad=0.1, fontsize=7)

            #
            mean = np.mean(traces_all[start+k],axis=0)
            t = np.arange(0, mean.shape[0], 1)/30
            plt.plot(t, mean)
            
            # plot vertical line at 900 and 1100
            plt.plot([t[200],t[200]],[0,1],'--',c='grey',linewidth=1, alpha=0.4)
            plt.plot([t[230],t[230]],[0,1],'--',c='grey',linewidth=1, alpha=0.4)

            # plot horizontal line at 0.3
            plt.plot([t[0],t[-1]],[0.3,0.3],'--',c='grey',linewidth=3, alpha=0.4)

            #
            if k==0:
                plt.title("Trial (5-30s) | Burst (1-5s)| Post (10s)",fontsize=7)

            if k>0:
                plt.yticks([])

            #if k==90:
            #else:
            plt.xticks([])
            plt.ylim(0,1)

            #
            plt.xlim(t[0],t[-1])

            #
            ctr+=1

    #
    def plot_warped_cell_ca_burst_only(self):

        n_cells = self.sessions[self.session_id].F_upphase_bin.shape[0]
        print ("n_cells: ", n_cells, "# rewarsds: ", len(self.ids_snippets_rewarded))

        #
        if self.burst_sign==1:
            ids_snippets_local = self.ids_snippets_rewarded.copy()
        else:
            ids_snippets_local = self.ids_snippets.copy()
            
        #
        if self.burst_to_threshold_only==False:
            #snippets = self.snippets.copy()
            #snippets_warped = self.snippets_warped.copy()
            ts2 = self.mean
            starts_ends_local = self.starts_ends.copy()[ids_snippets_local]
        else:
            #snippets = self.snippets_to_thresh.copy()
            #snippets_warped = self.snippets_to_thresh_warped.copy()
            ts2 = self.mean_to_thresh
            starts_ends_local = self.starts_ends_to_threshold.copy()
        #
        #print ("ts2: ", ts2.shape, " starts_ends_local: ", starts_ends_local)

        #
        cell_ids = np.arange(self.sessions[self.session_id].F_upphase_bin.shape[0])
        line = np.zeros(400)
        traces_raw = []
        traces_warped = []
        for ctr, cell_id in enumerate(cell_ids):
            
            trace = self.sessions[self.session_id].F_upphase_bin[cell_id].squeeze()

            #
            stack1 = []
            stack2 = []
            #for k in range(len(self.ids_snippets_rewarded)):
            for k in range(len(starts_ends_local)):

                # get start-ends from the saved snippets for the negative bursts
                #print ("k: ", k, " starts_ends_local[k]: ", starts_ends_local[k])
                se = starts_ends_local[k]

                try:
                    ts1 = trace[se[0]:se[1]]
                except:
                    continue
    
                temp = line.copy()*0
                temp[0:ts1.shape[0]] = ts1
                stack1.append(temp)

                # get the warp function of current snippet to average mean
                warp_function = warp_ts1_to_ts2(ts1, ts2)

                # apply warp function and resample to fixed points
                warped_ca = interpolate_time_warp(ts1, ts2, warp_function)
                stack2.append(warped_ca)

            #
            traces_raw.append(np.vstack(stack1))
            traces_warped.append(np.vstack(stack2))

        # make traces into arrays
        traces_raw = np.array(traces_raw)
        print ("traces-raw: ", traces_raw.shape)
        traces_warped = np.array(traces_warped)

        # compute mean in the 0th axis
        traces_raw_mean = np.mean(traces_raw, axis=1)
        print ("traces mean: ", traces_raw_mean.shape)
        traces_warped_mean = np.mean(traces_warped, axis=1)

        # find the max of each cell in the means and reorder the traces
        traces_raw_mean_max = np.max(traces_raw_mean, axis=1)
        traces_warped_mean_max = np.max(traces_warped_mean, axis=1)

        # sort the traces
        traces_raw_mean_max_sorted = np.argsort(traces_raw_mean_max)[::-1]
        traces_warped_mean_max_sorted = np.argsort(traces_warped_mean_max)[::-1]

        # reorder the traces
        traces_raw = traces_raw[traces_raw_mean_max_sorted]
        traces_warped = traces_warped[traces_warped_mean_max_sorted]

        # plot first 100 of each
        plt.figure(figsize=(10,10))
        ctr=0
        start = self.start_idx
        end = start+100
        for k in range(0,100,1):
            plt.subplot(10,10,ctr+1)
            plt.title(str(k),pad=0.1, fontsize=7)
            plt.imshow(traces_warped[start+k],
                    aspect='auto')
            plt.xticks([])
            plt.yticks([])
            ctr+=1

        # plot first 100 cells average curve
        plt.figure(figsize=(10,10))
        ctr=0
        for k in range(0,100,1):
            plt.subplot(10,10,ctr+1)

            # plot title with low padding
            plt.title(str(k),pad=0.1, fontsize=7)

            #
            temp = traces_warped[start+k]
            #print (temp.shape)
            temp = np.mean(temp, axis=0)
            plt.plot(temp)

            # plot horizontal line at 0.3
            plt.plot([0,temp.shape[0]],[0.3,0.3],'--',c='grey',linewidth=3, alpha=0.4)

            #
            plt.xticks([])
            if k>0:
                plt.yticks([])
            plt.ylim(0,1)
            ctr+=1


    def get_rewarded_trials_align_trial_aligned(self):

        #
        fps=30

        #
        buffer = self.buffer #10*30
        window = self.window + buffer

        #
        img_rewarded = np.zeros((len(self.trial_ids_rewarded),window),dtype=np.float32)
        img_non_rewarded = np.zeros((len(self.trial_ids_not_rewarded),window),dtype=np.float32)

        line = np.zeros(window)+np.nan

        plt.figure(figsize=(12,6))
        plt.suptitle("Ensemble state"+
                     "\n"+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(len(self.trial_ids_rewarded))
                    )
        ax1 = plt.subplot(121)
        ax1.set_title("Rewarded trials")
        ax1.set_xlabel("Time (s)")
        ax1.set_ylabel("Trial #")
        ax2 = plt.subplot(122)
        ax2.set_title("Non rewarded trials")
        ax2.set_xlabel("Time (s)")
        ax2.set_ylabel("Trial #")

        # add the c.ensemble_state triggered on the beggining and ends of each rewarded trial
        for k in range(len(self.trial_ids_rewarded)):
            # get trial id
            i = self.trial_ids_rewarded[k]

            # get times of trial
            idx = np.where(self.trials==i)[0]
            temp = self.ensemble_state[idx[0]:idx[-1]+buffer-1]
            line*=0 + np.nan
            line[:temp.shape[0]] =  temp
            img_rewarded[k,:] = line
 
            # plot reward end for each trial as a red dot at 
            if k ==0:
                ax1.plot([idx.shape[0]/fps, idx.shape[0]/fps],
                     [k-0.5,k+0.5],c='r', linewidth=3, label='Reward time')
            else:
                ax1.plot([idx.shape[0]/fps, idx.shape[0]/fps],
                     [k-0.5,k+0.5],c='r', linewidth=3)

            


        # same for non rewarded trials
        for k in range(len(self.trial_ids_not_rewarded)):
            # get trial id
            i = self.trial_ids_not_rewarded[k]

            # get times of trial
            idx = np.where(self.trials==i)[0]
            temp = self.ensemble_state[idx[0]:idx[-1]+buffer-1]
            line*=0 + np.nan
            line[:temp.shape[0]] =  temp
            img_non_rewarded[k,:] = line

            # 
            # draw a verttical red line of height 1 at idx.shape[0]/fps
            ax2.plot([idx.shape[0]/fps, idx.shape[0]/fps],
                     [k-0.5,k+0.5],c='r', linewidth=3)     
            
            
        # invert the images upside down
        img_rewarded = np.flipud(img_rewarded)
        img_non_rewarded = np.flipud(img_non_rewarded)
        
       
        vmax = np.median(self.threshold_state)
        ax1.imshow(img_rewarded,aspect='auto',
                extent=[0,window/fps,-0.5,len(self.trial_ids_rewarded)-0.5],
                    interpolation='none',
                    cmap = self.cmap,
                    vmin = -vmax,
                    vmax = vmax  
                    )
        ax1.legend()


        clb = ax2.imshow(img_non_rewarded,aspect='auto',
                extent=[0,window/fps,-0.5,len(self.trial_ids_not_rewarded)-0.5],
                    interpolation='none',
                    cmap = self.cmap,
                    vmin = -vmax,
                    vmax = vmax  
                    )

        # show colorbar for ax2 with title "DFF"
        clb = plt.colorbar(clb,ax=ax2,shrink=1).set_label("Ensemble state (DFF units)")
        
        
    def get_rewarded_trials_align_reward_aligned_warped(self):

        fps=30

        #
        buffer = self.buffer #10*30
        window = self.window + buffer

        #
        img_rewarded = np.zeros((len(self.trial_ids_rewarded),window),dtype=np.float32)
        img_non_rewarded = np.zeros((len(self.trial_ids_not_rewarded),window),dtype=np.float32)

        line = np.zeros(window)+np.nan

        plt.figure(figsize=(6,6))
        plt.suptitle("Ensemble state"+
                     "\n"+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(len(self.trial_ids_rewarded))
                    )
        ax1 = plt.subplot(111)
        ax1.set_title("Rewarded trials")
        ax1.set_xlabel("Time (s)")
        ax1.set_ylabel("Trial #")
        # ax2 = plt.subplot(122)
        # ax2.set_title("Non rewarded trials")
        # ax2.set_xlabel("Time (s)")
        # ax2.set_ylabel("Trial #")

        # add the c.ensemble_state triggered on the beggining and ends of each rewarded trial
        for k in range(len(self.trial_ids_rewarded)):
            # get trial id
            i = self.trial_ids_rewarded[k]

            # get times of trial
            idx = np.where(self.trials==i)[0]
            temp_pre = self.ensemble_state[idx[0]:idx[-1]]
            temp_post = self.ensemble_state[idx[-1]:idx[-1]+buffer]
            line*=0 + np.nan

            # resample temp_pre to have the same length as window
            temp_pre = scipy.signal.resample(temp_pre, self.window)
           # print ("temp_pre: ", temp_pre.shape, "line: ", line.shape)
            
            line[:self.window] = temp_pre
            line[self.window:self.window+temp_post.shape[0]] =  temp_post
            
            img_rewarded[k,:] = line
 
            # plot reward end for each trial as a red dot at 
            if k ==0:
                ax1.plot([self.window/fps, self.window/fps],
                     [k-0.5,k+0.5],c='r', linewidth=3, label='reward time')
            else:
                ax1.plot([self.window/fps, self.window/fps],
                     [k-0.5,k+0.5],c='r', linewidth=3)

            
           
        # invert the images upside down
        img_rewarded = np.flipud(img_rewarded)
        # img_non_rewarded = np.flipud(img_non_rewarded)
        
       
        vmax = np.median(self.threshold_state)
        clb = ax1.imshow(img_rewarded,aspect='auto',
                extent=[0,window/fps,-0.5,len(self.trial_ids_rewarded)-0.5],
                    interpolation='none',
                    cmap = self.cmap,
                    vmin = -vmax,
                    vmax = vmax  
                    )
        ax1.legend()


        # clb = ax2.imshow(img_non_rewarded,aspect='auto',
        #         extent=[0,window/fps,-0.5,len(self.trial_ids_not_rewarded)-0.5],
        #             interpolation='none',
        #             cmap = self.cmap,
        #             vmin = -vmax,
        #             vmax = vmax  
        #             )

        # show colorbar for ax2 with title "DFF"
        clb = plt.colorbar(clb,ax=ax1,shrink=1).set_label("Ensemble state (DFF units)")
        
        

    def get_rewarded_trials_align_reward_aligned(self):

        #
        fps=30

        #
        buffer = self.buffer #10*30
        window = self.window + buffer

        #
        img_rewarded = np.zeros((len(self.trial_ids_rewarded),window),dtype=np.float32)
        img_non_rewarded = np.zeros((len(self.trial_ids_not_rewarded),window),dtype=np.float32)

        line = np.zeros(window)+np.nan

        plt.figure(figsize=(6,6))
        plt.suptitle("Ensemble state"+
                     "\n"+str(self.animal_id)+
                    ", Session: "+str(self.session_id)+
                    ", # of rewards: "+str(len(self.trial_ids_rewarded))
                    )
        ax1 = plt.subplot(111)
        ax1.set_title("Rewarded trials")
        ax1.set_xlabel("Time (s)")
        ax1.set_ylabel("Trial #")
        # ax2 = plt.subplot(122)
        # ax2.set_title("Non rewarded trials")
        # ax2.set_xlabel("Time (s)")
        # ax2.set_ylabel("Trial #")

        # add the c.ensemble_state triggered on the beggining and ends of each rewarded trial
        for k in range(len(self.trial_ids_rewarded)):
            # get trial id
            i = self.trial_ids_rewarded[k]

            # get times of trial
            idx = np.where(self.trials==i)[0]
            temp = self.ensemble_state[idx[0]:idx[-1]+buffer-1]
            line*=0 + np.nan

            #
            #print (self.window, buffer, line.shape, idx.shape, temp.shape )
            # 900 90 (990,) (60,) (148,)

            line[self.window-idx.shape[0]:self.window] =  temp[:idx.shape[0]]
            line[self.window:self.window+temp[idx.shape[0]:].shape[0]] =  temp[idx.shape[0]:]
            
            img_rewarded[k,:] = line
 
            # plot reward end for each trial as a red dot at 
            if k ==0:
                ax1.plot([self.window/fps, self.window/fps],
                     [k-0.5,k+0.5],c='r', linewidth=3, label='reward time')
            else:
                ax1.plot([self.window/fps, self.window/fps],
                     [k-0.5,k+0.5],c='r', linewidth=3)

            


        # # same for non rewarded trials
        # for k in range(len(self.trial_ids_not_rewarded)):
        #     # get trial id
        #     i = self.trial_ids_not_rewarded[k]

        #     # get times of trial
        #     idx = np.where(self.trials==i)[0]
        #     temp = self.ensemble_state[idx[0]:idx[-1]+buffer-1]
        #     line*=0 + np.nan
        #     line[:temp.shape[0]] =  temp
        #     img_non_rewarded[k,:] = line

        #     # 
        #     # draw a verttical red line of height 1 at idx.shape[0]/fps
        #     ax2.plot([idx.shape[0]/fps, idx.shape[0]/fps],
        #              [k-0.5,k+0.5],c='r', linewidth=3)     
            
            
        # invert the images upside down
        img_rewarded = np.flipud(img_rewarded)
        # img_non_rewarded = np.flipud(img_non_rewarded)
        
       
        vmax = np.median(self.threshold_state)
        clb = ax1.imshow(img_rewarded,aspect='auto',
                extent=[0,window/fps,-0.5,len(self.trial_ids_rewarded)-0.5],
                    interpolation='none',
                    cmap = self.cmap,
                    vmin = -vmax,
                    vmax = vmax  
                    )
        ax1.legend()


        # clb = ax2.imshow(img_non_rewarded,aspect='auto',
        #         extent=[0,window/fps,-0.5,len(self.trial_ids_not_rewarded)-0.5],
        #             interpolation='none',
        #             cmap = self.cmap,
        #             vmin = -vmax,
        #             vmax = vmax  
        #             )

        # show colorbar for ax2 with title "DFF"
        clb = plt.colorbar(clb,ax=ax1,shrink=1).set_label("Ensemble state (DFF units)")
        
        
        

    def get_rewarded_trials(self):

        trials = self.trials.copy()

        # find trial ids
        trial_ids = np.unique(trials)
        trial_ids = trial_ids[trial_ids < 1000]

        # find which trials are rewarded
        trial_ids_rewarded = []
        trial_ids_not_rewarded = []
        for trial_id in trial_ids:
            idx = np.where(trials == trial_id)[0]
            
            trial_reward = self.trials_rewards[idx].mean()
            if trial_reward == 1:
                trial_ids_rewarded.append(trial_id)
            elif trial_reward == 0:
                trial_ids_not_rewarded.append(trial_id)
            else:
                print ("trial id: ", trial_id, " has issue")

        print ("rewarded trials: ", trial_ids_rewarded)
        print ("not rewarded trials: ", trial_ids_not_rewarded)

        # 
        self.trial_ids_rewarded = trial_ids_rewarded
        self.trial_ids_not_rewarded = trial_ids_not_rewarded
    
    #
    def fix_spreadsheet(self, plotting=False):

        #
        for session_name in self.session_ids[1:]:

            # make an fname_out directory if not already present
            self.fname_out = os.path.join(self.root_dir,
                                            self.animal_id,
                                            session_name,
                                            'results_fixed.xlsx')

            # check if fname_out exists
            if os.path.exists(self.fname_out)==False:
                #session_id = self.session_id

                #load the n_rewards
                fname = os.path.join(self.root_dir,
                                        self.animal_id,
                                        session_name,
                                        'results.npz')
                
                #
                results = np.load(fname, allow_pickle=True)
                rewards = results['reward_times'].T
                idx = np.where(rewards[:,1]>0)[0]

                #
                self.n_rewards = idx.shape[0]
                reward_times = rewards[idx,1]
                #print ("# of rewards: ", self.n_rewards, " rewar[:100])

                # Load the workbook
                wb = load_workbook(os.path.join(self.root_dir,
                                                self.animal_id,
                                                session_name,
                                                'results.xlsx'), read_only=False)

                # Access the worksheet
                ws = wb[wb.sheetnames[0]]  # assuming you want the first sheet, change as needed

                # get white noise state  
                rows = [4,7,8,9]          # white_noise, tone_state, ensemble_state, reward_state
                self.white_noise_state = []
                self.tone_state = []
                self.ensemble_state = []
                self.reward_state = []
                for row in ws.iter_rows(min_row=2,values_only=False):
                    self.white_noise_state.append(row[rows[0]-1].value)
                    self.tone_state.append(row[rows[1]-1].value)
                    self.ensemble_state.append(row[rows[2]-1].value)
                    self.reward_state.append(row[rows[3]-1].value)

                # make all lists into arrays
                self.white_noise_state = np.array(self.white_noise_state)
                self.tone_state = np.array(self.tone_state)
                self.ensemble_state = np.array(self.ensemble_state)
                self.reward_state = np.array(self.reward_state)

                ############################################################
                ####################### COMPUTE TRIALS #####################
                ############################################################
                # compute the trial starts and ends using the tone and reward state information
                trials = []
                start = 0
                trial_ended = False
                trial_started = True

                # loop over all entries in the spreadsheet
                for k in range(self.reward_state.shape[0]):
                    
                    # find next reward and save it
                    if self.reward_state[k]==1 and trial_ended==False and trial_started:
                        trial_ended = True        # this indicates that a reward was previusly given
                        trial_started = False
                        end = k
                        trials.append([start,end,1])  # triple indicating start,end and whether reward was given

                    # same but checking if white_noise state was reached
                    if self.white_noise_state[k]==1 and trial_ended==False and trial_started:
                        trial_ended = True
                        trial_started = False
                        end = k
                        trials.append([start,end,0])
                    
                    # SEARCH FOR THE NEXT TRIAL START
                    # find the next trial start
                    if trial_ended==True:         # check if a reward for the previous trial was already given
                        
                        # check if we're out of the reward state and also out of the white noise state
                        if self.reward_state[k]==0 and self.white_noise_state[k]==0:

                            # check to see if tone is back online and is playing an actualy tone after dynamic lockouts
                            #if self.tone_state[k]>100:
                            if np.all(self.tone_state[k:k+10]>100):   # this check that we've escaped the dynamic lockout
                                                                      # at least for several frames as the tone update can sometimes lag behind the penalty/lockouts for a few frames ( due to computer? not sure)
                                trial_ended = False       # reset the trial start
                                trial_started = True
                                start = k
                
                #
                trials = np.array(trials)
                print ("# of trials: ", trials.shape[0],
                       "# of rewarded trials: ", np.where(trials[:,2]==1)[0].shape[0])
                idx = np.where(trials[:,2]==1)[0]

                # check if 2 arrays are identical: reward_times and trials[idx][:,1]
                if np.array_equal(reward_times, trials[idx][:,1])==False:
                    print ("reward_times and trials[idx][:,1] are not identical")
                    print ("reward_times: ", reward_times)
                    print ("trials[idx][:,1]: ", trials[idx][:,1])
                    print ("Exiting...")
                else:
                    print ("reward_times detected correctly")
                    #break

                # convert trial structure to an array
                trial = np.zeros((self.reward_state.shape[0]), dtype=np.int32)+ 1000
                reward_flag = np.zeros((self.reward_state.shape[0]), dtype=np.int32) + 1000
                ctr=0
                for k in range(len(trials)):
                    trial[trials[k][0]:trials[k][1]]=ctr
                    reward_flag[trials[k][0]:trials[k][1]]=trials[k][2]   # here keep track of each trial whether it was rewarded
                    ctr+=1

                #print ("trial: ", trial[:300])
                trials = trial.copy()
            
                ############################################################
                ################## SAVE UPDATED SPREADSHEET ################
                ############################################################
                # Access the worksheet
                sheet = wb.active

                # Insert the new column. Here, we insert it at column A (index 1). Adjust as necessary.
                n = 10
                sheet.insert_cols(idx=n)
                sheet.cell(row=1, column=n).value = 'trials'
                for k in range(trials.shape[0]):
                    sheet.cell(row=k+2, column=n).value=trials[k]

                # Insert the new column. Here, we insert it at column A (index 1). Adjust as necessary.
                n = 11
                sheet.insert_cols(idx=n)
                sheet.cell(row=1, column=n).value = 'trial_rewarded'
                for k in range(trials.shape[0]):
                    sheet.cell(row=k+2, column=n).value=reward_flag[k]

                # Save the updated workbook
                wb.save(self.fname_out)

                #
                print ('')

            # exit after processing the 1st session for now
        print ("...done...")
            #break



    def binarize_ensemble_state(self, plotting):

        #
        self.sessions[self.session_id].percentile_threshold = 0.98

        #
        #scale_threshold = 0.2
        self.sessions[self.session_id].binarize_traces_single_trace(self.e_state,
                                                                    self.scale_threshold)

        # get locations of positive and negative values of e_state
        pos = np.where(self.e_state >= 0)[0]
        neg = np.where(self.e_state < 0)[0]

        # make an array of -1 and +1 to represent the sign of the e_state
        e_state_sign = np.zeros(self.e_state.shape)
        e_state_sign[pos] = 1
        e_state_sign[neg] = -1

        #
        self.e_state_bin = self.sessions[self.session_id].trace_upphase_bin

        #
        self.e_state_sign = e_state_sign.copy()

        #
        if plotting:
            plot_e_state_binarization(self.e_state,
                                        self.e_state_bin,
                                        e_state_sign,
                                        self.thresh,
                                        )
    #        
    def process_snippets(self):

        # This function detects the upphase bursts and finds the snippet locations

        # NOTE Starts and ends is for the ensemble state - not for single cells...
        # show extracted snippets
        starts_ends = np.array(self.sessions[self.session_id].starts_ends).squeeze()
        self.starts_ends = starts_ends

        # get the sign of each extracted snippet
        signs = []
        for k in range(starts_ends.shape[0]):
            signs.append(np.median(self.e_state_sign[starts_ends[k,0]:starts_ends[k,1]]))

        signs = np.int32(signs)
        
        # list total number of snippets and how many are positive
        
        #
        self.bursts = np.zeros(self.sessions[self.session_id].trace_upphase_bin.shape[0], dtype=np.int32)+1000
        self.bursts_to_threshold = np.zeros(self.sessions[self.session_id].trace_upphase_bin.shape[0], dtype=np.int32)+1000
        self.bursts_rewarded = np.zeros(self.sessions[self.session_id].trace_upphase_bin.shape[0], dtype=np.int32)+1000
        
        #
        print ("# of snippets: ", signs.shape[0], " # of detected snippets of sign: ", self.burst_sign, " ", np.where(signs==self.burst_sign)[0].shape[0])

        ctr_total = 0
        ctr_rewarded = 0
        reward_buffer = 15 # no. of frames to search forward for a reward
        fps = 30
        ids_snippets_rewarded = []
        ids_snippets = []
        self.starts_ends_to_threshold = []
        for k in range(starts_ends.shape[0]):

            if signs[k] == self.burst_sign:
                
                #
                ids_snippets.append(k)

                # extract indexes
                temp_starts_ends = starts_ends[k]
                
                # save all burst times
                self.bursts[temp_starts_ends[0]: temp_starts_ends[1]] = ctr_total

                # check if water reard was given during this period
                water_state = self.water_reward[temp_starts_ends[0]:temp_starts_ends[1]+reward_buffer]
                diff = water_state[1:]-water_state[:-1]
                idx = np.where(diff==1)[0]

                # check if the water state changed from a 0 to a 1 during this period
                if idx.shape[0]>0:
                    
                    # add a reward
                    self.bursts_rewarded[temp_starts_ends[0]:temp_starts_ends[1]]= 1

                    # also save the burst to threshold
                    self.bursts_to_threshold[temp_starts_ends[0]:temp_starts_ends[0]+idx[0]]= 1

                    #
                    self.starts_ends_to_threshold.append([temp_starts_ends[0],temp_starts_ends[0]+idx[0]])

                    # save rewarded burst times
                    ctr_rewarded+=1

                    #
                    ids_snippets_rewarded.append(k)
                else:
                    self.bursts_rewarded[temp_starts_ends[0]: temp_starts_ends[1]]= 0

                #
                ctr_total+=1
        
        #
        self.signs = signs
        self.ids_snippets_rewarded = ids_snippets_rewarded
        self.ids_snippets = ids_snippets

        # print how many of each burst types there are
        print ("# of bursts: ", ctr_total, " # of rewarded bursts: ", ctr_rewarded)
        print ("self.update_spreadsheet: ", self.update_spreadsheet)

        if self.update_spreadsheet==True:
            # read the results_fixed.xlsx file 
            fname = os.path.join(self.root_dir,
                                    self.animal_id,
                                    self.session_ids[self.session_id],
                                    'results_fixed.xlsx')
            #print ("updatinging spreadsheet: ", fname)
            df = pd.read_excel(fname)

            # check to see if any columns named "burst" exists and overwrite it
            if 'burst' in df.columns:
                if self.overwrite_spreadsheet:
                    df['burst'] = self.bursts
            else:
                df['burst'] = self.bursts

            # same for bursts_rewarded
            if 'burst_rewarded' in df.columns:
                if self.overwrite_spreadsheet:
                    df['burst_rewarded'] = self.bursts_rewarded
            else:
                df['burst_rewarded'] = self.bursts_rewarded

            # same for bursts_to_threshold
            if 'burst_to_threshold' in df.columns:
                if self.overwrite_spreadsheet:
                    df['burst_to_threshold'] = self.bursts_to_threshold
            else:
                df['burst_to_threshold'] = self.bursts_to_threshold
                
            # save the updated spreadsheet
            df.to_excel(fname, index=False)

    #
    def fix_missing_vals(self, array_in):

        # we repeat the last value for each of the arrays
        temp = np.zeros(90000)
        temp[:array_in.shape[0]] = array_in
        temp[array_in.shape[0]:] = array_in[-1]
        
        return temp

    #   
    def load_spreadsheet(self):

        # load data from spreassheetds
        df = pd.read_excel(os.path.join(
                            self.root_dir,
                            self.animal_id,
                            self.session_ids[self.session_id],
                            'results_fixed.xlsx'))

        # find a column in the dataframe with the name "current_high_threshold"
        # and get its values as a numpy array
        self.thresh = df.loc[:, 'current_high_threshold'].values

        if self.fix_spreadsheet_missing_vals:
            if self.thresh.shape[0] < 90000:
                
                # 
                print ("Found short spreadsheet: ", self.thresh.shape[0], " entries... fixing it...")

                #
                df = df.reindex(range(90000))

                # 1. Identify the last non-NaN entry in each column
                last_entries = df.apply(lambda col: col[col.last_valid_index()])
                print ("... repeating last_entries: ", last_entries)
                # 2. Fill the NaN values in the extended rows with the last entry
                for column in df.columns:
                    df[column].fillna(last_entries[column], inplace=True)

                # get the first column by number (as it doesn't have a name) and set it to go from 1 to 90000
                df.iloc[:,0] = np.arange(0,90000)

                # fix the "n_ttl" colun to go from 1 to 90000
                df['n_ttl'] = np.arange(1,90001)

                print ("DF: ", df.shape)
                df.to_excel(os.path.join(
                            self.root_dir,
                            self.animal_id,
                            self.session_ids[self.session_id],
                            'results_fixed.xlsx'), 
                        index=False)

                # reload the spreadsheet
                df = pd.read_excel(os.path.join(
                            self.root_dir,
                            self.animal_id,
                            self.session_ids[self.session_id],
                            'results_fixed.xlsx'))

                # and grab the first value again                
                self.thresh = df.loc[:, 'current_high_threshold'].values

        # same for white_noise_state
        self.white_noise_state = df.loc[:, 'white_noise_state'].values

        # same for tone_state
        self.tone_state = df.loc[:, 'tone_state'].values

        # same for trials
        self.trials = df.loc[:, 'trials'].values

        #
        self.water_reward = df.loc[:, 'water_reward'].values

        # same for trial_rewarded
        self.trials_rewards = df.loc[:, 'trial_rewarded'].values

        # same for a columan named "ensemble_state"
        e_state = df.loc[:, 'ensemble_state'].values

        # delete first few frames
        e_state[:100] = 0 

        # filter using savgol_filter
        self.e_state = savgol_filter(e_state, 31, 3)

        # check to see there are <90,000 entries
        if self.fix_spreadsheet_missing_vals:
            if self.thresh.shape[0] < 90000:

                self.thresh = self.fix_missing_vals(self.thresh)
                self.white_noise_state = self.fix_missing_vals(self.white_noise_state)
                self.tone_state = self.fix_missing_vals(self.tone_state)
                self.trials = self.fix_missing_vals(self.trials)
                self.water_reward = self.fix_missing_vals(self.water_reward)
                self.trials_rewards = self.fix_missing_vals(self.trials_rewards)
                self.e_state = self.fix_missing_vals(self.e_state)





    #    
    def get_tones(self, plotting=False):


        # make an fname_out directory if not already present
        self.fname_out = os.path.join(self.root_dir,
                                        self.animal_id,
                                        self.session_ids[self.session_id],
                                        'tone_data.npz')

        # check if fname_out exists
        if os.path.exists(self.fname_out)==False or self.recompute_tones:
            session_id = self.session_id

            # load the n_rewards
            fname = os.path.join(self.root_dir,
                                    self.animal_id,
                                    self.session_ids[session_id],
                                    'results.npz')
            results = np.load(fname, allow_pickle=True)
            rewards = results['reward_times'].T
            idx = np.where(rewards[:,1]>0)[0]
            self.n_rewards = idx.shape[0]
            self.reward_times = rewards[idx,1]
            #print ("n_rewards: ", self.n_rewards)

            # Load the workbook
            wb = load_workbook(os.path.join(self.root_dir,
                                            self.animal_id,
                                            self.session_ids[session_id],
                                        # 'data',
                                            'results.xlsx'), read_only=True)


            # Access the worksheet
            ws = wb[wb.sheetnames[0]]  # assuming you want the first sheet, change as needed

            # Get the n-th column
            n = 7  # change this to your desired column (Python uses 0-based indexing)

            # Create a numpy array from the n-th column, skipping the header row
            self.tones = np.array([row[n-1].value for row in ws.iter_rows(min_row=2, values_only=False)])

            #
            self.tones = np.int32(self.tones)

            #
            n = 8
            self.ensemble_state = np.array([row[n-1].value for row in ws.iter_rows(min_row=2, values_only=False)])

            # save fname_out
            np.savez(self.fname_out, 
                     tones = self.tones,
                     reward_times = self.reward_times,
                     n_rewards = self.n_rewards,
                     ensemble_state = self.ensemble_state)


            # plot the mode as a reference
            if plotting:
                plt.figure()
                # find the most frequent value in tone
                from scipy.stats import mode
                mode_tone = mode(self.tones)

                plt.plot(self.tones)
                plt.plot(np.ones(self.tones.shape)*mode_tone[0],'r--')

                #
                plt.xlabel("Frames")
                plt.ylabel("Tone")

                #
                plt.suptitle(c.animal_id+ " - " + c.session_ids[1])

                #
                plt.show()
        else:
            #print ("loading tones from: ", self.fname_out)
            data = np.load(self.fname_out, allow_pickle=True)
            self.tones = data['tones']
            self.n_rewards = data['n_rewards']
            self.reward_times = data['reward_times']

    # load single data [ca] data
    def load_single_data(self):

        #          
        data_dir = os.path.join(
                        self.root_dir,
                        self.animal_id,
                        self.session_ids[self.session_id],
                        'plane0'
                        )

        C = Calcium()       
        C.data_dir = data_dir

        #
        C.load_suite2p()

        #
        C.load_footprints()

        # load binarization
        C.detrend_model_type = 'polynomial'
        C.detrend_model_order = 2
        C.percentile_threshold = 0.999999
        
        # 
        print ("loading binarization")
        C.load_binarization()
        
        return C

    #
    def load_data(self):

        #for animal_id in self.animal_ids:
        self.sessions = []
        for session_ in tqdm(self.session_ids):


            # make a directory called 'cells' if not already present
            self.cells_dir = os.path.join(self.root_dir,
                                            self.animal_id,
                                            'cells')
            if not os.path.exists(self.cells_dir):
                os.makedirs(self.cells_dir)

                
            data_dir = os.path.join(
                            self.root_dir,
                            self.animal_id,
                            session_,
                            'plane0'
                            )

            #
            try:
                C = Calcium()       
                C.data_dir = data_dir

                #
                C.load_suite2p()

                #
                C.load_footprints()

                # load binarization
                C.detrend_model_type = 'polynomial'
                C.detrend_model_order = 2
                C.percentile_threshold = 0.999999
                
                # 
                print ("loading binarization")
                C.load_binarization()
            
            except:
                print ("could not load session: ", session_)
                C = None

            #
            self.sessions.append(C)

    #
    def plot_session_contours(self, 
                              clr,
                              x_shift=0,
                              y_shift=0,
                              theta =0,
                              theta_x = 256,
                              theta_y = 256,
                              scale_factor_x = 1,
                              scale_factor_y = 1,
                              ):

        # replotting 
        session_contours = self.sessions[self.session_id].contours

        if clr=='red':
            idxs = self.day_cell_idx
        else:
            idxs = self.cell_idxs

        for k in idxs:
            temp2 = session_contours[k].copy()
            temp= temp2.copy()

            # rescale all contours by this factor but from self.scale_x and self.scale_y as centre
            # only the centres of the cells shift not the shape of the cells...
            # get centres of cells
            centre = np.mean(temp,0)

            # scale centres by subtracing from the centre theta_x and theta_y and then scaling by scale_factor
            xx = (centre[0] - theta_x)*scale_factor_x + theta_x
            yy = (centre[1] - theta_y)*scale_factor_y + theta_y

            # move the contour from the centre to the new centre
            temp = temp - centre + np.array([xx,yy])

            #
            temp = rotate_points(temp, theta_x, theta_y, theta)

            # shift the contour by x_shift and y_shift
            temp[:,0] = temp[:,0] + x_shift
            temp[:,1] = temp[:,1] + y_shift

            #
            self.ax.plot(temp[:,0],
                        temp[:,1],
                        c=clr,
                        alpha=.5,
                        linewidth=2,
                        label="session: "+str(self.session_id) if k==0 else ""
                        )
            


    #
    def plot_quadrants(self):

        self.ax.plot([0,512],
                     [self.theta_x,self.theta_x],
                     c='black',
                     linestyle='--')
        
        # same for y vertical line
        self.ax.plot([self.theta_y,self.theta_y],
                        [0,512],
                        c='black',
                        linestyle='--')

    
    #
    def plot_ROIs_contours(self):
        clrs = ['blue','lightblue','red','pink']
        centres = []
        for k in range(len(self.contours_ROIs)):
            temp2 = self.contours_ROIs[k][0]
            temp= temp2.copy()
            temp[:,0] = temp2[:,1]
            temp[:,1] = temp2[:,0]
            #print (temp.shape)
            plt.plot(temp[:,0],
                    temp[:,1],
                    c=clrs[k],
                    alpha=.5,
                    linewidth=6,
                    label="ROI: "+str(k))
            centre = np.mean(temp,0)
            centres.append(centre)

        # make vstack array
        centres = np.vstack(centres)

        return centres

    def shift_contours(self, event):

        print ("shifting contours")
        self.x_shift-=1

        clr = 'blue'
        self.plot_session_contours(clr, 
                                   self.x_shift,
                                   self.y_shift)

        #
        clr = 'red'
        self.session_id=0
        self.plot_session_contours(clr)

    #
    def plot_day0_session_n_contours(self):
        import matplotlib.pyplot as plt
        from matplotlib.widgets import Button
    
        #
        self.x_shift = 0 
        self.y_shift = 0

        #
        plt.figure(figsize=(10,10))
        plt.subplot(1,1,1)

        #
        clr = 'blue'
        self.plot_session_contours(clr)

        #
        clr = 'red'
        self.session_id=0
        self.plot_session_contours(clr)

        #
        plt.xlim(0,512)
        plt.ylim(0,512)

        # Define a function to handle the button click event
        def on_button_click(event):
            print("Click")

        # Create a button widget
        button_ax = plt.axes([0.7, 0.05, 0.2, 0.075])  # [x, y, width, height]
        button = Button(button_ax, 'Click me!')

        # Connect the button click event to the function
        button.on_clicked(on_button_click)

        # add a clickable button to shift the contours in x and y
        # import Button class from matplotlib.widgets
        from matplotlib.widgets import Button
        axcut = plt.axes([0.05, 0.05, 0.05, 0.05])

        # make button linked to axcut that prints the string 'cut!' on the console
        bcut = Button(axcut, 'shift')
        bcut.on_clicked(self.shift_contours)

        button_ax = plt.axes([0.7, 0.05, 0.2, 0.075])  # [x, y, width, height]
        button = Button(button_ax, 'Click me!')
        
        def on_button_click(event):
            print("Button clicked.")

        # Connect the button click event to the function
        button.on_clicked(on_button_click)



        #
        plt.suptitle(self.animal_id)
        plt.show()

    #
    def plot_matching_contours(self):
        
        #
        plt.figure()
        self.session_contours = []
        for k in range(len(self.sessions)):

            #
            shift = self.shifts[k]

            #
            plt.subplot(2,5,k+1)

            #
            self.session_contours.append([])

            #
            session_contour = self.sessions[k].contours

            #
            centres = self.plot_ROIs_contours()

            # loop through each session and plot contours individually
            found=False
            for p in range(len(session_contour)):
                
                #
                temp = session_contour[p]
                centre = np.median(temp,0)

                # check to see if temp centre is close to any of the centres
                dist = np.linalg.norm(centres-centre, axis=1)
                idx = np.argmin(dist)

                if dist[idx]<self.contour_ROI_max_dist:
                #print ("session: ", k, "  cell: ", p, "  closest cell: ", idx, "  dist: ", dist[idx])
                    self.session_contours[k].append([idx,p])
                    if found==False:
                        plt.plot(temp[:,0]+shift[0],
                            temp[:,1]+shift[1],
                            #c=colors[k],
                            c='black',
                            alpha=1,
                            linewidth=2,
                            label='Suite2p nearest cell'
                            )
                        found=True
                    else:
                        plt.plot(temp[:,0]+shift[0],
                            temp[:,1]+  shift[1],
                            #c=colors[k],
                            c='black',
                            alpha=1,
                            linewidth=2,
                            )
            plt.xlim(0,512)
            plt.ylim(0,512)

            if k==0:
                plt.legend()
            plt.title(self.session_ids[k])

            # 
            temp = np.vstack(self.session_contours[k])
            # sort by first column
            temp = temp[temp[:,0].argsort()]
            self.session_contours[k] = temp



        plt.suptitle(self.animal_id)
        plt.show()




def get_reward_centered_traces(session_id,
                                root_dir,
                                animal_id,
                                reward_window,
                                filter = True,
                                recompute=True):

    if session_id=='day0':
        return

    #   
    fname_out = os.path.join(root_dir,
                        animal_id,
                        session_id,
                        'reward_centered_traces_rewardwindow_'+str(reward_window)+'.npy')
    
    if os.path.exists(fname_out)==False or recompute==True:

        # load reward times            
        fname = os.path.join(root_dir,
                            animal_id,
                            session_id,
                            'results.npz')
        
        #
        data = np.load(fname, allow_pickle=True)
        rewards = data['reward_times'].T[:,::-1][:,0]
        idx = np.where(rewards>0)[0]
        reward_switches = rewards[idx]

        # load calcium traces raw from suite2p
        if False: 
            fname = os.path.join(root_dir,
                                animal_id,
                                session_id,
                                'plane0',
                                'F.npy')

            F = np.load(fname, allow_pickle=True)

            #
            # filter the trace
            if filter:
                F = scipy.signal.savgol_filter(F, 31, 3)

            ########################################################

            # get f0 values first as the median of axis 1
            #print ("F shape: ", F.shape)
            f0s = np.nanmedian(F, axis=1)

            #
            F = (F-f0s[:,np.newaxis])/f0s[:,np.newaxis]

        # use the F_upphase instead
        else:
            fname = os.path.join(root_dir,
                    animal_id,
                    session_id,
                    'plane0',
                    'binarized_traces.npz')
            d = np.load(fname, allow_pickle=True)
                        
            F = d["F_upphase"]
            F = d["F_onphase"]
            F = d["F_detrended"]
            F = d["F_filtered"]

            #


        # split c.sessions[0].F around the reward_witches time with a window of 10 seconds
        temp = []
        temp_shuffled = []
        for k in range(reward_switches.shape[0]):
            
            #
            temp1 = F[:,reward_switches[k]-reward_window:reward_switches[k]+reward_window]
            
            # make temp2 with the index shuffled
            # make a random value from 500 to -500 in F
            idx = np.random.randint(500, F.shape[1]-500)
            temp2 = F[:,reward_switches[k]-reward_window+idx:reward_switches[k]+reward_window+idx]

            # check if the window is correct 
            if temp1.shape[1]==reward_window*2:
                temp.append(temp1)

            # check if the window is correct 
            if temp2.shape[1]==reward_window*2:
                temp_shuffled.append(temp2)


        #
        reward_centered_traces = np.array(temp)
        reward_centered_traces_shuffled = np.array(temp_shuffled)

        #
        print ("# of cells, times: ", F.shape, ", output: ", reward_centered_traces.shape)

        # save the traces
        np.save(fname_out, reward_centered_traces)
        np.save(fname_out[:-4]+'_shuffled.npy', reward_centered_traces_shuffled)


def get_img_ca(cell_id, 
            F_filtered,
            window,
            tone_starts,
            shuffle,
            smooth):

    #
    img=[]
    #img_shuffled=[]
    for starts in tone_starts:
        traces = [] 
        for s in starts:
            #temp = c.sessions[session_id].F_upphase_bin[cell_id,s-window:s+window]
            if shuffle==False:
                temp = F_filtered[cell_id,s-window:s+window]
                
            # same but shuffle the tone times
            else:
                s_shuffled = np.random.randint(2*window,F_filtered.shape[1]-2*window)
                temp = F_filtered[cell_id,s_shuffled-window:s_shuffled+window]

            if temp.shape[0]!=2*window:
                continue
            #
            if smooth:
                temp = scipy.signal.savgol_filter(temp, 31, 3)

            traces.append(temp)

        # plot the average of traces
        if len(traces)==0:
            img.append(np.zeros((2*window)))
        
        else:
            img.append(np.mean(traces,axis=0))

    #
    return np.array(img)

def get_img_tone(
                cell_id,
                F_filtered,
                window,
                tones,
                shuffle,
                smooth,
                tone_type,
                remove_base_tone
                ):


    # find 1s in the [ca] vector
    idx_starts_only = np.where(F_filtered[cell_id]==1)[0]

    # find the beginning and ends of idx_starts
    from scipy.signal import chirp, find_peaks, peak_widths
    peaks, _ = find_peaks(F_filtered[cell_id])  # middle of the pluse/peak
    widths, heights, starts, ends = peak_widths(F_filtered[cell_id], peaks)
    starts = np.int32(starts)
    
    if tone_type=='starts':
        idx_starts = starts
    elif tone_type=='all':
        idx_starts = idx_starts_only

    #
    unique_tones, counts = np.unique(tones, return_counts=True)
    #print ("unique_tones: ", unique_tones)
    #
    traces = []
    for ctr, s in enumerate(idx_starts):
    
        #temp = c.sessions[session_id].F_upphase_bin[cell_id,s-window:s+window]
        if shuffle==False:
            temp = tones[s-window:s+window]
            
        # same but shuffle the tone times
        else:
            s_shuffled = np.random.randint(2*window,F_filtered.shape[1]-2*window)
            temp = tones[s_shuffled-window:s_shuffled+window]

        if temp.shape[0]!=2*window:
            continue
        #
        if smooth:
            temp = scipy.signal.savgol_filter(temp, 31, 3)

        # remove 100 vals
        temp = np.float32(temp)
        idx = np.where(temp==100)[0]
        temp[idx]=np.nan

        # same for 0 values
        idx = np.where(temp==0)[0]
        temp[idx]=np.nan

        traces.append(temp)

        #
        #if ctr%1000==0:
        #    print ("cell_id: ", cell_id, "  s: ", s, "  temp: ", temp.shape)
        #    print (temp)

    # plot the average of traces
    if len(traces)==0:
        img = np.zeros((unique_tones.shape[0], 2*window)).T
    
    else:
        #img = np.nanmean(traces,axis=0)

        # make a histogram for every time point in the window
        img=[]
        for k in range(len(traces)):
            temp = traces[k]

            if remove_base_tone:
                idx = np.where(temp==4756)[0]
                temp = np.delete(temp, idx)

            #
            y = np.histogram(temp, bins=unique_tones)

            #
            img.append(y[0])

        img = np.vstack(img).T

        
    #
    return img

  

def plot_e_state_binarization(e_state,
                            e_state_bin,
                            e_state_sign,
                            thresh):
#
    t = np.arange(0, 90000, 1)/30.

    # plot ensemble state graph
    plt.figure(figsize=(10, 10))

    plt.plot(t,e_state, label = 'ensemble state')
    plt.plot(t,e_state_bin*e_state_sign, label = 'ensemble state bin')

    # plot horizontal line from 0 to 90000 at thresh
    plt.plot(t, thresh, 'k-', lw=2, label = 'thresh')

    #
    plt.xlabel('time (s)')

    # set xlims
    plt.xlim(t[0], t[-1])

    plt.legend()
    plt.show()

    #
def resample_snippet(snippet):

    #
    time_original = np.arange(snippet.shape[0])
    data_original = snippet

    #
    f = scipy.interpolate.interp1d(time_original, data_original, kind='linear')

    # Use this function to obtain new data samples
    time_new = np.linspace(0, time_original[-1], 10000)
    #idx = np.where(time_new < time_original[-1])[0]
    #time_new = time_new[idx]

    data_new = f(time_new)
    #print (data_new.shape)
    
    # take a total of 100 samples from the snippet
    idx = np.linspace(0, data_new.shape[0]-1, 30).astype(int)
    snippet_downsampled = data_new[idx]

    return snippet_downsampled



def warp_ts1_to_ts2(ts1, ts2):
    """
    Warps ts1 to best match ts2 using DTW.
    
    Args:
    - ts1 (array-like): First time series
    - ts2 (array-like): Second time series
    
    Returns:
    - numpy.ndarray: Warped version of ts1
    """

    #
    #print (ts1)
    #print (ts2)

    #
    alignment = dtw(ts1, ts2, keep_internals=True)
    
    # Re-sample ts1 based on the warping path's indices for ts1
    #basic_warped_ts1 = ts1[alignment.index1]

    #
    return alignment.index1

def interpolate_time_warp(ts1, ts2, warp_function):

    ts1 = ts1[warp_function]
    
    # Interpolate the basic_warped_ts1 to match the length of ts2
    x_new = np.linspace(0, len(ts1) - 1, len(ts2))
    f = interp1d(np.arange(len(ts1)), ts1, kind='linear')
    resampled_warped_ts1 = f(x_new)

    return resampled_warped_ts1

#
###############################
# Define a function to handle the button click event
def x_right_shift(event, c):

    #
    c.x_shift+=1

    #
    c.alignment_logger.append(['x_shift', 1])

    #
    update_plots(c)

# Define a function to handle the button click event
def x_left_shift(event, c):

    #
    c.x_shift-=1

    #
    c.alignment_logger.append(['x_shift', -1])

    #
    update_plots(c)


# Define a function to handle the button click event
def y_up_shift(event, c):
    
    #
    c.y_shift+=1
    c.alignment_logger.append(['y_shift', 1])

    #
    update_plots(c)

# Define a function to handle the button click event
def y_down_shift(event, c):

    global calcium_object

    calcium_object = c

    print ("YDOWN")
    
    #
    calcium_object.y_shift-=1
    calcium_object.alignment_logger.append(['y_shift', -1])

    #
    update_plots(c)

# Define a function to handle the button click event
def rotate_plus(event, c):

    #
    c.theta+=0.1

    c.alignment_logger.append(['theta', +0.1])

    #
    update_plots(c)

# Define a function to handle the button click event
def rotate_minus(event, c):
    
    #
    c.theta-=0.1
    c.alignment_logger.append(['theta', -0.1])

    update_plots(c)


# Define a function to handle the button click event
def scale_y_plus(event, c):
    
    # 
    c.scale_factor_y = c.scale_factor_y*1.01
    c.alignment_logger.append(['scale_y', 1.01])

    #
    update_plots(c)

def scale_y_minus(event, c):
    
    # 
    c.scale_factor_y = c.scale_factor_y*0.99
    c.alignment_logger.append(['scale_y', 0.99])

    #
    update_plots(c)




# Define a function to handle the button click event
def scale_x_plus(event, c):
    
    #c.theta-=0.1
    c.scale_factor_x = c.scale_factor_x*1.01
    c.alignment_logger.append(['scale_x', 1.01])

    update_plots(c)

#
def scale_x_minus(event, c):
    
    #c.theta-=0.1
    c.scale_factor_x = c.scale_factor_x*0.99
    c.alignment_logger.append(['scale_x', 0.99])

    update_plots(c)


#
def update_plots(c):

    #
    c.ax.clear()

    #
    c.session_id = c.session_selected
    clr= 'blue'
    c.plot_session_contours(clr, 
                            c.x_shift, 
                            c.y_shift, 
                            c.theta,
                            c.theta_x, 
                            c.theta_y,
                            c.scale_factor_x,
                            c.scale_factor_y)

    #
    c.session_id = 0
    clr= 'red'
    c.plot_session_contours(clr)

    #
    #
    c.ax.set_xlim(0,512)
    c.ax.set_ylim(0,512)

    #c.ax.set_title("Scale factor: "+str(c.scale_factor))
   # c.ax.legend()

    plt.show()


# Define a function to handle the button click event
def exit(event, c):
    c.exit_flag = True
    
    plt.close()

# Define a function to handle the button click event
def save_data(event, c):
    

    # print animal id, session id and x_shift and y_shift
    print ("animal_id: ", c.animal_id)
    print ("session_id: ", c.session_id)
    print ("x_shift: ", c.x_shift)
    print ("y_shift: ", c.y_shift)
    print ("theta: ", c.theta)
    print ("theta_x: ", c.theta_x)
    print ("theta_y: ", c.theta_y)
    print ("scale_factor x: ", c.scale_factor_x)
    print ("scale_factor y: ", c.scale_factor_y)

    #
    print ("Logger: ", c.alignment_logger)
    
    # save all the parameters in an .npz file
    fname_out = os.path.join(c.root_dir,
                             c.animal_id,
                             str(c.session_ids[c.session_selected]),
                            'alignment_parameters.npz')
    
    #
    np.savez(fname_out,
            animal_id = c.animal_id,
            session_id = c.session_id,
            x_shift = c.x_shift,
            y_shift = c.y_shift,
            theta = c.theta,
            theta_x = c.theta_x,
            theta_y = c.theta_y,
            scale_factor_x = c.scale_factor_x,
            scale_factor_y = c.scale_factor_y,
            alignment_logger = c.alignment_logger
            )

    # 

    fname_out = os.path.join(calcium_object.root_dir,
                                calcium_object.animal_id,
                                str(calcium_object.session_ids[calcium_object.session_selected]),
                                'alignment_saved.png')
    plt.savefig(fname_out,dpi=300)
    
    plt.close()


# rotate cells             
def rotate_points(points, cx, cy, theta):
    # Convert the angle to radians
    theta = np.radians(theta)
    
    # Create a rotation matrix
    rotation_matrix = np.array([
        [np.cos(theta), -np.sin(theta)],
        [np.sin(theta), np.cos(theta)]
    ])
    
    # Create a matrix of points
    points_matrix = np.column_stack((points[:, 0] - cx, points[:, 1] - cy))
    
    # Apply the rotation matrix to the points
    rotated_points_matrix = np.dot(points_matrix, rotation_matrix.T)
    
    # Translate the points back to the original coordinate system
    rotated_points = rotated_points_matrix + np.array([cx, cy])
    
    return rotated_points


def reload_alignment(c):


    d = np.load(os.path.join(c.root_dir,
                                c.animal_id,
                                c.session_ids[c.session_selected],
                                'alignment_parameters.npz'),
                allow_pickle=True)
    
    c.x_shift = d['x_shift']
    c.y_shift = d['y_shift']
    c.theta = d['theta']
    c.theta_x = d['theta_x']
    c.theta_y = d['theta_y']
    c.scale_factor_x = d['scale_factor_x']
    c.scale_factor_y = d['scale_factor_y']

    print ("x_shift: ", c.x_shift)
    print ("y_shift: ", c.y_shift)
    print ("theta: ", c.theta)
    print ("theta_x: ", c.theta_x)
    print ("theta_y: ", c.theta_y)
    print ("scale_factor x: ", c.scale_factor_x)
    print ("scale_factor y: ", c.scale_factor_y)

    ########################################
    plt.figure(figsize=(10,10))

    #
    c.ax = plt.subplot(1,1,1)

    #
    c.session_id = 0
    clr = 'red'
    c.plot_session_contours(clr)

    #
    c.session_id = c.session_selected
    clr = 'blue'
    c.plot_session_contours(clr,
                            c.x_shift,
                            c.y_shift,
                            c.theta,
                            c.theta_x,
                            c.theta_y,
                            c.scale_factor_x,
                            c.scale_factor_y
                            )

    #
    plt.legend()

    fname_out = os.path.join(calcium_object.root_dir,
                                calcium_object.animal_id,
                                str(calcium_object.session_ids[calcium_object.session_selected]),
                                'alignment_reloaded.png')
    plt.savefig(fname_out,dpi=300)

    plt.close()



#
def on_mouse_click(event):

    global calcium_object
    
    if event.button == 1:  # Check if left mouse button (button 1) is clicked

        if event.inaxes == calcium_object.ax:    

            # figure out which quadrant you are in relative to [0,512] and [0,512] plot and x,y centre coordinates

            if event.xdata < calcium_object.theta_x and event.ydata < calcium_object.theta_y:
                print ("bottom left quadrant")
            elif event.xdata < calcium_object.theta_x and event.ydata > calcium_object.theta_y:
                print ("top left quadrant")
            elif event.xdata > calcium_object.theta_x and event.ydata < calcium_object.theta_y:   
                print ("bottom right quadrant")
            elif event.xdata > calcium_object.theta_x and event.ydata > calcium_object.theta_y:
                print ("top right quadrant")
                
            # plot dashed axies

    if event.button == 2:  # Check if middel button pressed
        print("Middle button pressed")

        if event.inaxes == calcium_object.ax:
            x, y = event.xdata, event.ydata

            #
            print("Setting centre to: ", x, y)

            #
            calcium_object.theta_x = x
            calcium_object.theta_y = y

            #
            calcium_object.scale_x = x
            calcium_object.scale_y = y

            calcium_object.alignment_logger.append(['scale', 0.999])

            calcium_object.plot_quadrants()

#
def align_gui_local(ca_object):

    global calcium_object

    calcium_object = ca_object

    #
    calcium_object.x_shift = 0 
    calcium_object.y_shift = 0
    calcium_object.theta = 0
    calcium_object.theta_x = 256
    calcium_object.theta_y = 256
    calcium_object.scale_x = 256
    calcium_object.scale_y = 256
    calcium_object.scale_factor_x = 1
    calcium_object.scale_factor_y = 1
    calcium_object.n_cells_show = 200
    calcium_object.exit_flag = False


    #
    calcium_object.cell_idxs = np.random.choice(len(calcium_object.sessions[calcium_object.session_selected].contours), 
                                size=min(calcium_object.n_cells_show, 
                                            len(calcium_object.sessions[calcium_object.session_selected].contours)), 
                                replace=False)
    calcium_object.day_cell_idx = np.random.choice(len(calcium_object.sessions[0].contours),
                                        size=min(calcium_object.n_cells_show, 
                                                len(calcium_object.sessions[0].contours)),
                                        replace=False)

    #
    calcium_object.alignment_logger = []

    ########################################
    ########################################
    ########################################
    #
    fig = plt.figure(figsize=(10,10))

    # Connect the mouse click event to the callback function
    fig.canvas.mpl_connect('button_press_event', on_mouse_click)

    #
    calcium_object.ax = plt.subplot(1,1,1)

    #
    calcium_object.session_id = 0
    clr = 'red'
    calcium_object.plot_session_contours(clr)

    #
    print ("selected session: ", calcium_object.session_ids[calcium_object.session_selected])
    calcium_object.session_id = calcium_object.session_selected
    clr = 'blue'
    calcium_object.plot_session_contours(clr)

    #
    plt.legend()

    #
    plt.xlim(0,512)
    plt.ylim(0,512)

    # Create a button widget
    button_ax = plt.axes([0.12, 0.04, 0.05, 0.03])
    button = Button(button_ax, 'x_right')
    button.on_clicked(lambda event: x_right_shift(event, calcium_object))

    #
    button_ax1 = plt.axes([0.05, 0.04, 0.05, 0.03])
    button1 = Button(button_ax1, 'x_left')
    button1.on_clicked(lambda event: x_left_shift(event, calcium_object))

    #
    button_ax2 = plt.axes([0.08, 0.07, 0.05, 0.03])
    button2 = Button(button_ax2, 'y_up')
    button2.on_clicked(lambda event: y_up_shift(event, calcium_object))

    #
    button_ax3 = plt.axes([0.08, 0.01, 0.05, 0.03])
    button3 = Button(button_ax3, 'y_down')
    button3.on_clicked(lambda event: y_down_shift(event, calcium_object))

    #
    button_ax4 = plt.axes([0.8, 0.01, 0.05, 0.03])
    button4 = Button(button_ax4, 'save')
    button4.on_clicked(lambda event: save_data(event, calcium_object))

    button_ax41 = plt.axes([0.8, 0.05, 0.05, 0.03])
    button41 = Button(button_ax41, 'exit without saving')
    button41.on_clicked(lambda event: exit(event, calcium_object))

    #
    button_ax5 = plt.axes([0.64, 0.01, 0.05, 0.03])
    button5 = Button(button_ax5, 'rotate +')
    button5.on_clicked(lambda event: rotate_plus(event, calcium_object))

    #
    button_ax6 = plt.axes([0.7, 0.01, 0.05, 0.03])
    button6 = Button(button_ax6, 'rotate -')
    button6.on_clicked(lambda event: rotate_minus(event, calcium_object))

    #######################################
    button_ax7 = plt.axes([0.4, 0.01, 0.05, 0.03])
    button7 = Button(button_ax7, 'scale y - ')
    button7.on_clicked(lambda event: scale_y_minus(event, calcium_object))

    button_ax8 = plt.axes([0.4, 0.07, 0.05, 0.03])
    button8 = Button(button_ax8, 'scale y + ')
    button8.on_clicked(lambda event: scale_y_plus(event, calcium_object))

    #
    button_ax9 = plt.axes([0.35, 0.035, 0.05, 0.03])
    button9 = Button(button_ax9, 'scale x - ')
    button9.on_clicked(lambda event: scale_x_minus(event, calcium_object))

    button_ax10 = plt.axes([0.45, 0.035, 0.05, 0.03])
    button10 = Button(button_ax10, 'scale x + ')
    button10.on_clicked(lambda event: scale_x_plus(event, calcium_object))

    plt.show(block=True)

    #
    if calcium_object.exit_flag==False:
        reload_alignment(calcium_object)